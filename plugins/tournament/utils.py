import aiohttp
import math
import numpy as np
import pandas as pd
import random

from core import report
from io import BytesIO
from matplotlib import pyplot as plt, patches
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from PIL import Image, ImageDraw, ImageFont, ImageFilter
from trueskill import Rating
from typing import Optional


def create_elimination_matches(squadrons: list[tuple[int, float]]) -> list[tuple[int, int]]:
    """
    Create tournament matches using snake pairing system.

    Args:
        squadrons: List of tuples (squadron_id, trueskill_rating)

    Returns:
        List of tuples (squadron1_id, squadron2_id) representing matches
    """
    # Sort squadrons by TrueSkill rating in descending order
    sorted_squadrons = sorted(squadrons, key=lambda x: x[1], reverse=True)

    # Check if we have an even number of squadrons
    if len(sorted_squadrons) % 2 != 0:
        raise ValueError("Need an even number of squadrons for the tournament")

    matches = []
    n = len(sorted_squadrons)
    half = n // 2

    # Create matches using snake pairing
    for i in range(half):
        squad1 = sorted_squadrons[i][0]  # Get squadron_id from tuple
        squad2 = sorted_squadrons[-(i + 1)][0]  # Get opponent from bottom, moving upwards
        matches.append((squad1, squad2))

    return matches


def create_groups(squadrons: list[tuple[int, float]], num_groups: int) -> list[list[int]]:
    """
    Create random groups for the group phase of the tournament.

    Args:
        squadrons: List of squadron IDs
        num_groups: Number of groups to create

    Returns:
        List of groups, where each group is a list of squadron IDs

    Raises:
        ValueError: If number of groups is invalid or if there aren't enough squadrons
    """
    if num_groups <= 0:
        raise ValueError("Number of groups must be positive")

    if len(squadrons) < num_groups * 2:
        raise ValueError(f"Need at least {num_groups * 2} squadrons for {num_groups} groups (minimum 2 per group)")

    # Create a copy of the squadron list to shuffle
    squadrons_copy = [x[0] for x in squadrons]
    random.shuffle(squadrons_copy)

    # Calculate minimum squadrons per group
    min_per_group = len(squadrons_copy) // num_groups
    # Calculate how many groups get an extra squadron (if uneven division)
    extras = len(squadrons_copy) % num_groups

    groups = []
    current_idx = 0

    for group_num in range(num_groups):
        # Calculate the size for this group
        group_size = min_per_group + (1 if group_num < extras else 0)
        # Create the group
        group = squadrons_copy[current_idx:current_idx + group_size]
        groups.append(group)
        current_idx += group_size

    return groups


def create_group_matches(groups: list[list[int]]) -> list[tuple[int, int]]:
    """
    Create matches for group phase where each squadron plays against all other squadrons in their group.

    Args:
        groups: List of groups, where each group is a list of squadron IDs
               (output from create_balanced_groups)

    Returns:
        List of tuples (squadron1_id, squadron2_id) representing matches

    Example:
        For a group [1, 2, 3], it creates matches [(1,2), (1,3), (2,3)]
    """
    matches = []

    # For each group
    for group in groups:
        # Create matches between each pair of squadrons in the group
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                matches.append((group[i], group[j]))

    return matches


async def download_image(image_url: str) -> bytes:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9'
    }

    timeout = aiohttp.ClientTimeout(total=30)  # 30 seconds total timeout
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(image_url, headers=headers) as response:
            if not response.headers.get('content-type', '').startswith('image/'):
                raise ValueError(f"URL does not point to an image: {image_url}")
            img_data = await response.read()
            if len(img_data) > 10 * 1024 * 1024:  # 10MB limit
                raise ValueError(f"Image too large: {image_url}")
            return img_data


async def create_versus_image(team_blue_image_url: str, team_red_image_url: str, winner: str = None) -> Optional[bytes]:
    """
    Create a versus image or winner image depending on if winner is specified.
    :param team_blue_image_url: Blue team image URL
    :param team_red_image_url: Red team image URL
    :param winner: Optional - either 'blue' or 'red' to indicate winner
    """
    img1_bio = None
    img2_bio = None
    try:
        img1_data = await download_image(team_blue_image_url)
        img2_data = await download_image(team_red_image_url)
    except ValueError:
        return None

    try:
        img1_bio = BytesIO(img1_data)
        img2_bio = BytesIO(img2_data)

        # Open images from binary data and convert to RGBA
        with Image.open(img1_bio) as img1, Image.open(img2_bio) as img2:

            img1 = img1.convert('RGBA')
            img2 = img2.convert('RGBA')

            if winner is None:
                # Original versus logic
                standard_size = (200, 200)
                img1.thumbnail(standard_size, Image.Resampling.LANCZOS)
                img2.thumbnail(standard_size, Image.Resampling.LANCZOS)
            else:
                # Winner/loser logic
                winner_size = (200, 200)
                loser_size = (150, 150)

                if winner.lower() == 'blue':
                    img1.thumbnail(winner_size, Image.Resampling.LANCZOS)
                    img2.thumbnail(loser_size, Image.Resampling.LANCZOS)
                    winner_img = img1
                    loser_img = img2
                else:  # 'red'
                    img1.thumbnail(loser_size, Image.Resampling.LANCZOS)
                    img2.thumbnail(winner_size, Image.Resampling.LANCZOS)
                    winner_img = img2
                    loser_img = img1

                # Create glow effect for winner
                glow_color = (255, 215, 0, 100)  # Golden color with alpha
                glow_size = 10
                winner_with_glow = Image.new('RGBA',
                                             (winner_img.width + 2 * glow_size,
                                              winner_img.height + 2 * glow_size),
                                             (0, 0, 0, 0))

                # Create glow effect using multiple passes
                glow = winner_img.copy()
                glow = glow.convert('RGBA')
                for i in range(glow_size):
                    glow_layer = Image.new('RGBA', winner_with_glow.size, (0, 0, 0, 0))
                    glow_layer.paste(glow, (i, i))
                    glow_layer = glow_layer.filter(ImageFilter.GaussianBlur(glow_size - i))
                    for x in range(glow_layer.width):
                        for y in range(glow_layer.height):
                            r, g, b, a = glow_layer.getpixel((x, y))
                            if a > 0:
                                glow_layer.putpixel((x, y),
                                                    (glow_color[0], glow_color[1],
                                                     glow_color[2], min(a, glow_color[3])))
                    winner_with_glow = Image.alpha_composite(winner_with_glow, glow_layer)

                # Paste the original winner image in the center of the glow
                winner_with_glow.paste(winner_img, (glow_size, glow_size), winner_img)

                # Replace original images with processed ones
                if winner.lower() == 'blue':
                    img1 = winner_with_glow
                    img2 = loser_img
                else:
                    img1 = loser_img
                    img2 = winner_with_glow

            # Add spacing for VS text or spacing between winner/loser
            spacing = 100
            total_width = img1.width + img2.width + spacing
            max_height = max(img1.height, img2.height)

            # Create new image with transparent background
            combined_image = Image.new('RGBA', (total_width, max_height), (255, 255, 255, 0))

            # Calculate vertical positions
            y1 = (max_height - img1.height) // 2
            y2 = (max_height - img2.height) // 2

            # Paste images
            combined_image.paste(img1, (0, y1), img1)
            combined_image.paste(img2, (img1.width + spacing, y2), img2)

            # Add text
            draw = ImageDraw.Draw(combined_image)
            try:
                font = ImageFont.truetype("arial.ttf", 32 if winner is None else 24)
            except:
                font = ImageFont.load_default()

            if winner is None:
                # Original VS text logic
                vs_text = " vs "
                text_bbox = draw.textbbox((0, 0), vs_text, font=font)
                text_width = text_bbox[2] - text_bbox[0]
                text_height = text_bbox[3] - text_bbox[1]
                x = (total_width - text_width) // 2
                y = (max_height - text_height) // 2

                # Draw VS text with outline
                draw.text((x - 1, y - 1), vs_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x + 1, y - 1), vs_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x - 1, y + 1), vs_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x + 1, y + 1), vs_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x, y), vs_text, font=font, fill=(255, 255, 255, 255))
            else:
                # Add WINNER text under winning image
                winner_text = "WINNER"
                text_bbox = draw.textbbox((0, 0), winner_text, font=font)
                text_width = text_bbox[2] - text_bbox[0]

                if winner.lower() == 'blue':
                    x = (img1.width - text_width) // 2
                    y = y1 + img1.height + 5
                else:
                    x = img1.width + spacing + (img2.width - text_width) // 2
                    y = y2 + img2.height + 5

                # Draw WINNER text with golden color and outline
                draw.text((x - 1, y - 1), winner_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x + 1, y - 1), winner_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x - 1, y + 1), winner_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x + 1, y + 1), winner_text, font=font, fill=(0, 0, 0, 255))
                draw.text((x, y), winner_text, font=font, fill=(255, 215, 0, 255))

            # Save to binary buffer
            buffer = BytesIO()
            combined_image.save(buffer, format='PNG')
            buffer.seek(0)
            image_bytes = buffer.getvalue()
            buffer.close()

            return image_bytes

    finally:
        # Clean up BytesIO objects
        if img1_bio:
            img1_bio.close()
        if img2_bio:
            img2_bio.close()


async def create_winner_image(winner_image_url: str) -> Optional[bytes]:
    """
    Create a special victory image for the tournament winner with enhanced visual effects.
    :param winner_image_url: URL of the winning squadron's image
    """
    winner_bio = None
    try:
        winner_data = await download_image(winner_image_url)
    except ValueError:
        return None

    try:
        winner_bio = BytesIO(winner_data)

        with Image.open(winner_bio) as winner_img:

            winner_img = winner_img.convert('RGBA')

            # Make the winner image larger for tournament victory
            winner_size = (300, 300)  # Bigger size for the tournament winner
            winner_img.thumbnail(winner_size, Image.Resampling.LANCZOS)

            # Create a larger canvas for effects
            padding = 100  # Extra space for effects and text
            extra_bottom_space = 150
            canvas_size = (winner_img.width + padding * 2,
                           winner_img.height + padding * 2 + extra_bottom_space)
            final_image = Image.new('RGBA', canvas_size, (0, 0, 0, 0))

            # Create multiple layers of the golden glow with different intensities
            glow_colors = [
                (255, 215, 0, 100),  # Golden
                (255, 223, 0, 80),  # Lighter golden
                (255, 200, 0, 60),  # Darker golden
            ]

            for i, glow_color in enumerate(glow_colors):
                glow_size = 20 - i * 5  # Decreasing glow size for each layer
                glow_layer = Image.new('RGBA', canvas_size, (0, 0, 0, 0))

                # Create star-like rays
                draw = ImageDraw.Draw(glow_layer)
                center = (canvas_size[0] // 2, canvas_size[1] // 2)
                for angle in range(0, 360, 45):  # 8 rays
                    end_x = center[0] + int(math.cos(math.radians(angle)) * (winner_size[0] // 2 + 50))
                    end_y = center[1] + int(math.sin(math.radians(angle)) * (winner_size[1] // 2 + 50))
                    draw.line([center, (end_x, end_y)], fill=glow_color, width=10)

                # Add circular glow
                glow = winner_img.copy()
                glow = glow.filter(ImageFilter.GaussianBlur(glow_size))
                mask = Image.new('L', glow.size, 0)
                draw_mask = ImageDraw.Draw(mask)
                draw_mask.ellipse([(0, 0), glow.size], fill=255)
                glow.putalpha(mask)

                # Position the glow in the center
                glow_pos = ((canvas_size[0] - glow.width) // 2,
                            (canvas_size[1] - glow.height) // 2)
                final_image.paste(glow, glow_pos, glow)

            # Add the main image in the center
            winner_pos = ((canvas_size[0] - winner_img.width) // 2,
                          (canvas_size[1] - winner_img.height) // 2)
            final_image.paste(winner_img, winner_pos, winner_img)

            # Add text
            draw = ImageDraw.Draw(final_image)
            try:
                title_font = ImageFont.truetype("arial.ttf", 48)
                subtitle_font = ImageFont.truetype("arial.ttf", 36)
            except:
                title_font = ImageFont.load_default()
                subtitle_font = ImageFont.load_default()

            # Add "TOURNAMENT CHAMPION" text
            title_text = "TOURNAMENT"
            subtitle_text = "CHAMPION"

            # Calculate text positions - moved lower with extra spacing
            base_text_y = canvas_size[1] - extra_bottom_space + 20  # Start text higher up from bottom

            # Calculate text positions
            for i, (text, font) in enumerate([(title_text, title_font), (subtitle_text, subtitle_font)]):
                bbox = draw.textbbox((0, 0), text, font=font)
                text_width = bbox[2] - bbox[0]
                x = (canvas_size[0] - text_width) // 2
                y = base_text_y + (i * 50)  # Stack the text lines

                # Draw text with golden gradient effect
                for offset in range(3):  # Create 3D effect
                    draw.text((x - offset, y - offset), text, font=font,
                              fill=(255 - offset * 20, 215 - offset * 20, offset * 20, 255))

            # Save to binary buffer
            buffer = BytesIO()
            final_image.save(buffer, format='PNG')
            buffer.seek(0)
            image_bytes = buffer.getvalue()
            buffer.close()

            return image_bytes

    finally:
        if winner_bio:
            winner_bio.close()


def calculate_point_multipliers(killer_rating: Rating, victim_rating: Rating) -> tuple[float, float]:
    """
    Calculate point multipliers for both killer and victim based on pre-match TrueSkill ratings.

    Args:
        killer_rating (Rating): TrueSkill rating of the killer before the match
        victim_rating (Rating): TrueSkill rating of the victim before the match

    Returns:
        tuple[float, float]: (killer_multiplier, victim_multiplier)
    """
    killer_conservative_skill = killer_rating.mu - 3 * killer_rating.sigma
    victim_conservative_skill = victim_rating.mu - 3 * victim_rating.sigma

    skill_difference = victim_conservative_skill - killer_conservative_skill

    total_uncertainty = killer_rating.sigma + victim_rating.sigma
    uncertainty_factor = 1.0 / (1.0 + total_uncertainty / 8.33)

    # Add a small threshold for considering skills equal
    if abs(skill_difference) < 0.0001:
        # Skills are effectively equal - use base multiplier
        return 1.0, 1.0

    # Killer multiplier
    if skill_difference > 0:
        # Killer has lower skill than victim - they get bonus points
        killer_raw_multiplier = 1.0 + math.log(1 + skill_difference / 10) * 0.5
        killer_multiplier = min(killer_raw_multiplier * uncertainty_factor, 2.5)
    else:
        # Killer has higher skill - they get reduced points
        killer_raw_multiplier = 1.0 / (1 + abs(skill_difference / 20))
        killer_multiplier = max(killer_raw_multiplier * uncertainty_factor, 0.5)

    # Victim multiplier
    if skill_difference > 0:
        # Victim has higher skill - they lose fewer points
        victim_raw_multiplier = 1.0 / (1 + skill_difference / 20)
        victim_multiplier = max(victim_raw_multiplier * uncertainty_factor, 0.5)
    else:
        # Victim has lower skill - they lose more points
        victim_raw_multiplier = 1.0 + math.log(1 + abs(skill_difference) / 10) * 0.5
        victim_multiplier = min(victim_raw_multiplier * uncertainty_factor, 2.5)

    return killer_multiplier, victim_multiplier


def create_tournament_sheet(squadrons_df: pd.DataFrame, matches_df: pd.DataFrame, tournament_id: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tournament {tournament_id}"

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Write group stages
    groups = squadrons_df.sort_values(['group_number', 'squadron_id'])

    # Start writing from row 2 to leave space for headers
    row = 2

    # Write group stage header
    ws.cell(row=1, column=1, value="Groups").font = Font(bold=True)

    # Write groups and squadrons
    for group_num, group_data in groups.groupby('group_number'):
        ws.cell(row=row, column=1, value=f'Group {group_num}').font = Font(bold=True)
        row += 1

        for _, squadron in group_data.iterrows():
            cell = ws.cell(row=row, column=1, value=squadron['name'])
            cell.border = border
            row += 1

        row += 1  # Space between groups

    # Create squadron_id to name mapping for easier lookup
    squadron_names = dict(zip(squadrons_df['squadron_id'], squadrons_df['name']))

    # Write matches by stage
    matches_sorted = matches_df.sort_values(['stage', 'match_id'])
    max_row = row  # Remember the maximum row used by groups

    # Group matches by stage
    for stage, stage_matches in matches_sorted.groupby('stage'):
        if stage == 1:
            # Group stage matches - place them next to the groups
            column = 3  # Start group matches from column C
            ws.cell(row=1, column=column, value="Group Stage Matches").font = Font(bold=True)

            row = 2
            for _, match in stage_matches.iterrows():
                red_name = squadron_names[match['squadron_red']]
                blue_name = squadron_names[match['squadron_blue']]

                ws.cell(row=row, column=column,
                        value=f'{red_name} vs {blue_name}').border = border
                ws.cell(row=row + 1, column=column,
                        value=f'Result: {match["squadron_red_rounds_won"]} - {match["squadron_blue_rounds_won"]}').border = border

                if match['winner_squadron_id']:
                    winner_name = squadron_names[match['winner_squadron_id']]
                    ws.cell(row=row + 2, column=column,
                            value=f'Winner: {winner_name}').font = Font(italic=True)

                row += 4  # Space for next match
        else:
            # Elimination stages - place them progressively to the right
            column = 3 + (stage - 1) * 3  # Each stage moves 3 columns to the right
            ws.cell(row=1, column=column, value=f"Stage {stage}").font = Font(bold=True)

            row_spacing = 4 * (2 ** (stage - 2))  # Increase spacing between matches for each stage
            base_row = 2

            for idx, match in enumerate(stage_matches.itertuples()):
                current_row = base_row + (idx * row_spacing)

                red_name = squadron_names[match.squadron_red]
                blue_name = squadron_names[match.squadron_blue]

                ws.cell(row=current_row, column=column,
                        value=f'{red_name} vs {blue_name}').border = border
                ws.cell(row=current_row + 1, column=column,
                        value=f'Result: {match.squadron_red_rounds_won} - {match.squadron_blue_rounds_won}').border = border

                if match.winner_squadron_id:
                    winner_name = squadron_names[match.winner_squadron_id]
                    ws.cell(row=current_row + 2, column=column,
                            value=f'Winner: {winner_name}').font = Font(italic=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    xls_bytes = buf.getvalue()
    buf.close()

    return xls_bytes


def create_placeholder_icon(size=(32, 32)) -> Image:
    """Create a simple placeholder icon with a question mark"""
    img = np.ones((size[0], size[1], 4), dtype=np.uint8) * 255  # White background with alpha
    img[:, :, 3] = 255  # Full opacity
    # Convert to PIL Image
    pil_img = Image.fromarray(img)
    return pil_img


async def render_groups(groups: list[list[tuple[str, str]]]) -> bytes:
    n_groups = len(groups)
    max_members = max(len(group) for group in groups)

    width_per_group = 3
    height_per_member = 0.8
    icon_size = 32

    fig_width = max(8, n_groups * width_per_group)
    fig_height = max(4, (max_members + 1) ** height_per_member)

    dpi = 150
    plt.switch_backend('agg')
    fig = plt.figure(figsize=(fig_width, fig_height), dpi=dpi)
    ax = fig.add_subplot(111)

    ax.set_facecolor('#424242')
    fig.patch.set_facecolor('#424242')

    ax.set_xlim(0, n_groups)
    ax.set_ylim(-(max_members + 1), 0)
    ax.set_xticks([])
    ax.set_yticks([])

    for spine in ax.spines.values():
        spine.set_visible(False)

    header = patches.Rectangle((0, -1), n_groups, 1,
                               facecolor='#4472C4',
                               edgecolor='none')
    ax.add_patch(header)

    for i in range(n_groups):
        bg = patches.Rectangle((i, -(max_members + 1)), 1, max_members,
                               facecolor='#424242',
                               edgecolor='none')
        ax.add_patch(bg)

    for i in range(n_groups + 1):
        ax.plot([i, i], [-(max_members + 1), 0], 'gray', linewidth=0.5)
    for i in range(max_members + 2):
        ax.plot([0, n_groups], [-i, -i], 'gray', linewidth=0.5)

    placeholder_icon = create_placeholder_icon()

    for i in range(n_groups):
        group_letter = chr(65 + i)

        ax.text(i + 0.5, -0.5, f"Group {group_letter}",
                horizontalalignment='center',
                verticalalignment='center',
                color='white',
                fontweight='bold')

        for j, (name, image_url) in enumerate(groups[i]):
            y_pos = -(j + 1.5)

            # Moved text closer to images
            ax.text(i + 0.3, y_pos, name,
                    horizontalalignment='left',
                    verticalalignment='center',
                    color='white')

            try:
                if image_url:
                    try:
                        img_data = await download_image(image_url)
                    except ValueError:
                        img_data = None
                    if img_data:
                        figure_coords = ax.get_figure().transFigure.inverted()
                        data_coords = ax.transData

                        icon_left, icon_bottom = figure_coords.transform(
                            data_coords.transform((i + 0.05, y_pos - 0.15))
                        )
                        icon_right, icon_top = figure_coords.transform(
                            data_coords.transform((i + 0.25, y_pos + 0.15))
                        )

                        icon_ax = fig.add_axes((
                            icon_left,
                            icon_bottom,
                            icon_right - icon_left,
                            icon_top - icon_bottom
                        ))

                        img = Image.open(BytesIO(img_data))
                        if img.mode != 'RGBA':
                            img = img.convert('RGBA')

                        img = img.resize((icon_size, icon_size))
                        icon_ax.imshow(img)
                        icon_ax.axis('off')
                    else:
                        ax.imshow(placeholder_icon,
                                  extent=(i + 0.05, i + 0.25, y_pos - 0.15, y_pos + 0.15),
                                  aspect='auto')
                else:
                    ax.imshow(placeholder_icon,
                              extent=(i + 0.05, i + 0.25, y_pos - 0.15, y_pos + 0.15),
                              aspect='auto')

            except Exception as e:
                print(f"Error loading icon: {e}")
                ax.imshow(placeholder_icon,
                          extent=(i + 0.05, i + 0.25, y_pos - 0.15, y_pos + 0.15),
                          aspect='auto')

    buf = BytesIO()
    fig.savefig(buf,
                format='png',
                bbox_inches='tight',
                dpi=300,
                facecolor='#424242',
                edgecolor='none')
    plt.close(fig)

    buf.seek(0)
    image_bytes = buf.getvalue()
    buf.close()

    return image_bytes


class TimePreferences(report.GraphElement):

    async def render(self, tournament_id: Optional[int] = None):
        labels = []
        values = []
        inner_sql = "WHERE p.tournament_id = %(tournament_id)s" if tournament_id else ""
        async with self.apool.connection() as conn:
            cursor = await conn.execute(f"""
                SELECT t.start_time, count(p.squadron_id) AS num 
                FROM tm_squadron_time_preferences p 
                JOIN tm_available_times t on p.available_time_id = t.time_id
                {inner_sql}
                GROUP BY 1
                ORDER BY 2 DESC
            """, {"tournament_id": tournament_id})
            async for row in cursor:
                labels.insert(0, row[0].strftime('%H:%M'))
                values.insert(0, row[1])

        if values:
            def func(pct, allvals):
                absolute = int(round(pct / 100. * np.sum(allvals)))
                return f'{absolute}'

            patches, texts, pcts = self.axes.pie(values, labels=labels, autopct=lambda pct: func(pct, values),
                                                 wedgeprops={'linewidth': 3.0, 'edgecolor': 'black'}, normalize=True)
            plt.setp(pcts, color='black', fontweight='bold', fontsize=15)
            plt.setp(texts, color='white', fontsize=15)
            self.axes.set_title('Preferred Times', color='white', fontsize=25)
            self.axes.axis('equal')
        else:
            self.axes.set_visible(False)


class TerrainPreferences(report.GraphElement):

    async def render(self, tournament_id: Optional[int] = None):
        labels = []
        values = []
        inner_sql = "WHERE tournament_id = %(tournament_id)s" if tournament_id else ""
        async with self.apool.connection() as conn:
            cursor = await conn.execute(f"""
                SELECT terrain, count(*) AS num 
                FROM tm_squadron_terrain_preferences
                {inner_sql}
                GROUP BY 1
                ORDER BY 2 DESC
            """, {"tournament_id": tournament_id})
            async for row in cursor:
                labels.insert(0, row[0].replace('_terrain', ''))
                values.insert(0, row[1])

        if values:
            def func(pct, allvals):
                absolute = int(round(pct / 100. * np.sum(allvals)))
                return f'{absolute}'

            patches, texts, pcts = self.axes.pie(values, labels=labels, autopct=lambda pct: func(pct, values),
                                                 wedgeprops={'linewidth': 3.0, 'edgecolor': 'black'}, normalize=True)
            plt.setp(pcts, color='black', fontweight='bold', fontsize=15)
            plt.setp(texts, color='white', fontsize=15)
            self.axes.set_title('Preferred Terrains', color='white', fontsize=25)
            self.axes.axis('equal')
        else:
            self.axes.set_visible(False)
