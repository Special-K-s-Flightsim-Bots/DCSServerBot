import asyncio
import aiohttp
import discord
import os
import pandas as pd
import psycopg
import random
import re
import shutil
import warnings

from core import Plugin, Group, utils, get_translation, PluginRequiredError, Status, Coalition, yn_question, Server, \
    MizFile, DataObjectFactory, async_cache, Report, TRAFFIC_LIGHTS, THEATRES
from datetime import datetime, timezone, timedelta
from discord import app_commands, TextChannel, CategoryChannel, NotFound
from discord.ext import tasks, commands
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from psycopg.errors import UniqueViolation
from psycopg.rows import dict_row
from services.bot import DCSServerBot
from time import time
from typing import Optional

from .const import TOURNAMENT_PHASE
from .listener import TournamentEventListener
from .utils import create_versus_image, create_elimination_matches, create_group_matches, create_groups, \
    create_tournament_sheet, render_groups, create_winner_image, squadrons_to_groups
from .view import ChoicesView, ApplicationModal, ApplicationView, TournamentModal, SignupView
from ..competitive.commands import Competitive
from ..creditsystem.squadron import Squadron

# ruamel YAML support
from ruamel.yaml import YAML
yaml = YAML()

_ = get_translation(__name__.split('.')[1])


async def tournament_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    async with interaction.client.apool.connection() as conn:
        cursor = await conn.execute("""
            SELECT tournament_id, campaign 
            FROM tm_tournaments 
            WHERE campaign ILIKE %s ORDER BY campaign
        """, ('%' + current + '%', ))
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[1], value=row[0])
            async for row in cursor
        ]
        return choices[:25]


async def active_tournament_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    async with interaction.client.apool.connection() as conn:
        cursor = await conn.execute("""
            SELECT t.tournament_id, t.campaign 
            FROM tm_tournaments t JOIN campaigns c ON t.campaign = c.name 
            WHERE campaign ILIKE %s
            AND c.start <= NOW() AT TIME ZONE 'UTC'
            AND COALESCE(c.stop, NOW() AT TIME ZONE 'UTC') >= NOW() AT TIME ZONE 'UTC'
            ORDER BY campaign
        """, ('%' + current + '%', ))
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[1], value=row[0])
            async for row in cursor
        ]
        return choices[:25]


async def squadron_autocomplete(interaction: discord.Interaction, current: str,
                                config: str) -> list[app_commands.Choice[int]]:
    if not utils.check_roles(interaction.client.roles["GameMaster"], interaction.user):
        ucid = await interaction.client.get_ucid_by_member(interaction.user)
        squadron_admin_sql = """
        JOIN squadron_members sm ON sm.squadron_id = sq.id AND sm.player_ucid = %(ucid)s AND sm.admin IS TRUE
        """
    else:
        ucid = None
        squadron_admin_sql = ""

    tournament_id = utils.get_interaction_param(interaction, "tournament")
    if config == 'all':
        sub_query = ""
    elif config == 'reject':
        sub_query = "AND ts.status IN ('PENDING', 'ACCEPTED')"
    elif config == 'accepted':
        sub_query = "AND ts.status = 'ACCEPTED'"
    else:
        return []
    async with interaction.client.apool.connection() as conn:
        cursor = await conn.execute(f"""
            SELECT ts.squadron_id, sq.name, ts.status  
            FROM tm_squadrons ts JOIN squadrons sq ON ts.squadron_id = sq.id  
            {squadron_admin_sql}
            WHERE ts.tournament_id = %(tournament_id)s {sub_query} AND sq.name ILIKE %(name)s 
            ORDER BY CASE status
                WHEN 'PENDING' THEN 1
                WHEN 'REJECTED' THEN 2
                WHEN 'ACCEPTED' THEN 3
                ELSE 4
            END, sq.name
        """, {"tournament_id": tournament_id, "name": '%' + current + '%', "ucid": ucid})
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=f"{row[1]} ({row[2]})" if config == 'all' else row[1], value=row[0])
            async for row in cursor
        ]
        return choices[:25]

async def all_squadron_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    return await squadron_autocomplete(interaction, current, 'all')


async def reject_squadron_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    return await squadron_autocomplete(interaction, current, 'reject')


async def valid_squadron_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    return await squadron_autocomplete(interaction, current, 'accepted')


async def stage_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    tournament_id = utils.get_interaction_param(interaction, "tournament")
    async with interaction.client.apool.connection() as conn:
        cursor = await conn.execute("""
            SELECT COALESCE(MAX(stage), 1) 
            FROM tm_matches 
            WHERE tournament_id = %s
        """, (tournament_id,))
        stage = (await cursor.fetchone())[0]
    return [
        app_commands.Choice(name=stage, value=stage),
        app_commands.Choice(name=stage+1, value=stage+1),
    ]


async def server_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    tournament_id = utils.get_interaction_param(interaction, "tournament")
    async with interaction.client.apool.connection() as conn:
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[0], value=row[0])
            async for row in await conn.execute("""
                SELECT s.server_name 
                FROM tm_tournaments t 
                     JOIN campaigns c ON t.campaign = c.name
                     JOIN campaigns_servers s ON c.id = s.campaign_id 
                WHERE t.tournament_id = %s
                AND c.start <= NOW() AT TIME ZONE 'UTC'
                AND COALESCE(c.stop, NOW() AT TIME ZONE 'UTC') >= NOW() AT TIME ZONE 'UTC'
                AND s.server_name ILIKE %s
                ORDER BY s.server_name
            """, (tournament_id, '%' + current + '%', ))
        ]
        return choices[:25]


async def all_matches_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    tournament_id = utils.get_interaction_param(interaction, "tournament")
    async with interaction.client.apool.connection() as conn:
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[1] + ' vs ' + row[2], value=row[0])
            async for row in await conn.execute("""
                SELECT m.match_id, s1.name, s2.name 
                FROM tm_tournaments t
                     JOIN tm_matches m ON t.tournament_id = m.tournament_id 
                     JOIN squadrons s1 ON s1.id = m.squadron_blue
                     JOIN squadrons s2 ON s2.id = m.squadron_red
                WHERE t.tournament_id = %s
                ORDER BY 1
            """, (tournament_id, ))
            if not current or current.casefold() in row[1].casefold() or current.casefold() in row[2].casefold()
        ]
        return choices[:25]


async def active_matches_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    tournament_id = utils.get_interaction_param(interaction, "tournament")
    async with interaction.client.apool.connection() as conn:
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[1] + ' vs ' + row[2], value=row[0])
            async for row in await conn.execute("""
                SELECT m.match_id, s1.name, s2.name 
                FROM tm_tournaments t
                     JOIN tm_matches m ON t.tournament_id = m.tournament_id 
                     JOIN squadrons s1 ON s1.id = m.squadron_blue
                     JOIN squadrons s2 ON s2.id = m.squadron_red
                WHERE t.tournament_id = %s AND m.winner_squadron_id IS NULL
                AND (
                    m.round_number > 0 OR m.server_name NOT IN (
                        SELECT server_name FROM tm_matches
                        WHERE tournament_id = t.tournament_id
                        AND round_number > 0 
                        AND winner_squadron_id IS NULL
                    )
                )
                ORDER BY 1
            """, (tournament_id,))
            if not current or current.casefold() in row[1].casefold() or current.casefold() in row[2].casefold()
        ]
        return choices[:25]


async def match_squadrons_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    match_id = utils.get_interaction_param(interaction, "match")
    async with interaction.client.apool.connection() as conn:
        cursor = await conn.execute("SELECT squadron_blue, squadron_red FROM tm_matches WHERE match_id = %s",
                                    (match_id, ))
        squadron_blue_id, squadron_red_id = await cursor.fetchone()
        squadron_blue = utils.get_squadron(interaction.client.node, squadron_id=squadron_blue_id)
        squadron_red = utils.get_squadron(interaction.client.node, squadron_id=squadron_red_id)
        return [
            app_commands.Choice(name=squadron_blue['name'], value=squadron_blue_id),
            app_commands.Choice(name=squadron_red['name'], value=squadron_red_id)
        ]


async def mission_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    match_id = utils.get_interaction_param(interaction, "match")
    async with interaction.client.apool.connection() as conn:
        cursor = await conn.execute(f"""
            SELECT server_name FROM tm_matches WHERE match_id = %s
        """, (match_id,))
        server_name = (await cursor.fetchone())[0]
    if server_name:
        interaction.data['options'][0]['options'].append({"name": "server", "type": 4, "value": server_name})
        return await utils.mission_autocomplete(interaction, current)
    return []


async def date_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
    now = int(time())
    day_start = now - (now % 86400)
    return [
        app_commands.Choice(
            name=(datetime.fromtimestamp(day_start + (i * 86400), tz=timezone.utc)).strftime("%Y-%m-%d"),
            value=day_start + (i * 86400)
        )
        for i in range(0, 25)
    ]


async def time_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    tournament_id = utils.get_interaction_param(interaction, "tournament")
    async with interaction.client.apool.connection() as conn:
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[0].strftime("%H:%M"), value=row[0].hour * 3600 + row[0].minute * 60)
            async for row in await conn.execute("""
                SELECT start_time FROM tm_available_times WHERE tournament_id = %s
            """, (tournament_id, ))
        ]
        return choices[:25]


async def tickets_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    tournament_id = utils.get_interaction_param(interaction, "tournament")
    squadron_id = utils.get_interaction_param(interaction, "squadron")
    async with interaction.client.apool.connection() as conn:
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=row[0], value=row[0])
            async for row in await conn.execute("""
                SELECT ticket_name 
                FROM tm_tickets 
                WHERE tournament_id = %s AND squadron_id = %s AND ticket_count > 0 AND ticket_name ILIKE %s
            """, (tournament_id, squadron_id, '%' + current + '%'))
        ]
        return choices[:25]


class Tournament(Plugin[TournamentEventListener]):

    async def cog_load(self) -> None:
        await super().cog_load()
        if self.get_config().get('autostart_matches', False):
            self.match_scheduler.add_exception_type(psycopg.OperationalError)
            self.match_scheduler.add_exception_type(ValueError)
            self.match_scheduler.start()

    async def cog_unload(self) -> None:
        if self.get_config().get('autostart_matches', False):
            self.match_scheduler.cancel()
        await super().cog_unload()

    async def rename(self, conn: psycopg.AsyncConnection, old_name: str, new_name: str):
        await conn.execute('UPDATE tm_matches SET server_name = %s WHERE server_name = %s', (new_name, old_name))

    def get_admin_channel(self):
        config = self.get_config()
        channel_id = config.get('channels', {}).get('admin')
        if channel_id:
            channel = self.bot.get_channel(channel_id)
        else:
            channel = self.bot.get_admin_channel()
        return channel

    async def get_tournament(self, tournament_id: int) -> Optional[dict]:
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                await cursor.execute("""
                    SELECT t.tournament_id, c.name, c.description,
                           c.start AT TIME ZONE 'UTC' AS start, c.stop AT TIME ZONE 'UTC' AS stop, 
                           t.rounds, t.num_players
                    FROM campaigns c JOIN tm_tournaments t
                    ON c.name = t.campaign
                    WHERE t.tournament_id = %s
                """, (tournament_id, ))
                return await cursor.fetchone()

    async def get_match(self, match_id: int) -> Optional[dict]:
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                await cursor.execute("""
                    SELECT * FROM tm_matches WHERE match_id = %s
                """, (match_id, ))
                return await cursor.fetchone()

    def get_info_channel(self) -> Optional[discord.TextChannel]:
        config = self.get_config()
        channel_id = config.get('channels', {}).get('info')
        if channel_id and self.bot.check_channel(channel_id):
            return self.bot.get_channel(channel_id)
        return None

    async def get_squadron_channel(self, match_id: int, side: str) -> Optional[TextChannel]:
        async with self.apool.connection() as conn:
            cursor = await conn.execute(f"""
                SELECT squadron_{side}_channel FROM tm_matches WHERE match_id = %s
            """, (match_id, ))
            row = await cursor.fetchone()
            if row:
                channel_id = row[0]
                return self.bot.get_channel(channel_id)
        return None

    @async_cache
    async def get_squadron(self, tournament_id: int, squadron_id: int) -> Squadron:
        squadron = utils.get_squadron(node=self.node, squadron_id=squadron_id)
        async with self.node.apool.connection() as conn:
            cursor = await conn.execute("""
                SELECT c.id FROM campaigns c 
                JOIN tm_tournaments t ON t.campaign = c.name
                WHERE t.tournament_id = %s
            """, (tournament_id, ))
            campaign_id = (await cursor.fetchone())[0]
        return DataObjectFactory().new(Squadron, node=self.node, name=squadron['name'], campaign_id=campaign_id)

    async def inform_squadron(self, *, tournament_id: int, squadron_id: int, message: Optional[str] = None,
                              embed: Optional[discord.Embed] = None):
        async with self.apool.connection() as conn:
            async for row in await conn.execute("""
                SELECT p.discord_id
                FROM players p JOIN squadron_members m ON p.ucid = m.player_ucid
                WHERE m.squadron_id = %s AND m.admin IS TRUE
            """, (squadron_id,)):
                user = self.bot.get_user(row[0])
                if user:
                    tournament = await self.get_tournament(tournament_id)
                    squadron = utils.get_squadron(node=self.node, squadron_id=squadron_id)
                    dm_channel = await user.create_dm()
                    if message:
                        message = message.format(squadron=squadron['name'], tournament=tournament['name'])
                    await dm_channel.send(content=message, embed=embed)

    async def get_terrain_preferences(self, tournament_id: int, squadron_id: int) -> list[str]:
        terrains = []
        async with self.apool.connection() as conn:
            async for row in await conn.execute("""
                SELECT terrain
                FROM tm_squadron_terrain_preferences
                WHERE tournament_id = %s
                AND squadron_id = %s
            """, (tournament_id, squadron_id)):
                terrains.append(row[0])
        return terrains

    async def get_time_preferences(self, tournament_id: int, squadron_id: int) -> list[str]:
        times = []
        async with self.apool.connection() as conn:
            async for row in await conn.execute("""
                SELECT tp.available_time_id, tt.start_time
                FROM tm_squadron_time_preferences tp
                JOIN tm_available_times tt ON tt.time_id = tp.available_time_id
                WHERE tp.tournament_id = %s
                AND tp.squadron_id = %s
            """, (tournament_id, squadron_id)):
                times.append(row[1].replace(tzinfo=timezone.utc).strftime('%H:%M'))
        return times

    # New command group "/tournament"
    tournament = Group(name="tournament", description="Commands to manage tournaments")

    async def render_groups_image(self, tournament_id: int) -> Optional[bytes]:
        groups: list[list[int]] = []
        async with self.apool.connection() as conn:
            async for row in await conn.execute("""
                SELECT group_number, squadron_id 
                FROM tm_squadrons 
                WHERE tournament_id = %s
                ORDER BY group_number
            """, (tournament_id,)):
                if not row[0]:
                    return None
                if row[0] > len(groups):
                    groups.append([])
                groups[row[0] - 1].append(row[1])

        all_squadron_ids = [squad_id for group in groups for squad_id in group]
        squadron_data = {
            squad_id: utils.get_squadron(self.node, squadron_id=squad_id)
            for squad_id in all_squadron_ids
        }

        buffer = await render_groups([
            [
                (squadron_data[squad_id]['name'], squadron_data[squad_id]['image_url'])
                for squad_id in group
            ]
            for group in groups
        ])
        return buffer

    async def render_status_embed(self, tournament_id: int, *,
                                  phase: Optional[TOURNAMENT_PHASE] = None) -> None:
        tournament = await self.get_tournament(tournament_id)
        async with self.apool.connection() as conn:
            # read number of squadrons
            cursor = await conn.execute("""
                SELECT COUNT(*) FROM tm_squadrons WHERE tournament_id = %s AND status = 'ACCEPTED'
            """, (tournament_id,))
            num_squadrons = (await cursor.fetchone())[0]

            # check if we have groups defined
            cursor = await conn.execute("""
                SELECT COUNT(*)
                FROM tm_squadrons
                WHERE tournament_id = %s
                  AND group_number IS NOT NULL
            """, (tournament_id,))
            groups = (await cursor.fetchone())[0] > 0

            # check which stage/level we are in
            cursor = await conn.execute("""
                SELECT MAX(stage)
                FROM tm_matches
                WHERE tournament_id = %s
            """, (tournament_id,))
            level = (await cursor.fetchone())[0]

        embed = discord.Embed(color=discord.Color.blue(), title=f"Tournament {tournament['name']} Overview")
        buffer = None

        if not phase:
            if groups and level == 1:
                phase = TOURNAMENT_PHASE.START_GROUP_PHASE
            else:
                phase = TOURNAMENT_PHASE.START_ELIMINATION_PHASE

        if phase == TOURNAMENT_PHASE.SIGNUP:
            message = _("## :warning: Attention all Squadron Leaders! :warning:\n"
                        "A new tournament has been created:\n"
                        "\n"
                        "```{}```").format(tournament['description'])
            message += _("\nYou can use {} to sign up.").format(
                (await utils.get_command(self.bot, group=self.tournament.name, name=self.signup.name)).mention)
            embed.add_field(name=utils.print_ruler(ruler_length=27), value="_ _", inline=False)
            embed.add_field(name=_("Start Date"), value=f"<t:{int(tournament['start'].timestamp())}>")
            embed.add_field(name=_("# Players per Side"), value=str(tournament['num_players']))
            embed.add_field(name=_("# Signups"), value=str(num_squadrons))
            embed.set_footer(text=_("You need to be an admin of the respective squadron to sign up."))

        elif phase == TOURNAMENT_PHASE.START_GROUP_PHASE:
            message = _("The group phase is now running.")
            tmp = await self.render_matches(tournament=tournament)
            if tmp:
                for field in tmp.fields:
                    embed.add_field(name=field.name, value=field.value, inline=field.inline)
                buffer = await self.render_groups_image(tournament_id)

        elif phase == TOURNAMENT_PHASE.START_ELIMINATION_PHASE:
            message = _("The eliminiation phase is now running.")
            tmp = await self.render_matches(tournament=tournament)
            if tmp:
                for field in tmp.fields:
                    embed.add_field(name=field.name, value=field.value, inline=field.inline)

        elif phase == TOURNAMENT_PHASE.MATCH_RUNNING:
            message = _("A match is running.")
            tmp = await self.render_matches(tournament=tournament)
            if tmp:
                for field in tmp.fields:
                    embed.add_field(name=field.name, value=field.value, inline=field.inline)
                if groups and level == 1:
                    buffer = await self.render_groups_image(tournament_id)

        elif phase == TOURNAMENT_PHASE.TOURNAMENT_FINISHED:
            embed.title = _("THE TOURNAMENT HAS FINISHED!")
            embed.set_thumbnail(url=self.bot.guilds[0].icon.url)
            async with self.apool.connection() as conn:
                # check if we have flown matches already
                cursor = await conn.execute("""
                    SELECT winner_squadron_id FROM tm_matches 
                    WHERE tournament_id = %s 
                    ORDER BY stage DESC LIMIT 1
                """, (tournament_id,))
                winner_id = (await cursor.fetchone())[0]
            winner = utils.get_squadron(node=self.node, squadron_id=winner_id)
            winner_image = winner.get('image_url')
            if winner_image:
                buffer = await create_winner_image(winner_image)
            message = _("### Stand proud, {}!\n"
                        "_Through fire and thunder you prevailed,\n"
                        "When others faltered, you stood strong,\n"
                        "Victory is where you belong!_").format(winner['name'])

        else:
            return

        embed.description = message
        if buffer:
            file = discord.File(fp=BytesIO(buffer), filename=f"tournament_{tournament_id}.png")
            embed.set_image(url=f"attachment://tournament_{tournament_id}.png")
        else:
            file = None

        # create a persistent message
        channel_id = self.get_config().get('channels', {}).get('info')
        await self.bot.setEmbed(embed_name=f"tournament_status_{tournament_id}", embed=embed, file=file,
                                channel_id=channel_id)

    async def render_info_embed(self, tournament_id: int, *,
                                phase: TOURNAMENT_PHASE = TOURNAMENT_PHASE.SIGNUP,
                                match_id: Optional[int] = None) -> None:
        tournament = await self.get_tournament(tournament_id)
        embed = discord.Embed(color=discord.Color.blue(), title=f"Tournament {tournament['name']} Information")
        buffer = None

        if phase == TOURNAMENT_PHASE.MATCH_RUNNING:
            match = await self.get_match(match_id=match_id)
            message = _("A match is running on server {}!").format(utils.escape_string(match['server_name']))
            squadron_blue = utils.get_squadron(node=self.node, squadron_id=match['squadron_blue'])
            squadron_red = utils.get_squadron(node=self.node, squadron_id=match['squadron_red'])
            blue_image = squadron_blue['image_url']
            red_image = squadron_red['image_url']
            if blue_image and red_image:
                buffer = await create_versus_image(blue_image, red_image)
                if not buffer:
                    self.log.debug("Image was not created between {} and {}.".format(squadron_blue['name'],
                                                                                     squadron_red['name']))

            embed.add_field(name=_("Blue"), value=squadron_blue['name'])
            ratings_blue = await Competitive.read_squadron_member_ratings(self.node, match['squadron_blue'])
            ratings_red = await Competitive.read_squadron_member_ratings(self.node, match['squadron_red'])
            win_probability = Competitive.win_probability(ratings_blue, ratings_red)
            embed.add_field(
                name=_("Win-Chance"),
                value=_("Blue") if win_probability > 0.5 else _("Red") if win_probability < 0.5 else _("Draw")
            )
            embed.add_field(name=_("Red"), value=squadron_red['name'])

            embed.add_field(name=_("Round"), value=f"{match['round_number']} of {tournament['rounds']}")
            embed.add_field(name=_("Blue Wins"), value=str(match['squadron_blue_rounds_won']))
            embed.add_field(name=_("Red Wins"), value=str(match['squadron_red_rounds_won']))

        elif phase == TOURNAMENT_PHASE.MATCH_FINISHED:
            match = await self.get_match(match_id=match_id)
            winner = "blue" if match['winner_squadron_id'] == match['squadron_blue'] else "red"
            message = _("The match is over: {} won!").format(winner.upper())
            squadron_blue = utils.get_squadron(node=self.node, squadron_id=match['squadron_blue'])
            squadron_red = utils.get_squadron(node=self.node, squadron_id=match['squadron_red'])
            blue_image = squadron_blue['image_url']
            red_image = squadron_red['image_url']
            if blue_image and red_image:
                buffer = await create_versus_image(blue_image, red_image, winner)
                self.log.debug("Image was not created between {} and {}.".format(squadron_blue['name'],
                                                                                 squadron_red['name']))

            embed.add_field(name=_("Round"), value=f"{match['round_number']} of {tournament['rounds']}")
            embed.add_field(name=_("Blue Wins"), value=str(match['squadron_blue_rounds_won']))
            embed.add_field(name=_("Red Wins"), value=str(match['squadron_red_rounds_won']))

        elif TOURNAMENT_PHASE.TOURNAMENT_FINISHED:
            # remove the status embed
            channel_id = self.get_config().get('channels', {}).get('info')
            channel = self.bot.get_channel(channel_id)
            message = await self.bot.fetch_embed(embed_name=f"tournament_info_{tournament_id}", channel=channel)
            if message:
                await message.delete()
            return

        else:
            return

        embed.description = message
        if buffer:
            file = discord.File(fp=BytesIO(buffer), filename=f"tournament_{tournament_id}.png")
            embed.set_image(url=f"attachment://tournament_{tournament_id}.png")
        else:
            file = None
        # create a persistent message
        channel_id = self.get_config().get('channels', {}).get('info')
        await self.bot.setEmbed(embed_name=f"tournament_info_{tournament_id}", embed=embed, file=file,
                                channel_id=channel_id)

    @tournament.command(description=_('Create a tournament'))
    @app_commands.guild_only()
    @app_commands.autocomplete(campaign=utils.campaign_autocomplete)
    @utils.app_has_role('Admin')
    async def create(self, interaction: discord.Interaction, campaign: str):
        ephemeral = utils.get_ephemeral(interaction)

        modal = TournamentModal()
        # noinspection PyUnresolvedReferences
        await interaction.response.send_modal(modal)
        if await modal.wait():
            await interaction.followup.send(_("Aborted."), ephemeral=ephemeral)
            return
        if modal.error:
            await interaction.followup.send(modal.error, ephemeral=ephemeral)
            return

        campaign_details = await utils.get_campaign(self, campaign)
        if campaign_details['stop']:
            await interaction.followup.send(_("Error: This campaign is already stopped!"), ephemeral=True)
            return
        elif campaign_details['start'] < datetime.now(tz=timezone.utc):
            if not await yn_question(interaction, _("The provided campaign is already started.\n"
                                                    "Are you sure you want to create a tournament for it now?")):
                await interaction.followup.send(_("Aborted."), ephemeral=True)
                return

        try:
            async with self.apool.connection() as conn:
                async with conn.transaction():
                    cursor = await conn.execute("""
                        INSERT INTO tm_tournaments (campaign, rounds, num_players) 
                        VALUES (%s, %s, %s)
                        RETURNING tournament_id
                    """, (campaign, modal.num_rounds, modal.num_players))
                    tournament_id = (await cursor.fetchone())[0]
                    for time in modal.times:
                        async with conn.transaction():
                            await conn.execute("""
                                INSERT INTO tm_available_times (tournament_id, start_time)
                                VALUES (%s, %s::time)
                            """, (tournament_id, time))
            await self.bot.audit(f"created tournament {campaign}.", user=interaction.user)
            await interaction.followup.send(_("Tournament {} created.").format(campaign), ephemeral=ephemeral)
        except UniqueViolation:
            await interaction.followup.send(_("Tournament {} exists already!").format(campaign), ephemeral=True)
            return

        # inform players
        channel = self.get_info_channel()
        if not channel:
            return

        if not await yn_question(interaction, _("Do you want to inform players about the new tournament now?"),
                                 ephemeral=ephemeral):
            await interaction.followup.send(_("Aborted."), ephemeral=True)
            return

        await self.render_status_embed(tournament_id, phase=TOURNAMENT_PHASE.SIGNUP)

    @staticmethod
    def reset_serversettings(server: Server):
        filename = os.path.join(server.instance.home, 'Config', 'serverSettings.lua')
        orig_file = filename + '.orig'
        if os.path.exists(orig_file):
            shutil.copy2(orig_file, filename)

    @tournament.command(description=_('Finish a tournament'))
    @app_commands.guild_only()
    @utils.app_has_role('Admin')
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    async def finish(self, interaction: discord.Interaction, tournament_id: int):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer(ephemeral=True)

        messages = []
        async with self.apool.connection() as conn:
            # checking tournament
            cursor = await conn.execute("""
                SELECT COUNT(*) FROM tm_matches 
                WHERE tournament_id = %s
                AND winner_squadron_id IS NULL
            """, (tournament_id,))
            count = (await cursor.fetchone())[0]
            if count > 0 and not await yn_question(interaction, _("There are unfinished matches. "
                                                                  "Do you really want to finish this tournament?")):
                await interaction.followup.send(_("Aborted."), ephemeral=True)
                return

            # reset serverSettings.lua
            messages.append(_("Resetting all amended serverSettings.lua files ..."))
            msg = await interaction.followup.send(messages[0], ephemeral=utils.get_ephemeral(interaction))
            async for row in await conn.execute("""
                SELECT server_name
                FROM campaigns_servers s
                         JOIN campaigns c ON s.campaign_id = c.id
                         JOIN tm_tournaments t ON c.name = t.campaign
                WHERE t.tournament_id = %s
            """, (tournament_id,)):
                server = self.bot.servers[row[0]]
                self.reset_serversettings(server)

            # close campaign
            messages.append(_("Closing the underlying campaign ..."))
            await msg.edit(content='\n'.join(messages))
            async with conn.transaction():
                cursor = await conn.execute("""
                    UPDATE campaigns SET stop = NOW() AT TIME ZONE 'UTC' 
                    WHERE name = (
                        SELECT name FROM tm_tournaments WHERE tournament_id = %s
                    )
                    RETURNING name
                """, (tournament_id,))
                name = (await cursor.fetchone())[0]

        await self.bot.audit(f"finished tournament {name} and closed the underlying campaign.",
                             user=interaction.user)
        messages.append(_("The tournament {} is finished.").format(name))
        await msg.edit(content='\n'.join(messages))

    @tournament.command(description=_('Delete a tournament'))
    @app_commands.guild_only()
    @utils.app_has_role('Admin')
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    async def delete(self, interaction: discord.Interaction, tournament_id: int):
        if not await yn_question(interaction, _("Do you really want to delete this tournament and all its data?"),
                                 ephemeral=utils.get_ephemeral(interaction)):
            await interaction.followup.send(_("Aborted."), ephemeral=True)
            return
        try:
            async with self.apool.connection() as conn:
                async with conn.transaction():
                    async with conn.cursor(row_factory=dict_row) as cursor:
                        # delete all temp channels
                        async for row in await cursor.execute("""
                            SELECT squadron_blue_channel, squadron_red_channel 
                            FROM tm_matches
                            WHERE tournament_id = %s
                        """, (tournament_id,)):
                            for side in ['blue', 'red']:
                                try:
                                    channel = self.bot.get_channel(row[f'squadron_{side}_channel'])
                                    if channel:
                                        await channel.delete()
                                except NotFound:
                                    pass
                        # delete all data
                        await cursor.execute("""
                            DELETE FROM tm_tournaments 
                            WHERE tournament_id = %s
                            RETURNING *
                        """, (tournament_id, ))
                        row = await cursor.fetchone()

            # delete the info embed
            for what in ['status', 'info']:
                msg = await self.bot.fetch_embed(embed_name=f"tournament_{what}_{tournament_id}", channel=self.get_info_channel())
                try:
                    if msg:
                        await msg.delete()
                except discord.NotFound:
                    pass

            await self.bot.audit(f"deleted tournament {row['campaign']}.", user=interaction.user)
            # noinspection PyUnresolvedReferences
            await interaction.followup.send(_("Tournament {} deleted.").format(row['campaign']),)
        except Exception as ex:
            # noinspection PyUnresolvedReferences
            await interaction.followup.send(_("Error deleting tournament: {}").format(ex), ephemeral=True)

    @tournament.command(description=_('Signup to a tournament'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    @app_commands.rename(squadron_id="squadron")
    @app_commands.autocomplete(squadron_id=utils.squadron_autocomplete_admin)
    @utils.squadron_role_check()
    async def signup(self, interaction: discord.Interaction, tournament_id: int, squadron_id: int):
        config = self.get_config()
        if config.get('use_signup_form', False):
            modal = ApplicationModal()
            # noinspection PyUnresolvedReferences
            await interaction.response.send_modal(modal, ephemeral=True)
            if await modal.wait():
                await interaction.followup.send(_("Aborted."), ephemeral=True)
                return
        else:
            modal = None
            # noinspection PyUnresolvedReferences
            await interaction.response.defer()

        try:
            squadron = utils.get_squadron(self.node, squadron_id=squadron_id)
            if not squadron['role']:
                # noinspection PyUnresolvedReferences
                await interaction.followup.send(
                    _(":warning: To participate in the tournament, your squadron must have an assigned role.\n"
                      "Contact the tournament host to receive your role assignment."), ephemeral=True)

            tournament = await self.get_tournament(tournament_id)
            async with self.apool.connection() as conn:
                cursor = await conn.execute("SELECT COUNT(*) FROM squadron_members WHERE squadron_id = %s",
                                            (squadron_id,))
                row = await cursor.fetchone()
                if row[0] < tournament['num_players']:
                    # noinspection PyUnresolvedReferences
                    await interaction.followup.send(
                        _(":warning: Your squadron does not have enough players to participate in the tournament yet."),
                        ephemeral=True)

                # read the preferred times
                times_options: list[discord.SelectOption] = []
                async for row in await conn.execute("""
                    SELECT time_id, start_time FROM tm_available_times WHERE tournament_id = %s
                """, (tournament_id,)):
                    times_options.append(discord.SelectOption(label=str(row[1]), value=str(row[0])))

                # get available maps
                if not THEATRES:
                    await asyncio.to_thread(MizFile)
                terrain_options = [
                    discord.SelectOption(label=x, value=x)
                    for x in THEATRES.keys()
                    if x not in ['Caucasus', 'MarianaIslands']
                ]
                view = SignupView(times_options, terrain_options)
                msg = await interaction.followup.send(view=view, ephemeral=True)
                try:
                    await view.wait()
                finally:
                    await msg.delete()

                if view.times is None:
                    await interaction.followup.send(_("Aborted."), ephemeral=True)
                    return

                async with conn.transaction():
                    await conn.execute("""
                        INSERT INTO tm_squadrons(tournament_id, squadron_id, application) VALUES (%s, %s, %s)
                    """, (tournament_id, squadron_id, modal.application_text.value if modal else None))
                    for value in view.times:
                        await conn.execute("""
                            INSERT INTO tm_squadron_time_preferences (tournament_id, squadron_id, available_time_id)
                            VALUES (%s, %s, %s)
                        """, (tournament_id, squadron_id, int(value)))
                    for value in view.terrains:
                        await conn.execute("""
                            INSERT INTO tm_squadron_terrain_preferences (tournament_id, squadron_id, terrain)
                            VALUES (%s, %s, %s)
                        """, (tournament_id, squadron_id, value))
            # noinspection PyUnresolvedReferences
            await interaction.followup.send(
                _("Squadron application handed in for tournament {}.").format(tournament['name']), ephemeral=True)
            admin_channel = self.get_admin_channel()
            if admin_channel:
                await admin_channel.send(_("Squadron {} signed up for tournament {}, you can now {} them.").format(
                    squadron['name'], tournament['name'],
                    (await utils.get_command(self.bot, group='tournament', name='verify')).mention))
        except UniqueViolation:
            # noinspection PyUnresolvedReferences
            await interaction.followup.send(_("Squadron already signed up for tournament."), ephemeral=True)

    @tournament.command(description=_('Withdraw from a tournament'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    @app_commands.rename(squadron_id="squadron")
    @app_commands.autocomplete(squadron_id=reject_squadron_autocomplete)
    @utils.squadron_role_check()
    async def withdraw(self, interaction: discord.Interaction, tournament_id: int, squadron_id: int):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()

        tournament = await self.get_tournament(tournament_id)
        if tournament['start'] <= datetime.now(timezone.utc) and not await utils.yn_question(
                interaction, _("You are going to withdraw from a running tournament!\n"
                               "All already scheduled matches will be marked as lost.")):
            # noinspection PyUnresolvedReferences
            await interaction.followup.send("Abort.", ephemeral=True)
            return

        squadron = utils.get_squadron(self.node, squadron_id=squadron_id)
        admin_channel = self.get_admin_channel()
        if admin_channel:
            await admin_channel.send(
                _("Squadron {} forfeit and withdrew from tournament {}.").format(
                    squadron['name'], tournament['name']))

        async with self.apool.connection() as conn:
            async with conn.transaction():
                await conn.execute("""
                    UPDATE tm_squadrons SET status = 'WITHDRAW' 
                    WHERE tournament_id = %s AND squadron_id = %s
                    """, (tournament_id, squadron_id))
                for side in ['red', 'blue']:
                    opposite = 'blue' if side == 'red' else 'red'
                    cursor = await conn.execute(f"""
                        UPDATE tm_matches SET winner_squadron_id = squadron_{opposite} 
                        WHERE tournament_id = %s AND squadron_{side} = %s AND winner_squadron_id IS NULL
                        RETURNING match_id, winner_squadron_id
                    """, (tournament_id, squadron_id))
                    for match_id, winner_squadron_id  in await cursor.fetchall():
                        if admin_channel:
                            winner_squadron = utils.get_squadron(self.node, squadron_id=winner_squadron_id)
                            await admin_channel.send(_("- Match {} vs {} was marked as won by {} due to a forfeit.").format(
                                squadron['name'], winner_squadron['name'], winner_squadron['name']))

        await self.bot.audit(f"withdrew squadron {squadron['name']} from tournament {tournament['name']}.",
                             user=interaction.user)
        await interaction.followup.send(
            _("Your squadron has been withdrawn from tournament {}.").format(tournament['name']), ephemeral=True)
        await self.eventlistener.check_tournament_finished(tournament_id)

    @tournament.command(description=_('Verfiy Applications'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    @app_commands.rename(squadron_id="squadron")
    @app_commands.autocomplete(squadron_id=all_squadron_autocomplete)
    @utils.app_has_role('GameMaster')
    async def verify(self, interaction: discord.Interaction, tournament_id: int, squadron_id: int):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                await cursor.execute("""
                    SELECT t.*, s.name, s.image_url, s.role, s.description, COALESCE(member_count, 0) as member_count
                    FROM tm_squadrons t 
                    JOIN squadrons s ON t.squadron_id = s.id 
                    LEFT JOIN (
                        SELECT squadron_id, COUNT(*) as member_count 
                        FROM squadron_members 
                        GROUP BY squadron_id
                    ) sm ON sm.squadron_id = s.id
                    WHERE t.squadron_id = %s
                """, (squadron_id, ))
                row = await cursor.fetchone()
                if not row:
                    await interaction.followup.send(_("Squadron ID not found."), ephemeral=True)
                    return

                embed = discord.Embed(color=discord.Color.blue(), title=_('Application for Squadron "{}"').format(
                    utils.escape_string(row['name'])))
                embed.description = row['application']
                if row['image_url']:
                    embed.set_thumbnail(url=row['image_url'])
                embed.add_field(name=_("# Members"), value=str(row['member_count']))
                embed.add_field(name=_("Role"), value=self.bot.get_role(row['role']).name if row['role'] else _("n/a"))
                embed.add_field(name=_("State"), value=row['status'])

                terrains = await self.get_terrain_preferences(tournament_id, squadron_id)
                if terrains:
                    embed.add_field(name=_("Terrain Preferences"), value='\n'.join([f"- {x}" for x in terrains]),
                                    inline=False)

                times = await self.get_time_preferences(tournament_id, squadron_id)
                if times:
                    embed.add_field(name=_("Time Preferences"), value='\n'.join([f"- {x}" for x in times]),
                                    inline=False)

        view = ApplicationView(self, tournament_id=tournament_id, squadron_id=squadron_id)
        msg = await interaction.followup.send(embed=embed, view=view, ephemeral=utils.get_ephemeral(interaction))
        try:
           await view.wait()
        finally:
            try:
                await msg.delete()
            except discord.NotFound:
                pass

    @tournament.command(description=_('Generate bracket'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    @utils.app_has_role('GameMaster')
    async def bracket(self, interaction: discord.Interaction, tournament_id: int):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()

        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                await cursor.execute("SELECT * FROM tm_matches WHERE tournament_id = %s", (tournament_id,))
                matches_df = pd.DataFrame(await cursor.fetchall())
                await cursor.execute("""
                    SELECT ts.squadron_id, ts.group_number, s.name 
                    FROM tm_squadrons ts JOIN squadrons s ON ts.squadron_id = s.id
                    WHERE ts.tournament_id = %s
                    AND ts.status = 'ACCEPTED'
                """, (tournament_id,))
                squadrons_df = pd.DataFrame(await cursor.fetchall())

        buffer: bytes = create_tournament_sheet(squadrons_df, matches_df, tournament_id)
        filename = f"tournament_{tournament_id}.xlsx"
        file = discord.File(BytesIO(buffer), filename=filename)
        # noinspection PyUnresolvedReferences
        await interaction.followup.send(file=file, ephemeral=utils.get_ephemeral(interaction))

    @tournament.command(description=_('Export matches'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    @utils.app_has_role('GameMaster')
    async def export(self, interaction: discord.Interaction, tournament_id: int, unflown: Optional[bool] = False):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()

        subquery = "AND winner_squadron_id IS NULL" if unflown else ""
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                await cursor.execute(f"""
                    SELECT match_id, server_name, squadron_red, squadron_blue, round_number, 
                           squadron_red_rounds_won, squadron_blue_rounds_won, winner_squadron_id,
                           match_time
                    FROM tm_matches 
                    WHERE tournament_id = %s
                    {subquery}
                    ORDER BY match_time
                """, (tournament_id,))
                matches_df = pd.DataFrame(await cursor.fetchall())
                await cursor.execute("""
                                     SELECT ts.squadron_id, ts.group_number, s.name
                                     FROM tm_squadrons ts
                                              JOIN squadrons s ON ts.squadron_id = s.id
                                     WHERE ts.tournament_id = %s
                                       AND ts.status = 'ACCEPTED'
                                     ORDER BY group_number
                                     """, (tournament_id,))
                squadrons_df = pd.DataFrame(await cursor.fetchall())
                await cursor.execute("""
                    SELECT server_name FROM campaigns_servers cs
                        JOIN campaigns c ON cs.campaign_id = c.id
                        JOIN tm_tournaments t ON c.name = t.campaign
                    WHERE t.tournament_id = %s         
                """, (tournament_id,))
                servers_df = pd.DataFrame(await cursor.fetchall())

        # Create the mapping dictionary from squadron_id to name
        squadron_name_map = dict(zip(squadrons_df['squadron_id'], squadrons_df['name']))

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Export each dataframe to a different sheet
            matches_df.to_excel(writer, sheet_name='Matches', index=False)
            squadrons_df.to_excel(writer, sheet_name='Squadrons', index=False)
            servers_df.to_excel(writer, sheet_name='Servers', index=False)

            # Get the matches worksheet
            matches_sheet = writer.sheets['Matches']

            # List of columns that contain squadron IDs
            squadron_columns = ['squadron_red', 'squadron_blue', 'winner_squadron_id']

            # Find the column indices
            squadron_cols = {}
            server_col = None
            for idx, col in enumerate(matches_df.columns):
                if col in squadron_columns:
                    squadron_cols[col] = idx + 1  # +1 because Excel columns start at 1
                elif col == 'server_name':
                    server_col = idx + 1

            # Create the data validation if we found any squadron columns
            if squadron_cols:
                # Get the number of rows in matches_df
                num_rows = len(matches_df) + 1  # +1 for header

                # Create the formula for the dropdown (references the name column in Squadrons sheet)
                formula = f'=Squadrons!$C$2:$C${len(squadrons_df) + 1}'  # Assuming 'name' is in column C

                # Create data validation object
                dv = DataValidation(type="list", formula1=formula)
                matches_sheet.add_data_validation(dv)

                # Apply data validation to squadron columns
                for row in range(2, num_rows + 1):  # Start from row 2 to skip header
                    for col_name, col_index in squadron_cols.items():
                        cell = matches_sheet.cell(row=row, column=col_index)
                        cell.value = squadron_name_map.get(matches_df.iloc[row - 2][col_name])
                        dv.add(cell)

            if server_col:
                # Create the formula for server dropdown
                server_formula = f'=Servers!$A$2:$A${len(servers_df) + 1}'  # Assuming server_name is in column A

                # Create data validation object for servers
                server_dv = DataValidation(type="list", formula1=server_formula)
                matches_sheet.add_data_validation(server_dv)

                # Apply server data validation
                for row in range(2, num_rows + 1):  # Start from row 2 to skip the header
                    cell = matches_sheet.cell(row=row, column=server_col)
                    cell.value = matches_df.iloc[row - 2]['server_name']
                    server_dv.add(cell)

            # hide the match_id
            matches_sheet.column_dimensions['A'].hidden = True

            for worksheet in [matches_sheet, writer.sheets['Squadrons'], writer.sheets['Servers']]:
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    # Setting width with some padding
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        buffer.seek(0)
        try:
            # noinspection PyUnresolvedReferences
            await interaction.followup.send(file=discord.File(fp=buffer, filename=f'tournament_{tournament_id}.xlsx'))
        finally:
            buffer.close()

    @tournament.command(description=_('Show squadron preferences'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=tournament_autocomplete)
    @utils.app_has_role('GameMaster')
    async def preferences(self, interaction: discord.Interaction, tournament_id: Optional[int] = None):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()
        report = Report(bot=self.bot, plugin=self.plugin_name, filename="preferences.json")
        env = await report.render(tournament_id=tournament_id)
        try:
            await interaction.followup.send(file=discord.File(env.buffer, filename=env.filename), ephemeral=True)
        finally:
            env.buffer.close()

    # New command group "/match"
    match = Group(name="match", description=_("Commands to manage matches in a tournament"))

    async def generate_group_stage(self, interaction: discord.Interaction, tournament_id: int,
                                   num_groups: int) -> list[tuple[int, int]]:
        async with self.apool.connection() as conn:
            cursor = await conn.execute("""
                SELECT squadron_id, group_number
                FROM tm_squadrons
                WHERE tournament_id = %s
                  AND status = 'ACCEPTED'
            """, (tournament_id,))

            squadrons: list[tuple[int, int]] = await cursor.fetchall()

            if squadrons[0][1] is None or await utils.yn_question(
                    interaction, question=_("Do you want to re-generate your groups?")):
                groups = create_groups(squadrons, num_groups)
                # update the groups, if any
                for idx, group in enumerate(groups):
                    for squadron_id in group:
                        await conn.execute("""
                            UPDATE tm_squadrons
                            SET group_number = %s
                            WHERE tournament_id = %s
                             AND squadron_id = %s
                                           """, (idx + 1, tournament_id, squadron_id))
            else:
                groups = squadrons_to_groups(squadrons)

        return create_group_matches(groups)

    async def generate_elimination_stage(self, interaction: discord.Interaction, tournament_id: int,
                                         level: int) -> list[tuple[int, int]]:
        async with self.apool.connection() as conn:
            cursor = await conn.execute("SELECT COUNT(DISTINCT(group_number)) FROM tm_squadrons WHERE tournament_id = %s",
                                        (tournament_id,))
            has_groups = (await cursor.fetchone())[0] > 0

            if level == 0:
                cursor = await conn.execute("""
                    SELECT squadron_id, 0 AS rank 
                    FROM tm_squadrons 
                    WHERE tournament_id = %s AND status = 'ACCEPTED'
                """, (tournament_id,))
            elif level == 1 and has_groups:
                cursor = await conn.execute("""
                    WITH squadron_stats AS (SELECT s.squadron_id,
                                                   s.group_number,
                                                   m.stage,
                                                   COUNT(CASE WHEN m.winner_squadron_id = s.squadron_id THEN 1 END) as matches_won,
                                                   SUM(CASE
                                                           WHEN m.winner_squadron_id = m.squadron_blue
                                                               THEN m.squadron_blue_rounds_won
                                                           WHEN m.winner_squadron_id = m.squadron_red
                                                               THEN m.squadron_red_rounds_won
                                                           ELSE 0 END) as total_rounds_won,
                                                   ROW_NUMBER() OVER (
                                                       PARTITION BY s.group_number
                                                       ORDER BY 4, 5
                                                       )  as rank
                                            FROM tm_matches m
                                                     LEFT JOIN tm_squadrons s
                                                               ON (m.tournament_id =
                                                                   s.tournament_id AND
                                                                   s.squadron_id =
                                                                   m.winner_squadron_id)
                                            WHERE m.tournament_id = %s
                                              AND m.winner_squadron_id IS NOT NULL
                                            GROUP BY s.squadron_id, s.group_number, m.stage)
                    SELECT squadron_id, group_number
                    FROM squadron_stats
                    WHERE rank <= 2
                    ORDER BY group_number, rank
                """, (tournament_id,))
            else:
                cursor = await conn.execute("""
                    SELECT winner_squadron_id, 0 AS rank 
                    FROM tm_matches 
                    WHERE tournament_id = %s
                    AND stage = %s
                """, (tournament_id, level))

            config = self.get_config()
            squadrons: list[tuple[int, float]] = []
            # read all squadrons and their ratings
            for row in await cursor.fetchall():
                if config.get('match_generation', 'trueskill') == 'trueskill':
                    rating = await Competitive.trueskill_squadron(self.node, row[0])
                    squadrons.append((row[0], Competitive.calculate_rating(rating)))
                else:
                    squadrons.append((row[0], row[1]))

        return create_elimination_matches(squadrons)

    @match.command(description=_('Generate matches'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @utils.app_has_role('GameMaster')
    async def generate(self, interaction: discord.Interaction, tournament_id: int, num_groups: Optional[int] = 4):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()
        tournament = await self.get_tournament(tournament_id)

        async with self.apool.connection() as conn:
            # check if we have flown matches already
            cursor = await conn.execute("""
                SELECT stage AS level, COUNT(*) AS all_matches, 
                       COALESCE(SUM(CASE WHEN winner_squadron_id IS NOT NULL THEN 1 ELSE 0 END), 0) AS flown_matches
                FROM tm_matches 
                WHERE tournament_id = %s
                GROUP BY stage
                ORDER BY 1 DESC LIMIT 1
            """, (tournament_id,))
            row = await cursor.fetchone()
            if row:
                level, all_matches, flown_matches = row
            else:
                level = all_matches = flown_matches = 0

            if all_matches == flown_matches and flown_matches == 1:
                await interaction.followup.send(_("This tournament is finished already!"))
                return

            elif 0 < flown_matches < all_matches:
                await interaction.followup.send(
                    _("Can't generate new matches, not all matches of this stage have been flown yet."), ephemeral=True)
                return

            elif all_matches > 0 and flown_matches == 0 and not await utils.yn_question(
                    interaction,
                    question=_("Do you want to generate new matches for tournament {}?").format(tournament['name']),
                    message=_("This will overwrite any existing **unflown** matches!")):
                # noinspection PyUnresolvedReferences
                await interaction.followup.send("Abort.", ephemeral=True)
                return

            # read all available servers
            servers: list[str] = []
            async for row in await conn.execute("""
                SELECT server_name FROM campaigns_servers s
                JOIN campaigns c ON s.campaign_id = c.id
                JOIN tm_tournaments t ON c.name = t.campaign
                WHERE t.tournament_id = %s
            """, (tournament_id,)):
                servers.append(row[0])

            try:
                if flown_matches > 0 and flown_matches == all_matches:
                    await interaction.followup.send(_("Generating knockout matches for this tournament ..."))
                    phase = TOURNAMENT_PHASE.START_ELIMINATION_PHASE
                    matches = await self.generate_elimination_stage(interaction, tournament_id, level)
                elif level <= 1:
                    if await utils.yn_question(interaction, "Do you want to generate a group stage?"):
                        await interaction.followup.send(_("Generating group matches for this tournament ..."))
                        phase = TOURNAMENT_PHASE.START_GROUP_PHASE
                        matches = await self.generate_group_stage(interaction, tournament_id, num_groups)
                    else:
                        await interaction.followup.send(_("Generating knockout matches for this tournament ..."))
                        phase = TOURNAMENT_PHASE.START_ELIMINATION_PHASE
                        matches = await self.generate_elimination_stage(interaction, tournament_id, 0)

                async with conn.transaction():
                    # delete old matches
                    cursor = await conn.execute("""
                        DELETE FROM tm_matches WHERE tournament_id = %s AND winner_squadron_id IS NULL
                        RETURNING *
                    """, (tournament_id, ))
                    # we have not deleted anything, so we are creating a new stage
                    if cursor.rowcount == 0:
                        level += 1

                    # store new matches in the database
                    for idx, match in enumerate(matches):
                        server = servers[idx % len(servers)]
                        await conn.execute("""
                           INSERT INTO tm_matches(tournament_id, stage, server_name, squadron_red, squadron_blue)
                           VALUES (%s, %s, %s, %s, %s)
                        """, (tournament_id, level, server, match[0], match[1]))

                    if level > 1:
                        # check for eliminations
                        async for row in await conn.execute(f"""
                            WITH active_squadrons AS (
                                SELECT DISTINCT squadron_id
                                FROM (
                                    SELECT squadron_blue AS squadron_id
                                    FROM tm_matches
                                    WHERE tournament_id = %(tournament_id)s
                                    AND winner_squadron_id IS NULL
                                    UNION ALL
                                    SELECT squadron_red AS squadron_id
                                    FROM tm_matches
                                    WHERE tournament_id = %(tournament_id)s
                                    AND winner_squadron_id IS NULL
                                ) all_active
                            ),
                            previous_stage_squadrons AS (
                                SELECT DISTINCT squadron_id
                                FROM (
                                    SELECT squadron_blue AS squadron_id
                                    FROM tm_matches
                                    WHERE tournament_id = %(tournament_id)s
                                    AND winner_squadron_id IS NOT NULL
                                    UNION ALL
                                    SELECT squadron_red AS squadron_id
                                    FROM tm_matches
                                    WHERE tournament_id = %(tournament_id)s
                                    AND winner_squadron_id IS NOT NULL
                                ) all_previous
                            )
                            SELECT squadron_id 
                            FROM previous_stage_squadrons
                            WHERE squadron_id NOT IN (SELECT squadron_id FROM active_squadrons)
                        """, {"tournament_id": tournament_id}):
                            embed = discord.Embed(title=_("You have been eliminated!"), color=discord.Color.blue())
                            embed.set_thumbnail(url="https://upload.wikimedia.org/wikipedia/commons/thumb/d/df/Uncle_Sam_%28pointing_finger%29.png/960px-Uncle_Sam_%28pointing_finger%29.png")
                            embed.description = _("You have been eliminated from the tournament {}.\n"
                                                  "Thank you for your participation!").format(tournament['name'])
                            await self.inform_squadron(tournament_id=tournament_id, squadron_id=row[0], embed=embed)

                await self.bot.audit(f"generated matches for tournament {tournament['name']}.",
                                     user=interaction.user)
            except ValueError as ex:
                await interaction.followup.send(f"Error: {ex}")
                return

        asyncio.create_task(self.render_status_embed(tournament_id, phase=phase))
        embed = await self.render_matches(tournament=tournament, unflown=True)
        await interaction.followup.send(_("{} matches generated:").format(len(matches)), embed=embed,
                                        ephemeral=utils.get_ephemeral(interaction))

    @match.command(description=_('Create a manual match'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.describe(stage=_("The level of your match. If all matches on one level are finished, increase the level by one."))
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @app_commands.autocomplete(stage=stage_autocomplete)
    @app_commands.autocomplete(server_name=server_autocomplete)
    @app_commands.autocomplete(squadron_blue=valid_squadron_autocomplete)
    @app_commands.autocomplete(squadron_red=valid_squadron_autocomplete)
    @app_commands.autocomplete(day=date_autocomplete)
    @app_commands.autocomplete(time=time_autocomplete)
    @utils.app_has_role('GameMaster')
    async def create(self, interaction: discord.Interaction, tournament_id: int, stage: app_commands.Range[int, 1, 7],
                     server_name: str, squadron_blue: int, squadron_red: int, day: int, time: int):
        match_time = datetime.fromtimestamp(day, tz=timezone.utc) + timedelta(seconds=time)
        async with self.apool.connection() as conn:
            async with conn.transaction():
                await conn.execute("""
                    INSERT INTO tm_matches(tournament_id, stage, match_time, server_name, squadron_red, squadron_blue) 
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (tournament_id, stage, match_time, server_name, squadron_red, squadron_blue))

        tournament = await self.get_tournament(tournament_id)
        blue = utils.get_squadron(self.node, squadron_id=squadron_blue)
        red = utils.get_squadron(self.node, squadron_id=squadron_red)
        await self.bot.audit(f"created a match for tournament {tournament['name']} "
                               f"between squadrons {blue['name']} and {red['name']}.", user=interaction.user)
        # noinspection PyUnresolvedReferences
        await interaction.response.send_message(_("Match created."), ephemeral=utils.get_ephemeral(interaction))

    async def render_matches(self, tournament: dict, unflown: Optional[bool] = False) -> Optional[discord.Embed]:
        embed = discord.Embed(color=discord.Color.blue())
        embed.title = _("Matches for Tournament {}").format(tournament['name'])#
        embed.set_thumbnail(url=self.bot.guilds[0].icon.url)
        squadrons_blue = []
        squadrons_red = []
        status = []
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                await cursor.execute("""
                    SELECT * FROM tm_matches 
                    WHERE tournament_id = %(tournament_id)s
                    AND stage = (
                        SELECT MAX(stage) 
                        FROM tm_matches 
                        WHERE tournament_id = %(tournament_id)s
                    )
                    ORDER BY winner_squadron_id DESC, round_number DESC, match_time
                """, tournament)
                for row in await cursor.fetchall():
                    if row['winner_squadron_id'] and unflown:
                        continue
                    squadrons_blue.append(utils.get_squadron(self.node, squadron_id=row['squadron_blue'])['name'])
                    squadrons_red.append(utils.get_squadron(self.node, squadron_id=row['squadron_red'])['name'])
                    if row['winner_squadron_id']:
                        winner = utils.get_squadron(self.node, squadron_id=row['winner_squadron_id'])
                        status.append(_("Winner: {}").format(winner['name']))
                    elif row['round_number'] == 0:
                        status.append("Not started" if row['match_time'] is None else
                                      f"<t:{int(row['match_time'].replace(tzinfo=timezone.utc).timestamp())}:R>")
                    else:
                        status.append(_("Round: {}, {} : {}").format(row['round_number'],
                                                                     row['squadron_blue_rounds_won'],
                                                                     row['squadron_red_rounds_won']))
        # no data
        if not len(squadrons_blue):
            return None

        embed.add_field(name=_("Blue"), value='\n'.join(squadrons_blue), inline=True)
        embed.add_field(name=_("Red"), value='\n'.join(squadrons_red), inline=True)
        embed.add_field(name=_("Status"), value='\n'.join(status), inline=True)
        return embed

    @match.command(name="list", description=_('List matches'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @utils.app_has_role('DCS')
    async def _list(self, interaction: discord.Interaction, tournament_id: int, unflown: bool = False):
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()
        tournament = await self.get_tournament(tournament_id)
        if not tournament:
            await interaction.followup.send(_("Tournament not found."), ephemeral=True)
            return

        embed = await self.render_matches(tournament, unflown)
        if not embed:
            await interaction.followup.send(_("No matches for tournament {} found.").format(tournament['name']),
                                            ephemeral=True)
            return
        await interaction.followup.send(embed=embed)

    async def setup_server_for_match(self, msg: discord.Message, embed: discord.Embed, server: Server, match: dict,
                                     channels: dict):
        config = self.get_config(server)
        squadrons = {
            'blue': utils.get_squadron(self.node, squadron_id=match['squadron_blue']),
            'red': utils.get_squadron(self.node, squadron_id=match['squadron_red'])
        }

        # backup the serversettings.lua
        filename = os.path.join(server.instance.home, 'Config', 'serverSettings.lua')
        orig_file = filename + '.orig'
        if not os.path.exists(orig_file):
            shutil.copy2(filename, orig_file)

        # enable / overwrite any coalition settings
        server.locals["coalitions"] = {
            "lock_time": "1 day",
            "allow_players_pool": False,
            "blue_role": squadrons['blue']['role'],
            "red_role": squadrons['red']['role']
        }

        # set coalition channels
        server.locals['channels']['blue'] = channels['blue']
        server.locals['channels']['red'] = channels['red']

        # setup streamer channel (replicates all events from red and blue)
        streamer_channel = config.get('channels', {}).get('streamer')
        if streamer_channel:
            server.locals['channels']['blue_events'] = streamer_channel
            server.locals['channels']['red_events'] = streamer_channel

        # dirty but works
        server._channels.clear()

        # Server should start paused
        advanced = server.settings['advanced']
        advanced |= {
            "resume_mode": 0,
            "pause_on_load": True,
            "sav_autosave": False,
            "maxPing": 300
        }

        # sanitize the server
        server.settings['require_pure_textures'] = True
        server.settings['require_pure_models'] = True
        server.settings['require_pure_clients'] = True
        server.settings['require_pure_scripts'] = True
        server.settings['listShuffle'] = False
        server.settings['listLoop'] = False
        if not config.get('allow_exports', False):
            advanced |= {
                "allow_ownship_export": False,
                "allow_object_export": False,
                "allow_sensor_export": False,
                "allow_players_pool": False,
                "disable_events": True
            }

        # set coalition passwords
        if config.get('coalition_passwords'):
            embed.description += _("\n- Setting coalition passwords...")
            await msg.edit(embed=embed)
            for coalition in [Coalition.BLUE, Coalition.RED]:
                password = str(random.randint(100000, 999999))
                await server.setCoalitionPassword(coalition, password)
                channel = self.bot.get_channel(channels[coalition.value])
                _embed = discord.Embed(color=discord.Color.blue(), title=_("**Get your team ready!**\n"))
                if squadrons[coalition.value]['image_url']:
                    _embed.set_thumbnail(url=squadrons[coalition.value]['image_url'])
                _embed.add_field(name=_("Coalition"), value=coalition.value.upper(), inline=True)
                _embed.add_field(name=_("Password"), value=password, inline=True)
                _embed.set_footer(text=_("You must not share the password with anyone outside your squadron!\n"
                                        "You will stay on the {} side throughout the whole match.").format(
                    coalition.value))
                await channel.send(embed=_embed)

        # assign all members of the respective squadrons to the respective side
        embed.description += _("\n- Setting coalitions for players...")
        await msg.edit(embed=embed)
        async with self.apool.connection() as conn:
            async with conn.transaction():
                await conn.execute("DELETE FROM coalitions WHERE server_name = %s", (server.name, ))
                for coalition in ['blue', 'red']:
                    await conn.execute(f"""
                        INSERT INTO coalitions(server_name, player_ucid, coalition) 
                            SELECT %s, s.player_ucid, '{coalition}' 
                            FROM squadron_members s JOIN tm_matches m ON s.squadron_id = m.squadron_{coalition}
                            WHERE m.match_id = %s
                        ON CONFLICT (server_name, player_ucid) DO UPDATE SET coalition = '{coalition}'
                    """, (server.name, match['match_id']))

    async def prepare_mission(self, server: Server, match_id: int, round_number: int,
                              mission_id: Optional[int] = None) -> str:
        config = self.get_config(server)
        # set startindex or use last mission
        if mission_id is not None and server.settings['listStartIndex'] != mission_id + 1:
            await server.setStartIndex(mission_id + 1)

        # change the mission
        filename = await server.get_current_mission_file()
        # create a writable mission
        new_filename = utils.create_writable_mission(filename)
        # get the orig file
        orig_filename = utils.get_orig_file(new_filename)
        # and copy the orig file over
        shutil.copy2(orig_filename, new_filename)

        self.log.debug(f"Changing mission {new_filename}")
        miz = await asyncio.to_thread(MizFile, new_filename)
        preset_file = os.path.join(self.node.config_dir, config.get('presets', {}).get('file', 'presets.yaml'))
        # apply the initial presets
        for preset in config.get('presets', {}).get('initial', []):
            self.log.debug(f"Applying initial preset: {preset} ...")
            await asyncio.to_thread(
                miz.apply_preset,
                utils.get_preset(self.node, preset, filename=preset_file)
            )

        if round_number %2 == 0:
            for preset in config.get('presets', {}).get('even', []):
                self.log.debug(f"Applying even-round preset: {preset} ...")
                await asyncio.to_thread(
                    miz.apply_preset,
                    utils.get_preset(self.node, preset, filename=preset_file)
                )
        else:
            for preset in config.get('presets', {}).get('uneven', []):
                self.log.debug(f"Applying uneven-round preset: {preset} ...")
                await asyncio.to_thread(
                    miz.apply_preset,
                    utils.get_preset(self.node, preset, filename=preset_file)
                )

        async with self.apool.connection() as conn:
            async with conn.transaction():
                # apply the squadron presets
                for side in ['blue', 'red']:
                    # apply persistent presets
                    async for row in await conn.execute(f"""
                        SELECT preset, config FROM tm_persistent_choices c
                        JOIN tm_matches m ON c.match_id = m.match_id AND c.squadron_id = m.squadron_{side}
                        WHERE m.match_id = %s 
                    """, (match_id,)):
                        self.log.debug(f"Applying persistent preset for side {side}: {row[0]} ...")
                        await asyncio.to_thread(
                            miz.apply_preset,
                            utils.get_preset(self.node, row[0], filename=preset_file),
                            side=side, **row[1]
                        )
                    # apply choices
                    async for row in await conn.execute(f"""
                        SELECT preset, config FROM tm_choices c 
                        JOIN tm_matches m ON c.match_id = m.match_id AND c.squadron_id = m.squadron_{side}
                        WHERE m.match_id = %(match_id)s AND m.choices_{side}_ack = TRUE
                    """, {"match_id": match_id}):
                        self.log.debug(f"Applying custom preset for side {side}: {row[0]} ...")
                        await asyncio.to_thread(
                            miz.apply_preset,
                            utils.get_preset(self.node, row[0], filename=preset_file),
                            side=side, **row[1]
                        )

                # delete the choices from the database and update the acknoledgement
                await conn.execute("DELETE FROM tm_choices WHERE match_id = %s", (match_id,))
                await conn.execute("""
                    UPDATE tm_matches 
                    SET choices_blue_ack=FALSE, choices_red_ack=FALSE 
                    WHERE match_id = %s
                """, (match_id,))

        miz.save(new_filename)
        if new_filename != filename:
            self.log.info(f"  => New mission written: {new_filename}")
            await server.replaceMission(int(server.settings['listStartIndex']), new_filename)
        return new_filename

    async def change_tacview_output(self, server: Server, results: int):
        extensions = await server.list_extension()
        if 'Tacview' in extensions:
            await server.run_on_extension(
                extension='Tacview',
                method='change_config',
                config={
                    "target": f"<id:{results}>"
                }
            )

    def find_mission_in_list(self, mission_list: list[str], mission: str) -> int:
        for idx, mission_file in enumerate(mission_list):
            if mission in mission_file:
                return idx
        else:
            raise IndexError(f"Mission {mission} not found in mission list.")

    async def get_mission(self, server: Server, tournament_id: int, match: dict) -> Optional[str]:
        config = self.get_config(server)
        if isinstance(config.get('mission'), str):
            return config['mission']
        prefs_red = set(await self.get_terrain_preferences(tournament_id, match['squadron_red']))
        prefs_blue = set(await self.get_terrain_preferences(tournament_id, match['squadron_blue']))

        common_maps = prefs_red & prefs_blue
        common_maps.add('Caucasus')
        common_maps.add('MarianaIslands')

        if isinstance(config.get('mission'), list):
            all_missions = await server.getMissionList()
            missions = {}
            for mission in config['mission']:
                try:
                    mission_id = self.find_mission_in_list(all_missions, mission)
                    miz: MizFile = await asyncio.to_thread(MizFile, all_missions[mission_id])
                    missions[mission] = miz.theatre
                except IndexError:
                    self.log.warning(f"Mission {mission} not found in mission list, skipping ...")
                except Exception:
                    self.log.error(f"Can't read mission {mission}, skipping ...")
            config['mission'] = missions

        if isinstance(config.get('mission'), dict):
            valid_missions = [
                mission for mission, terrain in config['mission'].items()
                if terrain in common_maps
            ]
            return random.choice(valid_missions) if valid_missions else None
        return None

    async def start_match(self, server: Server, tournament_id: int, match_id: int, mission_id: Optional[int] = None,
                          round_number: Optional[int] = None):
        tournament = await self.get_tournament(tournament_id)
        match = await self.get_match(match_id)

        squadrons = {
            'blue': utils.get_squadron(self.node, squadron_id=match['squadron_blue']),
            'red': utils.get_squadron(self.node, squadron_id=match['squadron_red'])
        }

        # set the correct round number
        if not round_number:
            if match['round_number'] == 0:
                round_number = 1
            else:
                round_number = match['round_number'] if not match['winner_squadron_id'] else match['round_number'] + 1

        # Start the next round
        async with self.apool.connection() as conn:
            async with conn.transaction():
                await conn.execute("""
                    UPDATE tm_matches 
                    SET round_number = %s, 
                        match_time = NOW() AT TIME ZONE 'UTC',
                        choices_blue_ack = FALSE, 
                        choices_red_ack = FALSE
                    WHERE match_id = %s
                """, (round_number, match_id))

        channel_id = self.get_config(server).get('channels', {}).get('admin')
        channel = self.bot.get_channel(channel_id) or self.bot.get_admin_channel(server)
        if not server:
            await channel.send(_("Server {} not found.").format(match['server_name']))
            return

        embed = discord.Embed(color=discord.Color.blue(), title=_("Match Setup"))
        embed.description = _("- Creating the squadron channels ...")
        embed.set_thumbnail(url=TRAFFIC_LIGHTS['red'])
        msg = await channel.send(embed=embed)
        try:
            # open the channels
            channels = await self.open_channel(match_id, server)
        except (ValueError, PermissionError) as ex:
            await channel.send(_("Error during opening the channels: {}").format(ex))
            return

        # inform the squadrons that they can choose
        embed.description += _("\n- Inform the squadrons and wait for their initial choice ...")
        await msg.edit(embed=embed)
        self.eventlistener.tournaments[server.name] = tournament

        config = self.get_config(server)
        min_costs = min(choice['costs'] for choice in config['presets']['choices'].values())
        async with self.apool.connection() as conn:
            async with conn.transaction():
                for side in ['blue', 'red']:
                    squadron = await self.get_squadron(tournament_id, squadrons[side]['id'])
                    if squadron.points >= min_costs:
                        channel = self.bot.get_channel(channels[side])
                        await channel.send(_("You can now use {} to chose your customizations!\n"
                                             "If you do not want to change anything, "
                                             "please run it and say 'Skip this round'").format(
                            (await utils.get_command(self.bot, group=self.match.name,
                                                     name=self.customize.name)).mention))
                    else:
                        await conn.execute(f"UPDATE tm_matches SET choices_{side}_ack = TRUE WHERE match_id = %s",
                                           (match_id,))

        # wait until all choices are finished
        await self.eventlistener.wait_until_choices_finished(server)

        # preparing the server
        embed.description += _("\n- Preparing server {} for the match ...").format(match['server_name'])
        await msg.edit(embed=embed)

        # make sure the server is stopped
        if server.status not in [Status.STOPPED, Status.SHUTDOWN]:
            embed.description += _("\n- Shutting down server {} ...").format(match['server_name'])
            await msg.edit(embed=embed)
            await server.shutdown()
            await channel.send(_("Server {} shut down.").format(match['server_name']))

        # change settings
        await self.setup_server_for_match(msg, embed, server, match, channels)

        # find a mission
        mission_list = await server.getMissionList()
        if mission_id is None:
            mission = await self.get_mission(server, tournament_id, match)
            if isinstance(mission, int):
                mission_id = mission
            elif isinstance(mission, str):
                try:
                    mission_id = self.find_mission_in_list(mission_list, mission)
                except IndexError:
                    self.log.warning(f"Mission {mission} not found in mission list! Using default mission.")

        # mission id is still not set, use the default mission
        if mission_id is None:
            mission_id = server.settings['listStartIndex'] - 1

        # prepare the mission
        embed.description += _("\n- Preparing mission {} ...").format(os.path.basename(mission_list[mission_id]))
        await msg.edit(embed=embed)
        await self.prepare_mission(server, match_id, round_number, mission_id)

        # Starting the server up again
        embed.description += _("\n- Starting server {} ...").format(match['server_name'])
        embed.set_thumbnail(url=TRAFFIC_LIGHTS['amber'])
        await msg.edit(embed=embed)
        for i in range(0, 3):
            try:
                await server.startup(modify_mission=False, use_orig=False)
                break
            except (TimeoutError, asyncio.TimeoutError):
                addon = '. Retrying in 5 seconds ...' if i < 2 else '. Giving up.'
                self.log.warning(f"Timeout while starting server {server.name}{addon}")
                if i < 2:
                    await asyncio.sleep(5)
        else:
            embed.description = _("## Error during the startup of server\n{}: Timeout").format(server.name)
            embed.set_thumbnail(url=TRAFFIC_LIGHTS['red'])
            await msg.edit(embed=embed)
            return
        # Check if we need to forward Tacview
        results = config.get('channels', {}).get('results', -1)
        if results > 0:
            await self.change_tacview_output(server, results)
        embed.description += _("\n- Server {} started. Inform squadrons ...").format(match['server_name'])
        embed.set_thumbnail(url=TRAFFIC_LIGHTS['green'])
        await msg.edit(embed=embed)
        # inform everyone
        for side in ['blue', 'red']:
            channel: TextChannel = self.bot.get_channel(channels[side])
            embed = discord.Embed(color=discord.Color.blue(), title=_("The match is starting NOW!"))
            if squadrons[side]['image_url']:
                embed.set_thumbnail(url=squadrons[side]['image_url'])
            embed.description = _("You can **now** join the server.")
            embed.add_field(name=_("Server"), value=server.name)
            embed.add_field(name=_("IP:Port"), value=f"{server.node.public_ip}:{server.settings['port']}")
            embed.add_field(name=_("Password"), value=server.settings.get('password', ''))
            embed.add_field(name=_("Terrain"), value=server.current_mission.map)
            embed.set_footer(text=_("Please keep in mind that you can only use {} planes!").format(
                tournament['num_players']))
            await channel.send(embed=embed)
        info = self.get_info_channel()
        if info:
            asyncio.create_task(self.render_status_embed(tournament_id, phase=TOURNAMENT_PHASE.MATCH_RUNNING))
            asyncio.create_task(self.render_info_embed(tournament_id, phase=TOURNAMENT_PHASE.MATCH_RUNNING,
                                                       match_id=match_id))

    @match.command(description=_('Start a match'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @app_commands.rename(match_id="match")
    @app_commands.autocomplete(match_id=active_matches_autocomplete)
    @app_commands.rename(mission_id="mission")
    @app_commands.autocomplete(mission_id=mission_autocomplete)
    @utils.app_has_role('GameMaster')
    async def start(self, interaction: discord.Interaction, tournament_id: int, match_id: int,
                    mission_id: Optional[int] = None, round_number: Optional[int] = None):
        ephemeral = utils.get_ephemeral(interaction)
        # noinspection PyUnresolvedReferences
        await interaction.response.defer()
        match = await self.get_match(match_id)
        if match['winner_squadron_id']:
            await interaction.followup.send(_("This match is over already. You can not start another round."),
                                            ephemeral=True)
            return

        tournament = await self.get_tournament(tournament_id)
        squadrons = {
            'blue': utils.get_squadron(self.node, squadron_id=match['squadron_blue']),
            'red': utils.get_squadron(self.node, squadron_id=match['squadron_red'])
        }

        # set the correct round number
        if not round_number:
            if match['round_number'] == 0:
                round_number = 1
            else:
                round_number = match['round_number'] if not match['winner_squadron_id'] else match['round_number'] + 1

        if not await yn_question(
                interaction,
                _("Do you want to start round {} of the match between\n{} and {}?").format(
                    round_number, squadrons['blue']['name'], squadrons['red']['name']
                ), ephemeral=ephemeral):
            await interaction.followup.send(_("Aborted."), ephemeral=True)
            return

        server = self.bot.servers.get(match['server_name'])
        if not server:
            await interaction.followup.send(_("Server {} not found.").format(match['server_name']), ephemeral=True)
            return

        # start the match
        await self.start_match(server, tournament_id, match_id, mission_id, round_number)

        # audit event
        await self.bot.audit(f"started a match for tournament {tournament['name']} "
                             f"between squadrons {squadrons['blue']['name']} and {squadrons['red']['name']}.",
                             user=interaction.user)

    async def open_channel(self, match_id: int, server: Server) -> dict[str, int]:
        config = self.get_config(server)
        match = await self.get_match(match_id)
        category: CategoryChannel = self.bot.get_channel(config['channels']['category'])
        channels = {}
        for side in ['blue', 'red']:
            squadron = utils.get_squadron(self.node, squadron_id=match[f'squadron_{side}'])
            if not squadron['role']:
                raise ValueError(f"Squadron {squadron['name']} does not have a role set.")
            role = self.bot.get_role(squadron['role'])
            if not role:
                raise ValueError(f"Squadron {squadron['name']} has an invalid role.")

            async with self.apool.connection() as conn:
                cursor = await conn.execute(f"""
                    SELECT squadron_{side}_channel FROM tm_matches WHERE match_id = %s
                """, (match_id, ))
                channel_id = (await cursor.fetchone())[0]
                if channel_id == -1 or not self.bot.get_channel(channel_id):
                    opponent = utils.get_squadron(
                        self.node, squadron_id=match['squadron_red' if side == 'blue' else 'squadron_blue'])
                    channel_name = f"{squadron['name']} vs {opponent['name']}"
                    channel: TextChannel = await category.create_text_channel(name=channel_name)
                    channel_id = channel.id
                    overwrite = discord.PermissionOverwrite(
                        view_channel=True,
                        send_messages=True,
                        embed_links=True,
                        attach_files=True,
                        read_message_history=True,
                        use_application_commands=True
                    )
                    try:
                        await channel.set_permissions(role, overwrite=overwrite)
                        await channel.send(role.mention + _(", this is your private channel during the upcoming match!"))
                    except discord.Forbidden:
                        raise PermissionError("You need to give the bot the Manage Roles permission!")

                    async with conn.transaction():
                        await conn.execute(f"""
                            UPDATE tm_matches SET squadron_{side}_channel = %s 
                            WHERE match_id = %s
                        """, (channel_id, match_id))

            channels[side] = channel_id

        return channels

    async def close_channel(self, match_id: int):
        match = await self.get_match(match_id)
        for side in ['blue', 'red']:
            squadron = utils.get_squadron(self.node, squadron_id=match[f'squadron_{side}'])
            if not squadron['role']:
                raise ValueError(f"Squadron {squadron['name']} does not have a role set.")
            role = self.bot.get_role(squadron['role'])
            if not role:
                raise ValueError(f"Squadron {squadron['name']} has an invalid role.")
            try:
                channel = await self.get_squadron_channel(match_id, side)
                if channel:
                    await channel.edit(name=channel.name + " (closed)")
                await channel.set_permissions(role, overwrite=None)
            except discord.Forbidden:
                raise PermissionError("You need to give the bot the Manage Roles permission!")

    @match.command(description=_('Edit a match'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @app_commands.rename(match_id="match")
    @app_commands.autocomplete(match_id=all_matches_autocomplete)
    @app_commands.autocomplete(winner_squadron_id=match_squadrons_autocomplete)
    @utils.app_has_role('GameMaster')
    async def edit(self, interaction: discord.Interaction, tournament_id: int, match_id: int,
                   winner_squadron_id: Optional[int] = None):
        match = await self.get_match(match_id)
        modal = utils.ConfigModal(title="Match Results", config={
            "squadron_blue_rounds_won": {
                "type": int,
                "label": _("Squadron blue rounds won"),
                "required": True
            },
            "squadron_red_rounds_won": {
                "type": int,
                "label": _("Squadron red rounds won"),
                "required": True
            },
            "match_time": {
                "type": datetime,
                "label": _("Match time (yyyy-mm-dd hh:mm:ss)"),
                "required": False
            }
        }, old_values=match)
        # noinspection PyUnresolvedReferences
        await interaction.response.send_modal(modal)
        if await modal.wait():
            await interaction.followup.send(_("Aborted."), ephemeral=True)
            return

        if modal.value.get('squadron_blue_rounds_won') > modal.value.get('squadron_red_rounds_won'):
            winner_squadron_id = match['squadron_blue']
        elif modal.value.get('squadron_red_rounds_won') > modal.value.get('squadron_blue_rounds_won'):
            winner_squadron_id = match['squadron_red']

        match_time = modal.value.get('match_time')
        try:
            match_time = datetime.strptime(match_time, "%Y-%m-%d %H:%M:%S") if match_time else None
        except ValueError:
            await interaction.followup.send(_("Invalid date format. Use yyyy-mm-dd hh:mm:ss."), ephemeral=True)
            return
        async with self.apool.connection() as conn:
            async with conn.transaction():
                await conn.execute("""
                    UPDATE tm_matches 
                    SET squadron_blue_rounds_won = %s, squadron_red_rounds_won = %s, match_time = %s 
                    WHERE match_id = %s
                """, (modal.value.get('squadron_blue_rounds_won'),
                      modal.value.get('squadron_red_rounds_won'),
                      match_time, match_id))

                if winner_squadron_id:
                    squadron = utils.get_squadron(self.node, squadron_id=winner_squadron_id)
                    if await utils.yn_question(
                            interaction, _("Do you really want to finish the match and make {} the winner?").format(
                                squadron['name']), ephemeral=True):
                        cursor = await conn.execute("""
                            UPDATE tm_matches SET winner_squadron_id = %s WHERE match_id = %s
                            RETURNING server_name
                        """, (winner_squadron_id, match_id))
                        server_name = (await cursor.fetchone())[0]
                        server = self.bot.servers[server_name]
                        self.reset_serversettings(server)
                    else:
                        await interaction.followup.send(_("Winner not updated."), ephemeral=True)

        if match['winner_squadron_id'] != winner_squadron_id \
            or match['squadron_blue_rounds_won'] != modal.value.get('squadron_blue_rounds_won') \
            or match['squadron_red_rounds_won'] != modal.value.get('squadron_red_rounds_won')\
            or match['match_time'] != match_time:
            await self.bot.audit(f"updated match {match_id} for tournament {tournament_id}.",
                                 user=interaction.user)
            await interaction.followup.send(_("Match updated."), ephemeral=True)
        else:
            await interaction.followup.send(_("Match not updated."), ephemeral=True)

        if await self.eventlistener.check_tournament_finished(tournament_id):
            await interaction.followup.send(_("You just finished the tournament!"), ephemeral=True)
        elif winner_squadron_id:
            await self.render_info_embed(tournament_id, phase=TOURNAMENT_PHASE.MATCH_FINISHED, match_id=match_id)
        await self.render_status_embed(tournament_id)

    @match.command(description=_('Customize the next round'))
    @app_commands.guild_only()
    async def customize(self, interaction: discord.Interaction):
        ephemeral = utils.get_ephemeral(interaction)
        # noinspection PyUnresolvedReferences
        await interaction.response.defer(ephemeral=ephemeral)
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                for coalition in ['blue', 'red']:
                    await cursor.execute(f"""
                        SELECT m.tournament_id, m.match_id, 
                               squadron_{coalition} AS squadron_id, choices_{coalition}_ack AS ack, 
                               server_name 
                        FROM tm_matches m 
                        JOIN tm_tournaments t ON m.tournament_id = t.tournament_id 
                        JOIN campaigns c ON t.campaign = c.name
                        WHERE c.start <= NOW() AT TIME ZONE 'UTC' AND squadron_{coalition}_channel = %s
                        AND COALESCE(c.stop, NOW() AT TIME ZONE 'UTC') >= NOW() AT TIME ZONE 'UTC'
                        AND m.round_number > 0 and m.winner_squadron_id IS NULL
                    """, (interaction.channel.id, ))
                    row = await cursor.fetchone()
                    if not row:
                        continue
                    if row['ack']:
                        await interaction.followup.send(_("You already made your choice. Wait for the next round!"),
                                                        ephemeral=True)
                        return
                    tournament_id = row['tournament_id']
                    match_id = row['match_id']
                    squadron_id = row['squadron_id']
                    break
                else:
                    await interaction.followup.send(_("{} has to be used in the respective coalition channel.").format(
                        (await utils.get_command(self.bot, group=self.match.name, name=self.customize.name)).mention),
                        ephemeral=True)
                    return

        # check if a match is running
        server = self.bus.servers[row['server_name']]
        if server.status == Status.RUNNING:
            await interaction.followup.send(_("You can not choose during a running match!"), ephemeral=True)
            return

        admins = utils.get_squadron_admins(self.node, squadron_id)
        if interaction.user.id not in admins and not utils.check_roles(self.bot.roles['GameMaster'], interaction.user):
            await interaction.followup.send(
                f"You need to be an admin of the squadron {squadron_id} or a Game Master to use this command.",
                ephemeral=True
            )
            return

        view = ChoicesView(self, tournament_id=tournament_id, match_id=match_id, squadron_id=squadron_id,
                           config=self.get_config(server))
        embed = await view.render()
        # noinspection PyUnresolvedReferences
        if not view.children[0].options:
            await interaction.followup.send(_("You do not have enough squadron credits to buy a choice."),
                                            ephemeral=True)
            return
        msg = await interaction.followup.send(view=view, embed=embed, ephemeral=ephemeral)
        try:
            if await view.wait():
                return

            if view.acknowledged is True:
                embed = await view.render()
                embed.description = None
                embed.set_footer(text=None)

            if view.acknowledged is None or (view.acknowledged is True and not await utils.yn_question(
                    interaction, _("Are you sure?\nYour settings will be directly applied to the next round."),
                    embed=embed, ephemeral=True)):
                await interaction.followup.send(
                    _("Your choices were saved.\n"
                      "If you want them to be applied to the next round, press 'Confirm & Buy'."))
                return

            async with self.apool.connection() as conn:
                async with conn.transaction():
                    await conn.execute("""
                        UPDATE tm_matches
                        SET 
                            choices_blue_ack = CASE
                                WHEN squadron_blue = %(squadron_id)s THEN true
                                ELSE choices_blue_ack
                            END,
                            choices_red_ack  = CASE
                                WHEN squadron_red = %(squadron_id)s THEN true
                                ELSE choices_red_ack
                            END
                        WHERE 
                            (squadron_blue = %(squadron_id)s OR squadron_red = %(squadron_id)s)
                            AND match_id = %(match_id)s
                    """, {"match_id": match_id, "squadron_id": squadron_id})
            if not view.acknowledged:
                await interaction.followup.send(_("You decided to not buy any customizations in this round."))
            else:
                embed.title = _("Invoice")
                embed.set_footer(text=_("Thank you for your purchase!"))
                await interaction.followup.send(embed=embed)
        finally:
            try:
                await msg.delete()
            except discord.NotFound:
                pass

    # New command group "/tickets"
    tickets = Group(name="tickets", description=_("Commands to manage tickets in a tournament"))

    @tickets.command(name="list", description=_('List tickets of a squadron'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @app_commands.rename(squadron_id="squadron")
    @app_commands.autocomplete(squadron_id=valid_squadron_autocomplete)
    @utils.squadron_role_check()
    async def _list(self, interaction: discord.Interaction, tournament_id: int, squadron_id: int):
        embed = discord.Embed(colour=discord.Colour.blue(), title=_("Your Tickets"))
        ticket_names = []
        ticket_counts = []
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                async for row in await cursor.execute(
                        "SELECT * FROM tm_tickets WHERE tournament_id = %s AND squadron_id = %s",
                        (tournament_id, squadron_id)):
                    ticket_names.append(row['ticket_name'])
                    ticket_counts.append(row['ticket_count'])
        if not ticket_names:
            # noinspection PyUnresolvedReferences
            await interaction.response.send_message(_("Your squadron does not have any tickets."), ephemeral=True)
            return

        embed.add_field(name=_("Tickets"),
                        value="\n".join([f"{y} x {x}" for x, y in zip(ticket_names, ticket_counts)]))
        # noinspection PyUnresolvedReferences
        await interaction.response.send_message(embed=embed, ephemeral=True)

    @tickets.command(name="sell", description=_('Sell tickets for credits'))
    @app_commands.guild_only()
    @app_commands.rename(tournament_id="tournament")
    @app_commands.autocomplete(tournament_id=active_tournament_autocomplete)
    @app_commands.rename(squadron_id="squadron")
    @app_commands.autocomplete(squadron_id=valid_squadron_autocomplete)
    @app_commands.autocomplete(ticket_name=tickets_autocomplete)
    @utils.squadron_role_check()
    async def sell(self, interaction: discord.Interaction, tournament_id: int, squadron_id: int, ticket_name: str):
        squadron = await self.get_squadron(tournament_id, squadron_id)
        ticket = self.get_config().get('presets', {}).get('tickets', {}).get(ticket_name)
        if not ticket:
            # noinspection PyUnresolvedReferences
            await interaction.response.send_message(_("The ticket {} does not exist.").format(ticket_name),
                                                    ephemeral=True)
            return

        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                my_tickets = {}
                async for row in await cursor.execute("""
                    SELECT ticket_name, ticket_count
                    FROM tm_tickets
                    WHERE tournament_id = %s AND squadron_id = %s AND ticket_count > 0
                """, (tournament_id, squadron_id)):
                    my_tickets[row['ticket_name']] = row['ticket_count']

                if not my_tickets.get(ticket_name):
                    # noinspection PyUnresolvedReferences
                    await interaction.response.send_message(_("You do not have any tickets of this type."),
                                                            ephemeral=True)
                    return

                credits = ticket.get('credits', 0)
                if credits == 0:
                    # noinspection PyUnresolvedReferences
                    await interaction.response.send_message(
                        _("Tickets of type {} are not for sale.").format(ticket_name), ephemeral=True)
                    return

                if not await utils.yn_question(
                        interaction, _("Do you really want to sell a tickets of type {} for {} credits?").format(
                            ticket_name, credits)):
                    await interaction.followup.send(_("Aborted"))
                    return

                async with conn.transaction():
                    cursor = await cursor.execute("""
                        UPDATE tm_tickets SET ticket_count = ticket_count - 1
                        WHERE tournament_id = %s 
                          AND squadron_id = %s
                          AND ticket_name = %s
                        RETURNING ticket_count
                    """, (tournament_id, squadron_id, ticket_name))
                    ticket_count = (await cursor.fetchone())['ticket_count']
                    if ticket_count == 0:
                        await conn.execute("""
                            DELETE FROM tm_tickets WHERE tournament_id = %s AND squadron_id = %s AND ticket_name = %s
                        """, (tournament_id, squadron_id, ticket_name))
                    squadron.points += credits
                    squadron.audit(event='Ticket sale', points=credits, remark=f"{ticket_name} sold")
            # noinspection PyUnresolvedReferences
            await interaction.followup.send(
                _("You sold a ticket of type {} and got {} credits back.\n"
                  "Your total squadron balance is {} credits.").format(ticket_name, credits, squadron.points),
                ephemeral=True)

    @tasks.loop(minutes=1.0)
    async def match_scheduler(self):
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                # for all active tournaments
                await cursor.execute("""
                    SELECT t.tournament_id, t.campaign AS name
                    FROM tm_tournaments t JOIN campaigns c ON t.campaign = c.name 
                    WHERE c.start <= NOW() AT TIME ZONE 'UTC'
                    AND COALESCE(c.stop, NOW() AT TIME ZONE 'UTC') >= NOW() AT TIME ZONE 'UTC'
                    ORDER BY campaign
                """)
                tournaments = await cursor.fetchall()
                for tournament in tournaments:
                    # warn squadrons prior to their matches
                    await cursor.execute("""
                        SELECT match_id, match_time, squadron_blue, squadron_red
                        FROM tm_matches m 
                        WHERE tournament_id = %s
                        AND round_number = 0
                        AND winner_squadron_id IS NULL
                        AND match_time IS NOT NULL
                        AND 
                        (
                            DATE_TRUNC('minute', match_time) = DATE_TRUNC('minute', NOW() AT TIME ZONE 'UTC' - interval '24 hours')
                        ) OR (
                            DATE_TRUNC('minute', match_time) = DATE_TRUNC('minute', NOW() AT TIME ZONE 'UTC' - interval '1 hour')
                        )
                    """, (tournament['tournament_id'], ))
                    for match in await cursor.fetchall():
                        match_time = int(match['match_time'].replace(tzinfo=timezone.utc).timestamp())
                        squadron_blue = utils.get_squadron(self.node, squadron_id=match['squadron_blue'])
                        squadron_red = utils.get_squadron(self.node, squadron_id=match['squadron_red'])
                        await self.inform_squadron(
                            tournament_id=tournament['tournament_id'],
                            squadron_id=match['squadron_blue'],
                            message=_("Your next match against {name} starts <t:{time}:R>!").format(
                                name=squadron_red['name'], time=match_time))
                        await self.inform_squadron(
                            tournament_id=tournament['tournament_id'],
                            squadron_id=match['squadron_red'],
                            message=_("Your next match against {name} starts <t:{time}:R>!").format(
                                name=squadron_blue['name'], time=match_time))

                    # start scheduled matches if there is no one running already
                    await cursor.execute("""
                        SELECT match_id, server_name, squadron_blue, squadron_red
                        FROM tm_matches m 
                        WHERE tournament_id = %s
                        AND round_number = 0
                        AND winner_squadron_id IS NULL
                        AND match_time IS NOT NULL
                        AND match_time < NOW() AT TIME ZONE 'UTC'
                        AND server_name NOT IN (
                            SELECT server_name 
                            FROM tm_matches
                            WHERE tournament_id = m.tournament_id
                            AND round_number > 0
                            AND winner_squadron_id IS NULL
                        )
                    """, (tournament['tournament_id'], ))
                    server_names = []
                    for match in await cursor.fetchall():
                        # we can only start ONE match per server at a time
                        if match['server_name'] in server_names:
                            self.log.debug("Match Schedler: match found, but there is one running already!")
                            continue
                        server_names.append(match['server_name'])
                        server = self.bus.servers[match['server_name']]
                        # we must not start a match if the server is (still?) running
                        if server.status == Status.RUNNING:
                            self.log.debug("Match Scheduler: match found, but the server is running already!")
                            continue

                        # start the match
                        try:
                            await self.start_match(server,
                                                   tournament_id=tournament['tournament_id'],
                                                   match_id=match['match_id'])
                        except ValueError as ex:
                            self.log.warning(ex)
                            return

                        # audit event
                        squadron_blue = utils.get_squadron(self.node, squadron_id=match['squadron_blue'])
                        squadron_red = utils.get_squadron(self.node, squadron_id=match['squadron_red'])
                        await self.bot.audit(f"Scheduler started a match for tournament {tournament['name']} "
                                             f"between squadrons {squadron_blue['name']} and {squadron_red['name']}.")

    @match_scheduler.before_loop
    async def before_match_scheduler(self):
        await self.bot.wait_until_ready()

    async def import_tournament_data(self, buffer: BytesIO, tournament_id: int) -> pd.DataFrame:
        """
        Import tournament data from Excel buffer and convert squadron names back to IDs

        Args:
            buffer: BytesIO object containing the Excel file
            tournament_id: The tournament ID to verify data against

        Returns:
            DataFrame with the processed matches data
        """
        # Read all sheets from the Excel file
        xlsx = pd.ExcelFile(buffer)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            matches_df = pd.read_excel(xlsx, 'Matches')
            squadrons_df = pd.read_excel(xlsx, 'Squadrons')

        # Create reverse mapping from squadron name to ID
        name_squadron_map = dict(zip(squadrons_df['name'], squadrons_df['squadron_id']))

        # List of columns that contain squadron names that need to be converted back to IDs
        squadron_columns = ['squadron_red', 'squadron_blue', 'winner_squadron_id']

        # Sort the DataFrame by match_id to ensure consistent processing
        matches_df = matches_df.sort_values('match_id').reset_index(drop=True)

        # Convert squadron names back to IDs
        for column in squadron_columns:
            if column in matches_df.columns:
                if column == 'winner_squadron_id':
                    # For winner_squadron_id, handle NaN values specially
                    matches_df[column] = matches_df[column].map(
                        lambda x: name_squadron_map.get(x) if pd.notna(x) else None)
                else:
                    # For required columns (squadron_red, squadron_blue), use regular mapping
                    matches_df[column] = matches_df[column].map(name_squadron_map)

                # Check if any required mappings failed (would result in NaN)
                if column != 'winner_squadron_id' and matches_df[column].isna().any():
                    invalid_names = matches_df[matches_df[column].isna()][column].unique()
                    raise ValueError(f"Invalid squadron names found in {column}: {invalid_names}")

        # Verify the server names exist
        if 'server_name' in matches_df.columns:
            servers_df = pd.read_excel(xlsx, 'Servers')
            valid_servers = set(servers_df['server_name'].unique())
            invalid_servers = set(matches_df['server_name'].unique()) - valid_servers
            if invalid_servers:
                raise ValueError(f"Invalid server names found: {invalid_servers}")

        # Convert match_time back to datetime if it exists
        if 'match_time' in matches_df.columns:
            matches_df['match_time'] = pd.to_datetime(matches_df['match_time'])

        return matches_df

    async def import_tournament_data_to_db(self, match_df: pd.DataFrame, tournament_id: int):
        """
        Update tournament matches in the database based on the provided DataFrame.

        Args:
            match_df: DataFrame containing the match data
            tournament_id: ID of the tournament

        Raises:
            ValueError: If update fails or invalid data is encountered
        """
        # Get the column names from the DataFrame that we want to update
        updateable_columns = [
            'squadron_red', 'squadron_blue', 'round_number',
            'squadron_red_rounds_won', 'squadron_blue_rounds_won',
            'winner_squadron_id', 'server_name', 'match_time'
        ]

        # Filter to only columns that exist in the DataFrame
        columns_to_update = [col for col in updateable_columns if col in match_df.columns]

        if not columns_to_update:
            raise ValueError("No valid columns to update found in the DataFrame")

        # Construct the SQL UPDATE query dynamically based on available columns
        set_clause = ", ".join(f"{col} = %s" for col in columns_to_update)
        query = f"""
            UPDATE tm_matches 
            SET {set_clause}
            WHERE match_id = %s AND tournament_id = %s
        """

        try:
            async with self.apool.connection() as conn:
                async with conn.transaction():
                    # Process each row in the DataFrame
                    for _, row in match_df.iterrows():
                        # Extract values for the update in the correct order
                        update_values = []
                        for col in columns_to_update:
                            if col == 'match_time':
                                # Handle NaT values for datetime column
                                value = None if pd.isna(row[col]) else row[col]
                            else:
                                value = row[col]
                            update_values.append(value)

                        # Add match_id and tournament_id for the WHERE clause
                        update_values.extend([row['match_id'], tournament_id])

                        await conn.execute(query, tuple(update_values))

        except Exception as e:
            raise ValueError(f"Failed to update tournament data: {str(e)}")

    @commands.Cog.listener()
    async def on_message(self, message: discord.Message):
        if message.author.bot or not message.attachments:
            return

        att = message.attachments[0]
        filename = att.filename.lower()
        filetype = filename.lower().split('.')[-1]
        if not filename.startswith('tournament') or not filetype.startswith('xls'):
            return

        if not utils.check_roles(['DCS Admin', 'GameMaster'], message.author):
            await message.channel.send(_("You need to be DCS Admin or GameMaster to upload data."))
            return

        match = re.match(r'^tournament_(\d+)\.xlsx?$', filename)
        if not match:
            await message.channel.send(_("The filename has to be 'tournament_ID.xlsx'."))
            return

        tournament_id = int(match.group(1))
        tournament = await self.get_tournament(tournament_id)
        if not tournament:
            await message.channel.send(_("Tournament not found."))
            return

        ctx = await self.bot.get_context(message)
        if await self.eventlistener.is_tournament_finished(tournament_id):
            msg = _("This tournament is already finished!\n")
        else:
            msg = ""
        if not await utils.yn_question(ctx, msg + _("Do you want to import the data and overwrite your matches?")):
            await message.channel.send(_("Aborted."))
            return

        # read the file
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(att.url, proxy=self.node.proxy, proxy_auth=self.node.proxy_auth) as response:
                    response.raise_for_status()
                    match_df = await self.import_tournament_data(BytesIO(await response.read()), tournament_id)
                    await self.import_tournament_data_to_db(match_df, tournament_id)
                    embed = await self.render_matches(tournament)
                    embed.set_footer(text=_("The data has been imported successfully."))
                    await message.channel.send(embed=embed)
                    # check if the tournament has finished
                    if await self.eventlistener.check_tournament_finished(tournament_id):
                        await message.channel.send("The upload finished the tournament.")
                    await self.render_status_embed(tournament_id)
        except Exception as ex:
            self.log.exception(ex)
            await message.channel.send(_("Error while processing the file: {}").format(ex))


async def setup(bot: DCSServerBot):
    if 'competitive' not in bot.plugins:
        raise PluginRequiredError('competitive')

    await bot.add_cog(Tournament(bot, TournamentEventListener))
