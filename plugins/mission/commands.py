import aiohttp
import aiofiles
import asyncio
import discord
import importlib
import os
import pandas as pd
import psycopg
import random
import re
import traceback
import warnings

from contextlib import suppress
from core import utils, Plugin, Report, Status, Server, Coalition, Channel, Player, PluginRequiredError, MizFile, \
    Group, ReportEnv, command, PlayerType, DataObjectFactory, Member, DEFAULT_TAG, get_translation, \
    UnsupportedMizFileException, cache_with_expiration
from datetime import datetime, timezone
from discord import Interaction, app_commands, SelectOption
from discord.app_commands import Range, describe
from discord.ext import commands, tasks
from io import BytesIO
from openpyxl.utils import get_column_letter
from pathlib import Path
from psycopg.rows import dict_row
from services.bot import DCSServerBot
from typing import Literal, Type

from .airbase import Info
from .const import LIQUIDS
from .listener import MissionEventListener
from .upload import MissionUploadHandler
from .views import ServerView, PresetView, InfoView, ModifyView, AirbaseView, BanModal
from ..userstats.filter import PeriodFilter

# ruamel YAML support
from ruamel.yaml import YAML
yaml = YAML()

_ = get_translation(__name__.split('.')[1])

SHEET_TITLES = {
    "aircraft": "Aircraft",
    "weapon": "Weapons",
    "liquids": "Liquids",
}
REVERSE_LIQUIDS = {v: k for k, v in LIQUIDS.items()}


async def mizfile_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    try:
        server: Server = await utils.ServerTransformer().transform(interaction, interaction.namespace.server)
        if not server:
            return []
        base_dir = await server.get_missions_dir()
        ignore = ['.dcssb']
        if server.locals.get('ignore_dirs'):
            ignore.extend(server.locals['ignore_dirs'])
        installed_missions = [os.path.expandvars(x) for x in await utils.get_cached_mission_list(server)]
        exp_base, file_list = await server.node.list_directory(base_dir, pattern=['*.miz', '*.sav'], traverse=True,
                                                               ignore=ignore)
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=os.path.relpath(x, exp_base)[:-4], value=os.path.relpath(x, exp_base))
            for x in file_list
            if x not in installed_missions and os.path.join(os.path.dirname(x), '.dcssb', os.path.basename(
                x)) not in installed_missions and current.casefold() in os.path.relpath(x, base_dir).casefold()
        ]
        return choices[:25]
    except Exception as ex:
        interaction.client.log.exception(ex)
        return []

async def orig_mission_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    try:
        server: Server = await utils.ServerTransformer().transform(interaction, interaction.namespace.server)
        if not server:
            return []
        _, file_list = await server.node.list_directory(await server.get_missions_dir(), pattern='*.orig',
                                                        traverse=True)
        orig_files = [os.path.basename(x)[:-9] for x in file_list]
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=os.path.basename(x)[:-4], value=idx)
            for idx, x in enumerate(await utils.get_cached_mission_list(server))
            if os.path.basename(x)[:-4] in orig_files and (not current or current.casefold() in x[:-4].casefold())
        ]
        return choices[:25]
    except Exception as ex:
        interaction.client.log.exception(ex)
        return []


async def presets_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    try:
        choices: list[app_commands.Choice[str]] = [
            app_commands.Choice(name=x.name[:-5], value=str(x))
            for x in Path(interaction.client.node.config_dir).glob('presets*.yaml')
            if not current or current.casefold() in x.name[:-5].casefold()
        ]
        return choices[:25]
    except Exception as ex:
        interaction.client.log.exception(ex)
        return []


async def nosav_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[int]]:
    """
    Autocompletion of mission names from the current mission list of a server that has to be provided as an earlier
    parameter to the application command. The mission list can only be obtained by people with the DCS Admin role.
    """
    def get_name(base_dir: str, path: str):
        try:
            return os.path.relpath(path, base_dir).replace('.dcssb' + os.path.sep, '')[:-4]
        except ValueError:
            return os.path.basename(path)[:-4]

    if not await interaction.command._check_can_run(interaction):
        return []
    try:
        server: Server = await utils.ServerTransformer().transform(interaction, interaction.namespace.server)
        if not server:
            return []
        base_dir = await server.get_missions_dir()
        choices: list[app_commands.Choice[int]] = [
            app_commands.Choice(name=get_name(base_dir, x), value=idx)
            for idx, x in enumerate(await utils.get_cached_mission_list(server))
            if not x.endswith('.sav') and (not current or current.casefold() in get_name(base_dir, x).casefold())
        ]
        return sorted(choices, key=lambda choice: choice.name)[:25]
    except Exception as ex:
        interaction.client.log.exception(ex)
        return []

@cache_with_expiration(180)
async def get_airbase(server: Server, name: str) -> dict:
    return await server.send_to_dcs_sync({"command": "getAirbase", "name": name}, timeout=60)

async def wh_category_autocomplete(interaction: discord.Interaction, _current: str) -> list[app_commands.Choice[str]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    try:
        server: Server = await utils.ServerTransformer().transform(interaction, interaction.namespace.server)
        idx = interaction.namespace.airbase
        airbase: dict = server.current_mission.airbases[idx]
        data = await get_airbase(server, airbase['name'])
        return [
            app_commands.Choice(name=x.title(), value=x)
            for x in sorted(server.resources.keys())
            if not data['unlimited'][x]
        ]
    except Exception as ex:
        interaction.client.log.exception(ex)
        return []


async def wh_item_autocomplete(interaction: discord.Interaction, current: str) -> list[app_commands.Choice[str]]:
    if not await interaction.command._check_can_run(interaction):
        return []
    server: Server = await utils.ServerTransformer().transform(interaction, interaction.namespace.server)
    category = interaction.namespace.category
    try:
        choices: list[app_commands.Choice[str]] = [
            app_commands.Choice(name=x['name'], value=x['wstype'])
            for x in sorted(server.resources.get(category, {}), key=lambda x: x['name'])
            if not current or current.casefold() in x['name'].casefold()
        ]
        return choices[:25]
    except Exception as ex:
        interaction.client.log.exception(ex)
        return []


class DDTransformer(app_commands.Transformer):
    async def transform(self, interaction: discord.Interaction, value: str | float) -> float:
        try:
            return float(value)
        except ValueError:
            return utils.dms_to_dd(value)


class Mission(Plugin[MissionEventListener]):

    def __init__(self, bot: DCSServerBot, listener: Type[MissionEventListener] = None):
        super().__init__(bot, listener)
        self.lock = asyncio.Lock()

    async def cog_load(self) -> None:
        await super().cog_load()
        self.update_channel_name.add_exception_type(AttributeError)
        self.update_channel_name.start()
        self.afk_check.start()
        self.check_for_unban.add_exception_type(psycopg.DatabaseError)
        self.check_for_unban.start()
        self.expire_token.add_exception_type(psycopg.DatabaseError)
        self.expire_token.start()
        if self.bot.locals.get('autorole', {}):
            self.check_roles.add_exception_type(psycopg.DatabaseError)
            self.check_roles.add_exception_type(discord.errors.DiscordException)
            self.check_roles.start()

    async def cog_unload(self):
        if self.bot.locals.get('autorole', {}):
            self.check_roles.stop()
        self.expire_token.cancel()
        self.check_for_unban.cancel()
        self.afk_check.cancel()
        self.update_channel_name.cancel()
        await super().cog_unload()

    async def migrate(self, new_version: str, conn: psycopg.AsyncConnection | None = None) -> None:
        function_name = f"migrate_{new_version.replace('.', '_')}"
        migrate_module = importlib.import_module('.migrate', package=__package__)
        migrate_function = getattr(migrate_module, function_name, None)
        if callable(migrate_function):
            if asyncio.iscoroutinefunction(migrate_function):
                await migrate_function(self)
            else:
                migrate_function(self)
            self.locals = self.read_locals()

    async def prune(self, conn: psycopg.AsyncConnection, days: int) -> None:
        self.log.debug('Pruning Mission ...')
        await conn.execute(f"""
            DELETE FROM missions WHERE mission_end < (DATE(now() AT TIME ZONE 'utc') - %s::interval)
        """, (f'{days} days', ))
        self.log.debug('Mission pruned.')

    async def update_ucid(self, conn: psycopg.AsyncConnection, old_ucid: str, new_ucid: str) -> None:
        # check if the new ucid was banned already
        cursor = await conn.execute("""
            SELECT banned_by, reason, banned_at, banned_until 
            FROM bans WHERE ucid = %s
            AND banned_until > (NOW() AT TIME ZONE 'UTC')
        """,(new_ucid, ))
        if cursor.rowcount == 1:
            row = await cursor.fetchone()
            # if yes, create a ban for the old ucid also
            await conn.execute("""
                INSERT INTO bans (ucid, banned_by, reason, banned_at, banned_until)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (ucid) DO NOTHING
            """, (old_ucid, row[0], row[1], row[2], row[3]))
        else:
            # otherwise create a new ban if the old ucid was banned already
            await conn.execute(f"""
                INSERT INTO bans (ucid, banned_by, reason, banned_at, banned_until) 
                SELECT %(new_ucid)s, banned_by, reason, banned_at, banned_until FROM bans WHERE ucid = %(old_ucid)s
                ON CONFLICT (ucid) DO NOTHING
            """, {"new_ucid": new_ucid, "old_ucid": old_ucid})

    # New command group "/mission"
    mission = Group(name="mission", description=_("Commands to manage a DCS mission"))

    @mission.command(name="info", description=_('Info about the running mission'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS')
    async def mission_info(self, interaction: Interaction, server: app_commands.Transform[Server, utils.ServerTransformer]):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        report = Report(self.bot, self.plugin_name, 'serverStatus.json')
        env: ReportEnv = await report.render(server=server)
        try:
            file = discord.File(fp=env.buffer, filename=env.filename) if env.filename else discord.utils.MISSING
            await interaction.followup.send(embed=env.embed, file=file, ephemeral=ephemeral)
        finally:
            if env.buffer:
                env.buffer.close()

    @mission.command(description=_('Manage the active mission'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def manage(self, interaction: Interaction, server: app_commands.Transform[Server, utils.ServerTransformer(
                       status=[Status.RUNNING, Status.PAUSED, Status.STOPPED])]):
        view = ServerView(server)
        embed = await view.render(interaction)
        await interaction.response.send_message(embed=embed, view=view, ephemeral=utils.get_ephemeral(interaction))
        try:
            await view.wait()
        finally:
            await interaction.delete_original_response()

    @mission.command(description=_('Shows briefing of the active mission'))
    @utils.app_has_role('DCS')
    @app_commands.guild_only()
    async def briefing(self, interaction: discord.Interaction,
                       server: app_commands.Transform[Server, utils.ServerTransformer(
                           status=[Status.RUNNING, Status.PAUSED])]):
        async def read_passwords() -> dict:
            async with self.apool.connection() as conn:
                cursor = await conn.execute('SELECT blue_password, red_password FROM servers WHERE server_name = %s',
                                            (server.name,))
                row = await cursor.fetchone()
                return {"blue": row[0], "red": row[1]}

        if server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(server.display_name),
                                                    ephemeral=True)
            return
        await interaction.response.defer()
        mission_info = await server.send_to_dcs_sync({
            "command": "getMissionDetails"
        }, timeout=60)
        mission_info['passwords'] = await read_passwords()
        report = Report(self.bot, self.plugin_name, 'briefing.json')
        env = await report.render(mission_info=mission_info, server_name=server.name, interaction=interaction)
        msg = await interaction.original_response()
        await msg.edit(embed=env.embed, delete_after=self.bot.locals.get('message_autodelete'))

    @mission.command(description=_('Restarts the current active mission\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def restart(self, interaction: discord.Interaction,
                      server: app_commands.Transform[Server, utils.ServerTransformer(
                          status=[Status.RUNNING, Status.PAUSED, Status.STOPPED])],
                      delay: int | None = 120, reason: str | None = None, run_extensions: bool | None = True,
                      use_orig: bool | None = True):
        await self._restart(interaction, server, delay, reason, run_extensions, use_orig, rotate=False)

    @mission.command(description=_('Rotates to the next mission\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def rotate(self, interaction: discord.Interaction,
                     server: app_commands.Transform[Server, utils.ServerTransformer(
                          status=[Status.RUNNING, Status.PAUSED, Status.STOPPED])],
                     delay: int | None = 120, reason: str | None = None, run_extensions: bool | None = True,
                     use_orig: bool | None = True):
        await self._restart(interaction, server, delay, reason, run_extensions, use_orig, rotate=True)

    async def _restart(self, interaction: discord.Interaction,
                       server: app_commands.Transform[Server, utils.ServerTransformer(
                          status=[Status.RUNNING, Status.PAUSED, Status.STOPPED])],
                       delay: int | None = 120, reason: str | None = None, run_extensions: bool | None = True,
                       use_orig: bool | None = True, rotate: bool | None = False):
        what = "restart" if not rotate else "rotate"
        actions = {
            "restart": "restarted",
            "rotate": "rotated",
        }
        ephemeral = utils.get_ephemeral(interaction)
        if server.status not in [Status.RUNNING, Status.PAUSED, Status.STOPPED]:
            await interaction.response.send_message(
                _("Can't restart server {server} as it is {status}!").format(server=server.display_name,
                                                                             status=server.status.name), ephemeral=True)
            return
        if server.restart_pending and not await utils.yn_question(
                interaction, _('A restart is currently pending.\n'
                               'Would you still like to {} the mission?').format(_(what)),
                ephemeral=ephemeral):
            return
        else:
            server.on_empty = dict()
        if server.is_populated():
            result = await utils.populated_question(
                interaction, _("Do you really want to {} the mission?").format(_(what)), ephemeral=ephemeral)
            if not result:
                return
            elif result == 'later':
                server.on_empty = {
                    "method": what,
                    "run_extensions": run_extensions,
                    "use_orig": use_orig,
                    "user": interaction.user
                }
                server.restart_pending = True
                await interaction.followup.send(_('Mission will {}, when server is empty.').format(_(what)),
                                                ephemeral=ephemeral)
                return

        server.restart_pending = True
        if not interaction.response.is_done():
            await interaction.response.defer(ephemeral=ephemeral)
        if server.is_populated():
            if delay > 0:
                message = _("!!! Mission will be {what} in {when}!!!").format(what=_(actions.get(what)),
                                                                              when=utils.format_time(delay))
            else:
                message = _("!!! Mission will be {} NOW !!!").format(_(actions.get(what)))
            # have we got a message to present to the users?
            if reason:
                message += _(' Reason: {}').format(reason)

            msg = await interaction.followup.send(
                _('Mission will be {what} in {when} (warning users before)...').format(what=_(actions.get(what)),
                                                                                       when=utils.format_time(delay)),
                ephemeral=ephemeral)
            await server.sendPopupMessage(Coalition.ALL, message, sender=interaction.user.display_name)
            await asyncio.sleep(delay)
            await msg.delete()
        try:
            msg = await interaction.followup.send(_('Mission will {} now, please wait ...').format(_(what)),
                                                  ephemeral=ephemeral)
            if not server.locals.get('mission_rewrite', True) and server.status != Status.STOPPED:
                await server.stop()
            if rotate:
                await server.loadNextMission(modify_mission=run_extensions, use_orig=use_orig)
            else:
                await server.restart(modify_mission=run_extensions)
            await self.bot.audit(f'{actions.get(what)} mission', server=server, user=interaction.user)
            await msg.delete()
            await interaction.followup.send(_("Mission {}.").format(_(actions.get(what))), ephemeral=ephemeral)
        except (TimeoutError, asyncio.TimeoutError):
            await interaction.followup.send(
                _("Timeout while the mission {what}.\n"
                  "Please check with {command}, if the mission is running.").format(
                    what=_(actions.get(what)),
                    command=(await utils.get_command(self.bot, group=self.mission.name, name=self.info.name)).mention
                ), ephemeral=ephemeral)

    async def _load(self, interaction: discord.Interaction, server: Server, mission: int | str | None = None,
                    run_extensions: bool | None = False, use_orig: bool | None = True):
        ephemeral = utils.get_ephemeral(interaction)
        if server.status not in [Status.RUNNING, Status.PAUSED, Status.STOPPED]:
            await interaction.response.send_message(
                _("Can't load mission on server {server} as it is {status}!").format(
                    server=server.display_name, status=server.status.name), ephemeral=True)
            return
        if server.restart_pending and not await utils.yn_question(
                interaction,
                _('A restart is currently pending.\nWould you still like to {} the mission?').format(_("change")),
                ephemeral=ephemeral
        ):
            return
        else:
            server.on_empty = dict()

        if server.is_populated():
            result = await utils.populated_question(
                interaction,
                _("Do you really want to {} the mission?").format(_("change")),
                ephemeral=ephemeral
            )
            if not result:
                return
        else:
            result = "yes"

        if not interaction.response.is_done():
            await interaction.response.defer(ephemeral=ephemeral)
        if isinstance(mission, int):
            mission_id = mission
            mission = (await server.getMissionList())[mission_id]
        elif isinstance(mission, str):
            try:
                mission = os.path.join(await server.get_missions_dir(), mission)
                mission_id = (await server.getMissionList()).index(mission)
            except ValueError:
                mission_id = None
        else:
            await interaction.followup.send(_('You need to provide a mission!'), ephemeral=True)
            return
        if server.current_mission and mission == server.current_mission.filename:
            if result == 'later':
                server.on_empty = {
                    "method": "restart",
                    "run_extensions": run_extensions,
                    "use_orig": use_orig,
                    "user": interaction.user
                }
                server.restart_pending = True
                await interaction.followup.send(_('Mission will {}, when server is empty.').format(_('restart')),
                                                ephemeral=ephemeral)
            else:
                await server.restart(modify_mission=run_extensions)
                await interaction.followup.send(_('Mission {}.').format(_('restarted')), ephemeral=ephemeral)
        else:
            name = os.path.basename(mission[:-4])
            if mission_id is not None and result == 'later':
                # make sure we load that mission, independently of what happens to the server
                await server.setStartIndex(mission_id + 1)
                server.on_empty = {
                    "method": "load",
                    "mission_id": mission_id + 1,
                    "run_extensions": run_extensions,
                    "use_orig": use_orig,
                    "user": interaction.user
                }
                await interaction.followup.send(
                    _('Mission {} will be loaded when server is empty or on the next restart.').format(name),
                    ephemeral=ephemeral)
            else:
                msg = await interaction.followup.send(_('Loading mission {} ...').format(utils.escape_string(name)),
                                                      ephemeral=ephemeral)
                try:
                    if not server.locals.get('mission_rewrite', True) and server.status != Status.STOPPED:
                        await server.stop()
                    if not await server.loadMission(mission, modify_mission=run_extensions, use_orig=use_orig):
                        await msg.edit(content=_('Mission {} NOT loaded. '
                                                 'Check that you have installed the pre-requisites (terrains, mods).'
                                                 ).format(name))
                    else:
                        message = _('Mission {} loaded.').format(name)
                        if mission_id is None:
                            message += _('\nThis mission is NOT in the mission list and will not auto-load on server '
                                         'or mission restarts.\n'
                                         'If you want it to auto-load, use {}').format(
                                (await utils.get_command(self.bot, group=self.mission.name, name=self.add.name)).mention)
                        await msg.edit(content=message)
                        await self.bot.audit(f"loaded mission {utils.escape_string(name)}", server=server,
                                             user=interaction.user)
                except (TimeoutError, asyncio.TimeoutError):
                    await msg.edit(content=_('Timeout while loading mission {}!').format(name))
                except UnsupportedMizFileException as ex:
                    await msg.edit(content=ex)

    @mission.command(description=_('Loads a mission\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.rename(mission_id="mission")
    @app_commands.describe(use_orig="Change the mission based on the original uploaded mission file.")
    @app_commands.autocomplete(mission_id=utils.mission_autocomplete)
    @app_commands.autocomplete(alt_mission=mizfile_autocomplete)
    async def load(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(
                       status=[Status.STOPPED, Status.RUNNING, Status.PAUSED])],
                   mission_id: int | None = None, alt_mission: str | None = None,
                   run_extensions: bool | None = True, use_orig: bool | None = True):
        await self._load(
            interaction,
            server,
            mission_id if mission_id is not None else alt_mission,
            run_extensions,
            use_orig
        )

    @mission.command(description=_('Adds a mission to the list\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.autocomplete(path=mizfile_autocomplete)
    async def add(self, interaction: discord.Interaction,
                  server: app_commands.Transform[Server, utils.ServerTransformer], path: str,
                  autostart: bool | None = False):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)

        path = os.path.normpath(os.path.join(await server.get_missions_dir(), path))
        new_mission_list = await server.addMission(path, autostart=autostart)
        mission_name = utils.escape_string(os.path.basename(path))
        await interaction.followup.send(_('Mission "{}" added.').format(mission_name), ephemeral=ephemeral)
        mission_id = new_mission_list.index(path)
        if server.status not in [Status.RUNNING, Status.PAUSED, Status.STOPPED] or \
                not await utils.yn_question(interaction, _('Do you want to load this mission?'),
                                            ephemeral=ephemeral):
            return
        await self._load(interaction, server, mission_id, False)

    @mission.command(description=_('Deletes a mission from the list\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.rename(mission_id="mission")
    @app_commands.autocomplete(mission_id=utils.mission_autocomplete)
    async def delete(self, interaction: discord.Interaction,
                     server: app_commands.Transform[Server, utils.ServerTransformer],
                     mission_id: int):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        missions = await server.getMissionList()
        if mission_id >= len(missions):
            await interaction.followup.send(_("No mission found."))
            return
        filename = missions[mission_id]
        if server.status in [Status.RUNNING, Status.PAUSED, Status.STOPPED] and server.current_mission and \
                filename == server.current_mission.filename:
            await interaction.followup.send(_("You can't delete the running mission."), ephemeral=True)
            return
        mission_name = utils.escape_string(os.path.basename(filename[:-4]))

        if await utils.yn_question(interaction,
                                   _('Delete mission "{}" from the mission list?').format(mission_name),
                                   ephemeral=ephemeral):
            try:
                await server.deleteMission(mission_id + 1)
                await interaction.followup.send(_('Mission "{}" removed from list.').format(mission_name),
                                                ephemeral=ephemeral)
                if await utils.yn_question(interaction,
                                           _('Delete "{}" also from disk?').format(mission_name),
                                           ephemeral=ephemeral):
                    try:
                        await server.node.remove_file(filename)
                        if '.dcssb' in filename:
                            secondary = filename
                            primary = filename.replace(os.path.sep + '.dcssb', '')
                            await server.node.remove_file(primary)
                        else:
                            secondary = os.path.join(os.path.dirname(filename), '.dcssb', os.path.basename(filename))
                            await server.node.remove_file(secondary)
                        await server.node.remove_file(secondary + '.orig')
                        await interaction.followup.send(_('Mission "{}" deleted.').format(mission_name),
                                                        ephemeral=ephemeral)
                    except FileNotFoundError:
                        await interaction.followup.send(
                            _('Mission "{}" was already deleted.').format(mission_name),
                            ephemeral=ephemeral)
                await self.bot.audit(_("deleted mission {}").format(os.path.basename(filename[:-4])),
                                     user=interaction.user)
            except (TimeoutError, asyncio.TimeoutError):
                await interaction.followup.send(_("Timeout while deleting mission.\n"
                                                  "Please reconfirm that the deletion was successful."),
                                                ephemeral=ephemeral)

    @mission.command(description=_('Pauses the current running mission'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def pause(self, interaction: discord.Interaction,
                    server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])]):
        ephemeral = utils.get_ephemeral(interaction)
        if server.status == Status.RUNNING:
            await interaction.response.defer(thinking=True, ephemeral=ephemeral)
            await server.current_mission.pause()
            await interaction.followup.send(_('Mission on server "{}" paused.').format(server.display_name),
                                            ephemeral=ephemeral)
        else:
            await interaction.response.send_message(_('Server {} is not running.').format(server.display_name),
                                                    ephemeral=ephemeral)

    @mission.command(description=_('Resumes the running mission'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def unpause(self, interaction: discord.Interaction,
                      server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.PAUSED])]):
        ephemeral = utils.get_ephemeral(interaction)
        if server.status == Status.PAUSED:
            await interaction.response.defer(thinking=True, ephemeral=ephemeral)
            await server.current_mission.unpause()
            await interaction.followup.send(_('Mission on server "{}" resumed.').format(server.display_name),
                                            ephemeral=ephemeral)
        elif server.status == Status.RUNNING:
            await interaction.response.send_message(_('Server "{}" is not paused.').format(server.display_name),
                                                    ephemeral=ephemeral)
        else:
            await interaction.response.send_message(
                _("Server {server} is {status}, can't unpause.").format(server=server.display_name,
                                                                        status=server.status.name),
                ephemeral=ephemeral)

    async def simulate(self, interaction: discord.Interaction, server: Server, use_orig: bool, presets_file: str,
                       presets: list[str] | None, ephemeral: bool):

        presets = {x: utils.get_preset(self.node, x, filename=presets_file) for x in presets} if presets else None
        if not presets:
            await interaction.followup.send("No presets provided for simulation.")
            return

        mission_file = await server.get_current_mission_file()
        if use_orig:
            if server.is_remote:
                await interaction.followup.send(
                    "Simulation is currently only supported on local servers.", ephemeral=True)
                return
            mission_file = utils.get_orig_file(mission_file)
        old_mission: MizFile = await asyncio.to_thread(MizFile, mission_file)
        new_mission: MizFile = await asyncio.to_thread(MizFile, mission_file)

        for k, v in presets.items():
            try:
                await asyncio.to_thread(new_mission.apply_preset, v)
            except Exception as ex:
                self.log.exception(ex)
                await interaction.followup.send(
                    _("Error while applying preset {}: {}").format(k, ex), ephemeral=True)
                return

        changed = False
        if old_mission.mission != new_mission.mission:
            changed = True
            mission_change = utils.show_dict_diff(old_mission.mission, new_mission.mission)
            self.log.debug(f"Mission change:\n{mission_change}")
        else:
            mission_change = None
        if old_mission.warehouses != new_mission.warehouses:
            changed = True
            warehouses_change = utils.show_dict_diff(old_mission.warehouses, new_mission.warehouses)
            self.log.debug(f"Warehouses change:\n{warehouses_change}")
        else:
            warehouses_change = None
        if old_mission.options != new_mission.options:
            changed = True
            options_change = utils.show_dict_diff(old_mission.options, new_mission.options)
            self.log.debug(f"Options change:\n{options_change}")
        else:
            options_change = None

        if changed:
            view = ModifyView(presets, mission_change, warehouses_change, options_change)
            msg = await interaction.followup.send(embed=view.embed, view=view, ephemeral=ephemeral)
            try:
                await view.wait()
            finally:
                await msg.delete()
        else:
            await interaction.followup.send("Your mission was not changed.", ephemeral=ephemeral)
        return

    @mission.command(description=_('Modify mission with a preset\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.autocomplete(presets_file=presets_autocomplete)
    @app_commands.describe(presets_file=_('Chose an alternate presets file'))
    @app_commands.describe(use_orig="Change the mission based on the original uploaded mission file.")
    @app_commands.describe(simulate_only="This will only show you what would happen but not apply the preset")
    async def modify(self, interaction: discord.Interaction,
                     server: app_commands.Transform[Server, utils.ServerTransformer(
                         status=[Status.RUNNING, Status.PAUSED, Status.STOPPED, Status.SHUTDOWN])],
                     presets_file: str | None = None, use_orig: bool | None = True,
                     simulate_only: bool | None = False):
        ephemeral = utils.get_ephemeral(interaction)

        if presets_file is None:
            presets_file = os.path.join(self.node.config_dir, 'presets.yaml')
        try:
            with open(presets_file, mode='r', encoding='utf-8') as infile:
                presets = yaml.load(infile)
        except FileNotFoundError:
            await interaction.response.send_message(
                _('No presets available, please configure them in {}.').format(presets_file), ephemeral=True)
            return

        await interaction.response.defer(ephemeral=ephemeral)
        try:
            try:
                next((x for x in presets.values() if 'terrain' in x or 'terrains' in x), None)
                terrain = await server.get_current_mission_theatre()
            except StopIteration:
                terrain = ''
            options = [
                discord.SelectOption(label=k)
                for k, v in presets.items()
                if not isinstance(v, dict) or (
                        not v.get('hidden', False)
                        and v.get('terrain', terrain) == terrain
                        and terrain in v.get('terrains', [terrain])
                )
            ]
        except AttributeError:
            await interaction.followup.send(
                _("There is an error in your {}. Please check the file structure.").format(presets_file),
                ephemeral=True)
            return
        if len(options) > 25:
            self.log.warning("You have more than 25 presets created, you can only choose from 25!")
        elif not options:
            await interaction.followup.send(_("There are no presets to chose from."), ephemeral=True)

        if server.restart_pending and not await utils.yn_question(
                interaction,
                _('A restart is currently pending.\nWould you still like to modify the mission?'), ephemeral=ephemeral
        ):
            return

        server.on_empty = dict()
        result = None
        if server.status in [Status.PAUSED, Status.RUNNING]:
            question = _('Do you want to restart the server for a mission change?')
            if server.is_populated():
                result = await utils.populated_question(interaction, question, ephemeral=ephemeral)
            else:
                result = await utils.yn_question(interaction, question, ephemeral=ephemeral)
            if not result:
                return

        view = PresetView(options[:25])
        msg = await interaction.followup.send(view=view, ephemeral=ephemeral)
        try:
            if await view.wait() or view.result is None:
                return
        finally:
            await msg.delete()

        # noinspection PyUnreachableCode
        if simulate_only:
            await self.simulate(interaction, server, use_orig, presets_file, view.result, ephemeral)
            return

        if result == 'later':
            server.on_empty = {
                "method": "restart",
                "presets": presets_file,
                "settings": view.result,
                "use_orig": use_orig,
                "user": interaction.user
            }
            server.restart_pending = True
            await interaction.followup.send(_('Mission will be changed when server is empty.'), ephemeral=ephemeral)
            return
        else:
            server.on_empty = dict()
            startup = False
            msg = await interaction.followup.send(_('Changing mission ...'), ephemeral=ephemeral)
            # we need to stop the mission if rewrite is false
            if not server.locals.get('mission_rewrite', True) and server.status in [Status.PAUSED, Status.RUNNING]:
                await server.stop()
                startup = True
            filename = await server.get_current_mission_file()
            new_filename = await server.modifyMission(
                filename,
                [utils.get_preset(self.node, x, presets_file) for x in view.result],
                use_orig=use_orig
            )
            message = _('The following preset were applied: {}.').format(','.join(view.result))
            if new_filename != filename:
                self.log.info(f"  => {message}")
                self.log.info(f"  => New mission written: {new_filename}")
                await server.replaceMission(int(server.settings['listStartIndex']), new_filename)
            else:
                self.log.info(f"  => Mission {filename} overwritten.")
            if startup or server.status not in [Status.STOPPED, Status.SHUTDOWN]:
                try:
                    # if the filename has not changed, we can just restart the running mission
                    if filename == new_filename:
                        await server.restart(modify_mission=False)
                    # otherwise we load the new mission
                    else:
                        await server.loadMission(new_filename, modify_mission=False, use_orig=False)
                    message += _('\nMission reloaded.')
                    await self.bot.audit("changed preset {}".format(','.join(view.result)), server=server,
                                         user=interaction.user)
                except (TimeoutError, asyncio.TimeoutError):
                    message = _("Timeout during restart of mission!\n"
                                "Please check, if the mission is running or if it somehow got corrupted.")
            await msg.edit(content=message)

    @mission.command(description=_('Save mission preset\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def save_preset(self, interaction: discord.Interaction,
                          server: app_commands.Transform[Server, utils.ServerTransformer(
                              status=[Status.RUNNING, Status.PAUSED, Status.STOPPED])],
                          name: str):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        miz = await asyncio.to_thread(MizFile, server.current_mission.filename)
        config_file = os.path.join(self.node.config_dir, 'presets.yaml')
        if os.path.exists(config_file):
            with open(config_file, mode='r', encoding='utf-8') as infile:
                presets = yaml.load(infile)
        else:
            presets = {}
        if name in presets and \
                not await utils.yn_question(interaction,
                                            _('Do you want to overwrite the existing preset "{}"?').format(name),
                                            ephemeral=ephemeral):
            return
        presets[name] = {
            "start_time": miz.start_time,
            "date": miz.date.strftime('%Y-%m-%d'),
            "temperature": miz.temperature,
            "clouds": miz.clouds,
            "wind": miz.wind,
            "groundTurbulence": miz.groundTurbulence,
            "enable_dust": miz.enable_dust,
            "dust_density": miz.dust_density if miz.enable_dust else 0,
            "qnh": miz.qnh,
            "enable_fog": miz.enable_fog,
            "fog": miz.fog if miz.enable_fog else {"thickness": 0, "visibility": 0},
            "halo": miz.halo
        }
        with open(config_file, mode='w', encoding='utf-8') as outfile:
            yaml.dump(presets, outfile)
        await interaction.followup.send(_('Preset "{}" added.').format(name), ephemeral=ephemeral)

    @mission.command(description=_('Rollback to the original mission file after any modifications'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.rename(mission_id="mission")
    @app_commands.autocomplete(mission_id=orig_mission_autocomplete)
    async def rollback(self, interaction: discord.Interaction,
                       server: app_commands.Transform[Server, utils.ServerTransformer], mission_id: int):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        missions = await server.getMissionList()
        if mission_id >= len(missions):
            await interaction.followup.send(_("No mission found."), ephemeral=True)
            return
        filename = missions[mission_id]
        if server.status in [Status.RUNNING, Status.PAUSED] and filename == server.current_mission.filename:
            await interaction.followup.send(_("Please stop your server first to rollback the running mission."),
                                            ephemeral=True)
            return

        if '.dcssb' in filename:
            new_file = os.path.join(os.path.dirname(filename).replace('.dcssb', ''),
                                    os.path.basename(filename))
            orig_file = filename + '.orig'
        else:
            new_file = filename
            orig_file = os.path.join(os.path.dirname(filename), '.dcssb', os.path.basename(filename)) + '.orig'
        try:
            orig_file = orig_file.replace('.sav', '.miz')
            new_file = new_file.replace('.sav', '.miz')
            await server.node.rename_file(orig_file, new_file, force=True)
            if filename.endswith('.sav'):
                await server.node.remove_file(filename)
        except FileNotFoundError:
            # we should never be here, but just in case
            await interaction.followup.send(_('No ".orig" file there, the mission was never changed.'),
                                            ephemeral=True)
            return
        if new_file != filename:
            await server.replaceMission(mission_id + 1, new_file)
        await interaction.followup.send(_("Mission {} has been rolled back.").format(os.path.basename(filename)[:-4]),
                                        ephemeral=ephemeral)

    @mission.command(description=_('Sets fog in the running mission'))
    @app_commands.guild_only()
    @app_commands.describe(thickness=_("Thickness of the fog [100-5000]m, to disable, set 0."))
    @app_commands.describe(visibility=_("Visibility of the fog [100-100000]m, to disable, set 0."))
    @utils.app_has_role('DCS Admin')
    @utils.app_has_dcs_version("2.9.10")
    async def fog(self, interaction: discord.Interaction,
                  server: app_commands.Transform[Server, utils.ServerTransformer(
                      status=[Status.RUNNING, Status.PAUSED])],
                  thickness: app_commands.Range[int, 0, 5000] | None = None,
                  visibility: app_commands.Range[int, 0, 100000] | None = None):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        if thickness is None and visibility is None:
            ret = await server.send_to_dcs_sync({
                "command": "getFog"
            }, timeout=60)
        else:
            if thickness and thickness < 100:
                await interaction.followup.send(_("Thickness has to be in the range 100-5000"))
                return
            if visibility and visibility < 100:
                await interaction.followup.send(_("Visibility has to be in the range 100-100000"))
                return
            ret = await server.send_to_dcs_sync({
                "command": "setFog",
                "thickness": thickness if thickness is not None else -1,
                "visibility": visibility if visibility is not None else -1
            }, timeout=60)
        await interaction.followup.send(_("Current Fog Settings:\n- Thickness: {thickness:.2f}m\n- Visibility:\t{visibility:.2f}m").format(
            thickness=ret['thickness'], visibility=ret['visibility']), ephemeral=ephemeral)

    @mission.command(description=_('Runs a fog animation'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.autocomplete(presets_file=presets_autocomplete)
    @utils.app_has_dcs_version("2.9.10")
    async def fog_animation(self, interaction: discord.Interaction,
                            server: app_commands.Transform[Server, utils.ServerTransformer(
                                status=[Status.RUNNING, Status.PAUSED])],
                            presets_file: str | None = None):
        ephemeral = utils.get_ephemeral(interaction)
        if presets_file is None:
            presets_file = os.path.join(self.node.config_dir, 'presets.yaml')
        try:
            with open(presets_file, mode='r', encoding='utf-8') as infile:
                presets = yaml.load(infile)
        except FileNotFoundError:
            await interaction.response.send_message(
                _('No presets available, please configure them in {}.').format(presets_file), ephemeral=True)
            return
        try:
            options = [
                discord.SelectOption(label=k)
                for k, v in presets.items()
                if not v.get('hidden', False) and v.get('fog') and
                   (v['fog'].get('mode', None) == 'manual' or all(isinstance(y, int) for y in v['fog'].keys()))
            ]
        except AttributeError:
            await interaction.response.send_message(
                _("There is an error in your {}. Please check the file structure.").format(presets_file),
                ephemeral=True)
            return
        if len(options) > 25:
            self.log.warning("You have more than 25 presets created, you can only choose from 25!")
        elif not options:
            await interaction.response.send_message(_("There is no manual fog preset in your {}").format(presets_file),
                                                    ephemeral=True)
            return

        # select a preset
        view = PresetView(options[:25], multi=False)
        if interaction.response.is_done():
            msg = await interaction.followup.send(view=view, ephemeral=ephemeral)
        else:
            await interaction.response.send_message(view=view, ephemeral=ephemeral)
            msg = await interaction.original_response()
        try:
            if not await view.wait() and view.result is not None:
                fog = utils.get_preset(self.node, view.result[0], presets_file)['fog']
                fog.pop('mode', None)
                await server.send_to_dcs_sync(
                    {
                        'command': 'setFogAnimation',
                        'values': [
                            (key, value["visibility"], value["thickness"])
                            for key, value in fog.items()
                        ]
                    }, timeout=60)
                message = _('The following preset was applied: {}.').format(view.result[0])
                await interaction.followup.send(message, ephemeral=ephemeral)
        finally:
            with suppress(discord.NotFound):
                await msg.delete()

    @mission.command(description=_('Enables persistence'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.rename(mission_id="mission")
    @app_commands.autocomplete(mission_id=nosav_autocomplete)
    @utils.app_has_dcs_version("2.9.14")
    async def persistence(self, interaction: discord.Interaction,
                          server: app_commands.Transform[Server, utils.ServerTransformer], mission_id: int):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        missions = await server.getMissionList()
        if mission_id >= len(missions):
            await interaction.followup.send(_("No mission found."), ephemeral=True)
            return
        filename = missions[mission_id]
        new_file = filename.replace('.miz', '.sav')
        if server.status in [Status.RUNNING, Status.PAUSED] and filename == server.current_mission.filename:
            await interaction.followup.send(
                _("Please stop your server first to enable persistence of the running mission."), ephemeral=True)
            return
        try:
            utils.get_orig_file(filename)
            await server.node.rename_file(filename, new_file, force=True)
        except FileNotFoundError:
            # we should never be here, but just in case
            await interaction.followup.send(_('No ".orig" file found.'), ephemeral=True)
            return
        except PermissionError:
            await interaction.followup.send(_('{} is currently in use. Aborted.').format(os.path.basename(new_file)),
                                            ephemeral=True)
            return
        await server.replaceMission(mission_id + 1, new_file)
        await interaction.followup.send(
            _("Persistence for mission {} enabled.").format(os.path.basename(filename)[:-4]), ephemeral=ephemeral)

    # New command group "/airbase"
    airbase = Group(name='airbase', description=_('Commands to manage airbases'))

    @airbase.command(name="info", description=_('Information about a specific airbase'))
    @utils.app_has_role('DCS')
    @app_commands.guild_only()
    @app_commands.rename(_server='server')
    @app_commands.rename(idx=_('airbase'))
    @app_commands.describe(idx=_('Airbase for warehouse information'))
    @app_commands.autocomplete(idx=utils.airbase_autocomplete)
    async def airbase_info(self, interaction: discord.Interaction,
                           _server: app_commands.Transform[Server, utils.ServerTransformer(
                               status=[Status.RUNNING, Status.PAUSED])],
                           idx: int):
        if _server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(_server.display_name),
                                                    ephemeral=True)
            return

        await interaction.response.defer(ephemeral=utils.get_ephemeral(interaction))
        airbase = _server.current_mission.airbases[idx]
        data = await _server.send_to_dcs_sync({
            "command": "getAirbase",
            "name": airbase['name']
        }, timeout=60)
        colors = {
            0: "dark_gray",
            1: "red",
            2: "blue"
        }
        report = Report(self.bot, self.plugin_name, 'airbase.json')
        env = await report.render(
            interaction=interaction,
            server=_server,
            coalition=colors[data['coalition']],
            airbase=airbase,
            data=data
        )
        if utils.check_roles(set(self.bot.roles['DCS Admin'] + self.bot.roles['GameMaster']), interaction.user):
            view = AirbaseView(_server, airbase, data)
        else:
            view = discord.utils.MISSING
        msg = await interaction.followup.send(embed=env.embed, view=view)
        if view:
            try:
                await view.wait()
            finally:
                try:
                    await msg.delete()
                except discord.NotFound:
                    pass

    @airbase.command(description=_('Automatic Terminal Information Service (ATIS)'))
    @utils.app_has_role('DCS')
    @app_commands.guild_only()
    @app_commands.rename(_server='server')
    @app_commands.rename(idx=_('airbase'))
    @app_commands.describe(idx=_('Airbase for ATIS information'))
    @app_commands.autocomplete(idx=utils.airbase_autocomplete)
    async def atis(self, interaction: discord.Interaction,
                   _server: app_commands.Transform[Server, utils.ServerTransformer(
                       status=[Status.RUNNING, Status.PAUSED])],
                   idx: int):
        if _server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(_server.display_name),
                                                    ephemeral=True)
            return
        await interaction.response.defer()
        airbase = _server.current_mission.airbases[idx]
        data = await _server.send_to_dcs_sync({
            "command": "getWeatherInfo",
            "x": airbase['position']['x'],
            "y": airbase['position']['y'],
            "z": airbase['position']['z']
        }, timeout=60)
        report = Report(self.bot, self.plugin_name, 'atis.json')
        env = await report.render(airbase=airbase, data=data, server=_server)
        msg = await interaction.original_response()
        await msg.edit(embed=env.embed, delete_after=self.bot.locals.get('message_autodelete'))

    @airbase.command(description=_('Capture an airbase'))
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    @app_commands.guild_only()
    @app_commands.rename(idx=_('airbase'))
    @app_commands.describe(idx=_('Airbase to capture'))
    @app_commands.autocomplete(idx=utils.airbase_autocomplete)
    async def capture(self, interaction: discord.Interaction,
                      server: app_commands.Transform[Server, utils.ServerTransformer(
                          status=[Status.RUNNING, Status.PAUSED])],
                      idx: int, coalition: Literal['Red', 'Blue', 'Neutral']):
        if server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(server.display_name),
                                                    ephemeral=True)
            return

        await interaction.response.defer(ephemeral=utils.get_ephemeral(interaction))
        airbase = server.current_mission.airbases[idx]
        data = await server.send_to_dcs_sync({
            "command": "getAirbase",
            "name": airbase['name']
        }, timeout=60)
        ret_coalition = 'Red' if data['coalition'] == 1 else 'Blue' if data['coalition'] == 2 else 'Neutral'
        if ret_coalition == coalition:
            await interaction.followup.send(_('Airbase \"{}\" belonged to coalition {} already.').format(
                airbase['name'], coalition.lower()), ephemeral=True)
            return

        await server.send_to_dcs_sync({
            "command": "captureAirbase",
            "name": airbase['name'],
            "coalition": 1 if coalition == 'Red' else 2 if coalition == 'Blue' else 0
        }, timeout=60)
        await interaction.followup.send(
            _("Airbase \"{}\": Coalition changed to **{}**.\n:warning: Auto-capturing is now **disabled**!").format(
                airbase['name'], coalition.lower()))

    warehouse = Group(name='warehouse', description=_('Commands to manage warehouses'))

    @staticmethod
    async def manage_items(server: Server, airbase: dict, category: str, item: str | list[int],
                           value: int | None = None) -> dict:
        if value is None:
            if category == 'liquids':
                return await server.send_to_dcs_sync({
                    "command": "getWarehouseLiquid",
                    "name": airbase['name'],
                    "item": int(item)
                }, timeout=60)
            else:
                return await server.send_to_dcs_sync({
                    "command": "getWarehouseItem",
                    "name": airbase['name'],
                    "item": item
                }, timeout=60)
        else:
            if category == 'liquids':
                return await server.send_to_dcs_sync({
                    "command": "setWarehouseLiquid",
                    "name": airbase['name'],
                    "item": int(item),
                    "value": value * 1000
                }, timeout=60)
            else:
                return await server.send_to_dcs_sync({
                    "command": "setWarehouseItem",
                    "name": airbase['name'],
                    "item": item,
                    "value": value
                }, timeout=60)

    @staticmethod
    async def manage_category(server: Server, airbase: dict, category: str, value: int | None = None) -> None:
        tasks = []
        for item in [x['wstype'] for x in server.resources[category]]:
            _item = list(map(int, item.split('.'))) if isinstance(item, str) else item
            tasks.append(asyncio.create_task(
                Mission.manage_items(server, airbase, category, _item, value))
            )
        await asyncio.gather(*tasks)

    @staticmethod
    async def _manage_warehouse(interaction: discord.Interaction, _server: Server, idx: int,
                                category: str | None = None, item: str | None = None, value: int | None = None) -> None:
        if _server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(_server.display_name),
                                                    ephemeral=True)
            return

        await interaction.response.defer(ephemeral=True)
        airbase = _server.current_mission.airbases[idx]
        data = await _server.send_to_dcs_sync({
            "command": "getAirbase",
            "name": airbase['name']
        }, timeout=60)

        sides = utils.get_sides(interaction.client, interaction, _server)
        if ((data['coalition'] == 2 and Coalition.BLUE not in sides) or
                (data['coalition'] == 1 and Coalition.RED not in sides)):
            await interaction.followup.send(
                _("You are not allowed to view a warehouse item of the opposite coalition."))
            return

        embed = discord.Embed(title=_("Warehouse information for {}").format(airbase['name']),
                              color=discord.Color.blue())

        if item:
            _item = list(map(int, item.split('.'))) if isinstance(item, str) else item
            data = await Mission.manage_items(_server, airbase, category, _item, value)
            if data['value'] == 1000000:
                display = _("unlimited")
            elif category == 'liquids':
                display = _("{} tons").format(data['value'] / 1000)
            else:
                display = _("{} pcs").format(data['value'])

            item_name = next(
                (x['name'] for x in _server.resources.get(category) if str(x['wstype']) == item),
                'n/a'
            )
            embed.add_field(name=_("Inventory for {}").format(item_name), value="```" + display + "```")
            await interaction.followup.send(embed=embed)

        else:
            if value is not None:
                message = _("Do you really want to set all {} values in your warehouse to {}?").format(
                    category, value)
                if not await utils.yn_question(interaction, message):
                    await interaction.followup.send(_("Aborted."))
                    return

                await Mission.manage_category(_server, airbase, category, value)

            data = await _server.send_to_dcs_sync({
                "command": "getAirbase",
                "name": airbase['name']
            }, timeout=60)

            if not category or category == 'liquids':
                Info.render_liquids(embed, data)
            if not category or category == 'weapon':
                Info.render_weapons(embed, data)
            if not category or category == 'aircraft':
                Info.render_aircraft(embed, data)

            await interaction.followup.send(embed=embed, ephemeral=utils.get_ephemeral(interaction))

    @warehouse.command(description=_('Set warehouses items'))
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    @app_commands.guild_only()
    @app_commands.rename(_server='server')
    @app_commands.rename(idx=_('airbase'))
    @app_commands.describe(idx=_('Airbase for warehouse information'))
    @app_commands.autocomplete(idx=utils.airbase_autocomplete)
    @app_commands.autocomplete(category=wh_category_autocomplete)
    @app_commands.autocomplete(item=wh_item_autocomplete)
    async def set(self, interaction: discord.Interaction,
                        _server: app_commands.Transform[Server, utils.ServerTransformer(
                            status=[Status.RUNNING, Status.PAUSED])],
                        idx: int, category: str, item: str | None = None, value: int = 0):
        await self._manage_warehouse(interaction, _server, idx, category, item, value)

    @warehouse.command(description=_('Get warehouses items'))
    @utils.app_has_role('DCS')
    @app_commands.guild_only()
    @app_commands.rename(_server='server')
    @app_commands.rename(idx=_('airbase'))
    @app_commands.describe(idx=_('Airbase for warehouse information'))
    @app_commands.autocomplete(idx=utils.airbase_autocomplete)
    @app_commands.autocomplete(category=wh_category_autocomplete)
    @app_commands.autocomplete(item=wh_item_autocomplete)
    async def get(self, interaction: discord.Interaction,
                  _server: app_commands.Transform[Server, utils.ServerTransformer(
                      status=[Status.RUNNING, Status.PAUSED])],
                  idx: int, category: str | None = None, item: str | None = None):
        await self._manage_warehouse(interaction, _server, idx, category, item)

    @staticmethod
    async def download_warehouse(interaction: discord.Interaction, airbase: dict, data: dict):
        def dict_to_df(mapping: dict) -> pd.DataFrame:
            df = pd.DataFrame(list(mapping.items()), columns=["Name", "Count"])
            return df

        sheet_titles = {
            "aircraft": "Aircraft",
            "weapon": "Weapons",
            "liquids": "Liquids",
        }

        warehouse = data.get('warehouse', {})

        dataframes = {}
        for key, title in sheet_titles.items():
            inner_dict = warehouse.get(key)
            if key == "liquids":
                mapped = {LIQUIDS.get(k, k): v for k, v in inner_dict.items()}
            else:
                mapped = inner_dict
            df = dict_to_df(mapped)
            df.sort_values("Name", inplace=True)
            dataframes[title] = df

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for sheet_name, df in dataframes.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            wb = writer.book
            for sheet_name, df in dataframes.items():
                ws = wb[sheet_name]
                ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

                for col_idx, col_name in enumerate(df.columns, start=1):
                    max_len = max(
                        df[col_name].astype(str).map(len).max(),  # data
                        len(col_name)  # header
                    )
                    column_letter = get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = max_len + 2

        buffer.seek(0)
        try:
            code = utils.slugify(airbase.get('code', '') or airbase.get('name', 'XXXX'))
            await interaction.followup.send(
                file=discord.File(fp=buffer, filename=f"warehouse-{code.lower()}.xlsx"), ephemeral=True
            )
        finally:
            buffer.close()

    @warehouse.command(name="export", description=_('Export a warehouse'))
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    @app_commands.guild_only()
    @app_commands.rename(_server='server')
    @app_commands.rename(idx=_('airbase'))
    @app_commands.describe(idx=_('Airbase for warehouse information'))
    @app_commands.autocomplete(idx=utils.airbase_autocomplete)
    async def export(self, interaction: discord.Interaction,
                     _server: app_commands.Transform[Server, utils.ServerTransformer(
                         status=[Status.RUNNING, Status.PAUSED])],
                     idx: int):
        if _server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(_server.display_name),
                                                    ephemeral=True)
            return

        await interaction.response.defer(ephemeral=utils.get_ephemeral(interaction))
        airbase = _server.current_mission.airbases[idx]
        data = await _server.send_to_dcs_sync({
            "command": "getAirbase",
            "name": airbase['name']
        }, timeout=60)
        await self.download_warehouse(interaction, airbase, data)

    @warehouse.command(name="list", description=_('Export all possible wstypes'))
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    @app_commands.guild_only()
    @app_commands.rename(_server='server')
    async def wstypes(self, interaction: discord.Interaction,
                      _server: app_commands.Transform[Server, utils.ServerTransformer(
                         status=[Status.RUNNING, Status.PAUSED])]):
        if _server.status not in [Status.RUNNING, Status.PAUSED]:
            await interaction.response.send_message(_("Server {} is not running.").format(_server.display_name),
                                                    ephemeral=True)
            return

        await interaction.response.defer(ephemeral=utils.get_ephemeral(interaction))
        sheet_titles = {
            "aircraft": "Aircraft",
            "weapon": "Weapons",
            "liquids": "Liquids",
        }

        dataframes = {}
        for key, title in sheet_titles.items():
            dataframes[title] = pd.DataFrame.from_dict(_server.resources.get(key),
                                                       orient='index', columns=['wsType', 'Name'])

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for sheet_name, df in dataframes.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            wb = writer.book
            for sheet_name, df in dataframes.items():
                ws = wb[sheet_name]
                ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

                for col_idx, col_name in enumerate(df.columns, start=1):
                    max_len = max(
                        df[col_name].astype(str).map(len).max(),  # data
                        len(col_name)  # header
                    )
                    column_letter = get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = max_len + 2

        buffer.seek(0)
        try:
            await interaction.followup.send(
                file=discord.File(fp=buffer, filename=f"warehouse-items.xlsx"), ephemeral=True
            )
        finally:
            buffer.close()

    # New command group "/player"
    player = Group(name="player", description=_("Commands to manage DCS players"))

    @player.command(name='list', description=_('Lists the current players'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS')
    async def _list(self, interaction: discord.Interaction,
                    server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])]):
        if server.status != Status.RUNNING:
            await interaction.response.send_message(_("Server {} is not running.").format(server.display_name),
                                                    ephemeral=True)
            return
        report = Report(self.bot, self.plugin_name, 'players.json')
        env = await report.render(server=server, sides=utils.get_sides(interaction.client, interaction, server))
        await interaction.response.send_message(embed=env.embed, ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Kicks a player\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def kick(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                   player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)],
                   reason: str | None = 'n/a') -> None:
        if not player:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return
        await server.kick(player, reason)
        await self.bot.audit(f'kicked player {player.display_name} with reason "{reason}"', user=interaction.user)
        await interaction.response.send_message(
            _("Player {name} (ucid={ucid}) kicked.").format(name=player.display_name, ucid=player.ucid),
            ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Bans an active player'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def ban(self, interaction: discord.Interaction,
                  server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                  player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)]):

        if not player:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return
        await interaction.response.send_modal(BanModal(player.ucid))

    @player.command(description=_('Locks a player'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def lock(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                   player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)]):
        await player.lock()
        await interaction.response.send_message(_("Player {} has been locked.").format(player.display_name),
                                                ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Unlocks a player'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def unlock(self, interaction: discord.Interaction,
                     server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                     user: app_commands.Transform[
                         discord.Member | str, utils.UserTransformer(sel_type=PlayerType.PLAYER)
                     ]):
        if isinstance(user, discord.Member):
            ucid = await self.bot.get_ucid_by_member(user)
        else:
            ucid = user

        if not ucid:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return

        await server.send_to_dcs({
            "command": "unlock_player",
            "ucid": ucid
        })
        await interaction.response.send_message(_("Player has been unlocked."),
                                                ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Mutes a player'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def mute(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                   player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)]):
        await player.mute()
        await interaction.response.send_message(_("Player {} has been muted.").format(player.display_name),
                                                ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Unmutes a player'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def unmute(self, interaction: discord.Interaction,
                     server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                     player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)]):
        await player.unmute()
        await interaction.response.send_message(_("Player has been unmuted."),
                                                ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Moves a player to spectators\n'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def spec(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                   player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)],
                   reason: str | None = 'n/a') -> None:
        if not player:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return
        await server.move_to_spectators(player)
        if reason:
            await player.sendChatMessage(_("You have been moved to spectators. Reason: {}").format(reason),
                                         interaction.user.display_name)
        await self.bot.audit(f'moved player {player.name} to spectators with reason "{reason}".', user=interaction.user)
        await interaction.response.send_message(_('Player "{}" moved to spectators.').format(player.name),
                                                ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('List of AFK players'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def afk(self, interaction: discord.Interaction,
                  server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                  minutes: int | None = 10):
        if server.status != Status.RUNNING:
            await interaction.response.send_message(_("Server {} is not running.").format(server.display_name),
                                                    ephemeral=True)
            return
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        afk: list[Player] = list()
        for s in self.bot.servers.values():
            if server and s != server:
                continue
            for ucid, dt in s.afk.items():
                player = s.get_player(ucid=ucid, active=True)
                if not player:
                    continue
                if (datetime.now(tz=timezone.utc) - dt).total_seconds() > minutes * 60:
                    afk.append(player)

        if afk:
            title = 'AFK Players'
            if server:
                title += f' on {server.name}'
            embed = discord.Embed(title=title, color=discord.Color.blue())
            embed.description = _('These players are AFK for more than {} minutes:').format(minutes)
            for player in sorted(afk, key=lambda x: x.server.name):
                embed.add_field(name=_('Name'), value=player.display_name)
                embed.add_field(name=_('Time'),
                                value=utils.format_time(int((datetime.now(timezone.utc) -
                                                             player.server.afk[player.ucid]).total_seconds())))
                if server:
                    embed.add_field(name='_ _', value='_ _')
                else:
                    embed.add_field(name=_('Server'), value=player.server.display_name)
            await interaction.followup.send(embed=embed, ephemeral=ephemeral)
        else:
            await interaction.followup.send(_("No player is AFK for more than {} minutes.").format(minutes),
                                            ephemeral=ephemeral)

    @player.command(description=_('Exempt player from AFK kicks'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def exempt(self, interaction: discord.Interaction,
                     user: app_commands.Transform[
                         discord.Member | str, utils.UserTransformer(sel_type=PlayerType.PLAYER)
                     ],
                     server: app_commands.Transform[Server, utils.ServerTransformer] | None):
        ephemeral = utils.get_ephemeral(interaction)
        if isinstance(user, discord.Member):
            ucid = await self.bot.get_ucid_by_member(user)
        else:
            ucid = user
        config_file = os.path.join(self.node.config_dir, 'servers.yaml')
        if not server:
            section = DEFAULT_TAG
        else:
            section = server.name
        data = yaml.load(Path(config_file).read_text(encoding='utf-8'))
        if section not in data:
            data[section] = {}
        if 'afk' not in data[section]:
            data[section]['afk'] = {}
        if 'exemptions' not in data[section]['afk']:
            data[section]['afk']['exemptions'] = {}
        if 'ucid' not in data[section]['afk']['exemptions']:
            data[section]['afk']['exemptions']['ucid'] = []
        if ucid not in data[section]['afk']['exemptions']['ucid']:
            if not await utils.yn_question(interaction,
                                           _("Do you want to permanently add this user to the AFK exemption list?"),
                                           ephemeral=ephemeral):
                await interaction.followup.send("Aborted.", ephemeral=ephemeral)
                return
            data[section]['afk']['exemptions']['ucid'].append(ucid)
            await interaction.followup.send(_("User added to the exemption list."), ephemeral=ephemeral)
        else:
            if not await utils.yn_question(interaction,
                                           _("Player is on the list already. Do you want to remove them?")):
                await interaction.followup.send(_("Aborted."), ephemeral=ephemeral)
                return
            data[section]['afk']['exemptions']['ucid'].remove(ucid)
            await interaction.followup.send(_("User removed from the exemption list."), ephemeral=ephemeral)
        with open(config_file, 'w', encoding='utf-8') as outfile:
            yaml.dump(data, outfile)

    @player.command(description=_('Sends a popup to a player\n'))
    @app_commands.guild_only()
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    async def popup(self, interaction: discord.Interaction,
                    server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                    player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)],
                    message: str, time: Range[int, 1, 30] | None = -1):
        if not player:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return
        await player.sendPopupMessage(message, time, interaction.user.display_name)
        await interaction.response.send_message(_('Message sent.'), ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Sends a chat message to a player\n'))
    @app_commands.guild_only()
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    async def chat(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                   player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)], message: str):
        if not player:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return
        await player.sendChatMessage(message, interaction.user.display_name)
        await interaction.response.send_message(_('Message sent.'), ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Take a screenshot'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def screenshot(self, interaction: discord.Interaction,
                   server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                   player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)]) -> None:
        if not player:
            await interaction.response.send_message(_("Player not found."), ephemeral=True)
            return
        if not server.settings.get('advanced', {}).get('server_can_screenshot'):
            await interaction.response.send_message(_("Server can not take screenshots."), ephemeral=True)
            return
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        msg = await interaction.followup.send(_("Requesting screenshot ..."), ephemeral=ephemeral)
        try:
            old_screens = await player.getScreenshots()
            await player.makeScreenshot()
            timeout = 30 if server.node.locals.get('slow_system', False) else 10
            for i in range(1, timeout):
                await asyncio.sleep(1)
                new_screens = await player.getScreenshots()
                if len(new_screens) > len(old_screens):
                    break
            else:
                await msg.edit(content=_("Timeout while waiting for screenshot!"))
                return
        except (TimeoutError, asyncio.TimeoutError):
            await msg.edit(content=_("Timeout while waiting for screenshot!"))
            return
        key = new_screens[-1]
        # DCS 2.9.11+
        if 'screenshots' not in key:
            key = '/screenshots/' + key
        try:
            image_url = f"http://127.0.0.1:{server.instance.webgui_port}{key}"
            image_data = await server.node.read_file(image_url)
            file = discord.File(BytesIO(image_data), filename="screenshot.png")
            await msg.delete()
            embed = discord.Embed(color=discord.Color.blue(),
                                  title=_("Screenshot of Player {}").format(player.display_name))
            embed.set_image(url="attachment://screenshot.png")
            embed.add_field(name=_("Server"), value=server.display_name, inline=False)
            embed.add_field(name=_("Time"), value=f"<t:{int(datetime.now().timestamp())}>", inline=False)
            embed.add_field(name=_("Taken by"), value=interaction.user.display_name, inline=False)
            await interaction.followup.send(embed=embed, file=file, ephemeral=ephemeral)
        finally:
            await player.deleteScreenshot(key)

    watch = Group(name="watch", description="Commands to manage the watchlist")

    @watch.command(name='add', description=_('Puts a player onto the watchlist'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def _add(self, interaction: discord.Interaction,
                   user: app_commands.Transform[discord.Member | str, utils.UserTransformer(
                       sel_type=PlayerType.PLAYER, watchlist=False)], reason: str):
        if isinstance(user, discord.Member):
            ucid = await self.bot.get_ucid_by_member(user)
            if not ucid:
                await interaction.response.send_message(_("Member {} is not linked!").format(user.display_name),
                                                        ephemeral=True)
                return
        elif utils.is_ucid(user):
            ucid = user
        else:
            await interaction.response.send_message(_("User not found."), ephemeral=True)
            return

        try:
            async with self.apool.connection() as conn:
                await conn.execute("INSERT INTO watchlist (player_ucid, reason, created_by) VALUES (%s, %s, %s)",
                                   (ucid, reason, interaction.user.display_name))
            await interaction.response.send_message(_("Player {} is now on the watchlist.").format(
                user.display_name if isinstance(user, discord.Member) else ucid),
                ephemeral=utils.get_ephemeral(interaction))
        except psycopg.errors.UniqueViolation:
            await interaction.response.send_message(
                _("Player {} was already on the watchlist.").format(
                    user.display_name if isinstance(user, discord.Member) else ucid),
                ephemeral=utils.get_ephemeral(interaction))

    @watch.command(description=_('Removes a player from the watchlist'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def delete(self, interaction: discord.Interaction,
                     user: app_commands.Transform[discord.Member | str, utils.UserTransformer(
                         sel_type=PlayerType.PLAYER, watchlist=True)]):
        if isinstance(user, discord.Member):
            ucid = await self.bot.get_ucid_by_member(user)
            if not ucid:
                # we should never be here
                await interaction.response.send_message(_("Member {} is not linked!").format(user.display_name))
                return
        else:
            ucid = user
        async with self.apool.connection() as conn:
            await conn.execute("DELETE FROM watchlist WHERE player_ucid = %s", (ucid, ))
        await interaction.response.send_message(
            _("Player {} removed from the watchlist.").format(
                user.display_name if isinstance(user, discord.Member) else user),
            ephemeral=utils.get_ephemeral(interaction))

    @watch.command(name='list', description=_('Shows the watchlist'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def _list(self, interaction: discord.Interaction):
        ephemeral = utils.get_ephemeral(interaction)
        async with self.apool.connection() as conn:
            cursor = await conn.execute("""
                SELECT p.ucid, p.name, w.reason, w.created_by, w.created_at 
                FROM players p JOIN watchlist w ON (p.ucid = w.player_ucid)
            """)
            watches = await cursor.fetchall()
        if not watches:
            await interaction.response.send_message(_("The watchlist is currently empty."), ephemeral=ephemeral)
            return
        embed = discord.Embed(colour=discord.Colour.blue())
        embed.description = _("These players are currently on the watchlist:")
        names = created_by = ucids = ""
        for row in watches:
            names += utils.escape_string(row[1]) + '\n'
            ucids += row[0] + '\n'
            created_by += row[3] + '\n'
        embed.add_field(name=_("Name"), value=names)
        embed.add_field(name=_('UCID'), value=ucids)
        embed.add_field(name=_("Created by"), value=created_by)
        await interaction.response.send_message(embed=embed)

    # New command group "/group"
    group = Group(name="group", description="Commands to manage DCS groups")

    @group.command(description=_('Sends a popup to a group\n'))
    @app_commands.guild_only()
    @app_commands.autocomplete(group=utils.group_autocomplete)
    @utils.app_has_roles(['DCS Admin', 'GameMaster'])
    async def popup(self, interaction: discord.Interaction,
                    server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                    group: str, message: str, time: Range[int, 1, 30] | None = -1):
        await server.sendPopupMessage(group, message, time, interaction.user.display_name)
        await interaction.response.send_message(_('Message sent.'), ephemeral=utils.get_ephemeral(interaction))

    @command(description=_("Links a member to a DCS user"))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def link(self, interaction: discord.Interaction, member: discord.Member,
                   user: app_commands.Transform[discord.Member | str, utils.UserTransformer(
                       sel_type=PlayerType.PLAYER, linked=False)]
                   ):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        _member = DataObjectFactory().new(Member, name=member.name, node=self.node, member=member)
        if isinstance(user, discord.Member):
            _new_member = DataObjectFactory().new(Member, name=user.name, node=self.node, member=user)
            ucid = _new_member.ucid
            if ucid == _member.ucid:
                if _member.verified:
                    await interaction.followup.send(_("This member is linked to this UCID already."),
                                                    ephemeral=ephemeral)
                    return
            elif not await utils.yn_question(
                interaction, _("Member {name} is linked to another UCID ({ucid}) already. "
                               "Do you want to relink?").format(
                    name=utils.escape_string(user.display_name), ucid=ucid), ephemeral=ephemeral):
                return
            else:
                _new_member.unlink()
        else:
            ucid = user
        if _member.verified:
            if not await utils.yn_question(
                interaction, _("Member {name} is linked to another UCID ({ucid}) already. "
                               "Do you want to relink?").format(
                    name=utils.escape_string(member.display_name), ucid=_member.ucid), ephemeral=ephemeral):
                return
            else:
                _member.unlink()
        _member.link(ucid, verified=True)
        await interaction.followup.send(_('Member {name} linked to UCID {ucid}.').format(
            name=utils.escape_string(member.display_name), ucid=ucid), ephemeral=utils.get_ephemeral(interaction))
        await self.bot.audit(f'linked member {utils.escape_string(member.display_name)} to ucid {ucid}.',
                             user=interaction.user)
        # If autorole is enabled, give the user the role:
        autorole = self.bot.locals.get('autorole', {}).get('linked')
        if autorole:
            try:
                _role = self.bot.get_role(autorole)
                if not _role:
                    self.log.error(f'Role {autorole} not found!')
                    await interaction.followup.send(_("Role {} not found!").format(autorole), ephemeral=True)
                    return
                await member.add_roles(_role)
            except discord.Forbidden:
                await self.bot.audit(_('permission "Manage Roles" missing.'), user=self.bot.member)
        # Generate the onMemberLinked event
        for server_name, server in self.bot.servers.items():
            player = server.get_player(ucid=ucid, active=True)
            if player:
                player.member = self.bot.get_member_by_ucid(player.ucid)
                player.verified = True
                break
        else:
            server = None
        await self.bus.send_to_node({
            "command": "rpc",
            "service": "ServiceBus",
            "method": "propagate_event",
            "params": {
                "command": "onMemberLinked",
                "server": server.name if server else None,
                "data": {
                    "ucid": ucid,
                    "discord_id": member.id
                }
            }
        })

    @command(description=_('Unlinks a member or ucid'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    @app_commands.describe(user=_('Name of player, member or UCID'))
    async def unlink(self, interaction: discord.Interaction,
                     user: app_commands.Transform[discord.Member | str, utils.UserTransformer(linked=True)]):

        async def unlink_member(member: discord.Member, ucid: str):
            # change the link status of that member if they are an active player
            for server_name, server in self.bot.servers.items():
                player = server.get_player(ucid=ucid, active=True)
                if player:
                    player.member = None
                    player.verified = False
                    break
            else:
                await conn.execute('UPDATE players SET discord_id = -1, manual = FALSE WHERE ucid = %s', (ucid,))
                server = None
            await interaction.followup.send(_('Member {name} unlinked from UCID {ucid}.').format(
                name=utils.escape_string(member.display_name), ucid=ucid), ephemeral=ephemeral)
            await self.bot.audit(
                f'unlinked member {utils.escape_string(member.display_name)} from ucid {ucid}',
                user=interaction.user)
            await self.bus.send_to_node({
                "command": "rpc",
                "service": "ServiceBus",
                "method": "propagate_event",
                "params": {
                    "command": "onMemberUnlinked",
                    "server": server.name if server else None,
                    "data": {
                        "ucid": ucid,
                        "discord_id": member.id
                    }
                }
            })

        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        async with self.apool.connection() as conn:
            if isinstance(user, discord.Member):
                member = user
                cursor = await conn.execute('SELECT ucid FROM players WHERE discord_id = %s', (user.id, ))
                rows = await cursor.fetchall()
                for row in rows:
                    ucid = row[0]
                    await unlink_member(user, ucid)
            elif utils.is_ucid(user):
                ucid = user
                member = self.bot.get_member_by_ucid(ucid)
                if not member:
                    await interaction.followup.send(_('Player is not linked!'), ephemeral=True)
                    return
                await unlink_member(member, ucid)
            else:
                await interaction.followup.send(_('Unknown player / member provided'), ephemeral=True)
                return

        # If autorole is enabled, remove the role from the user:
        autorole = self.bot.locals.get('autorole', {}).get('linked')
        if autorole:
            try:
                await member.remove_roles(self.bot.get_role(autorole))
            except discord.Forbidden:
                await self.bot.audit(_('permission "Manage Roles" missing.'), user=self.bot.member)

    async def _find(self, interaction: discord.Interaction, name: str):
        ephemeral = utils.get_ephemeral(interaction)
        await interaction.response.defer(ephemeral=ephemeral)
        async with self.apool.connection() as conn:
            cursor = await conn.execute("""
                SELECT distinct ucid, name, max(last_seen) FROM (
                    SELECT ucid, name, last_seen FROM players
                    UNION
                    SELECT distinct ucid, name, time AS last_seen FROM players_hist
                ) x
                WHERE x.name ILIKE %s
                GROUP BY ucid, name
                ORDER BY 3 DESC
                LIMIT 25
            """, ('%' + name + '%', ))
            rows = await cursor.fetchall()
        # give back the database session
        last_seen_str = _('last seen')
        options = [
            SelectOption(label=f"{row[1]} ({last_seen_str}: {row[2]:%Y-%m-%d %H:%M})"[:100], value=str(idx))
            for idx, row in enumerate(rows)
        ]
        if not options:
            await interaction.followup.send(_("No user found."))
            return
        idx = await utils.selection(interaction, placeholder=_("Select a User"), options=options, ephemeral=ephemeral)
        if idx:
            await self._info(interaction, rows[int(idx)][0])

    @player.command(description=_('Find a player by name'))
    @utils.app_has_role('DCS Admin')
    @app_commands.guild_only()
    async def find(self, interaction: discord.Interaction, name: str):
        await self._find(interaction, name)

    @command(description=_('Find a player by name'))
    @utils.app_has_role('DCS Admin')
    @app_commands.guild_only()
    async def find(self, interaction: discord.Interaction, name: str):
        await self._find(interaction, name)

    async def _info(self, interaction: discord.Interaction, member: discord.Member | str):
        if not member:
            await interaction.response.send_message(
                _("This user does not exist. Try {} to find them in the historic data.").format(
                    (await utils.get_command(self.bot, name=self.find.name)).mention
                ), ephemeral=True)
            return
        ephemeral = utils.get_ephemeral(interaction)
        if not interaction.response.is_done():
            await interaction.response.defer(ephemeral=ephemeral)
        if isinstance(member, str):
            ucid = member
            member = self.bot.get_member_by_ucid(ucid)
        player: Player | None = None
        for server in self.bot.servers.values():
            if isinstance(member, discord.Member):
                player = server.get_player(discord_id=member.id, active=True)
            else:
                player = server.get_player(ucid=ucid, active=True)
            if player:
                break
        else:
            server = None

        view = InfoView(member=member or ucid, bot=self.bot, ephemeral=ephemeral, player=player, server=server)
        embed = await view.render()
        msg = await interaction.followup.send(embed=embed, view=view, ephemeral=ephemeral)
        try:
            await view.wait()
        finally:
            try:
                await msg.delete()
            except discord.NotFound:
                pass

    @player.command(name="info", description=_('Shows player information'))
    @utils.app_has_role('DCS')
    @app_commands.guild_only()
    async def player_info(self, interaction: discord.Interaction,
                          server: app_commands.Transform[Server, utils.ServerTransformer(status=[Status.RUNNING])],
                          player: app_commands.Transform[Player, utils.PlayerTransformer(active=True)]):
        report = Report(self.bot, 'mission', 'player-info.json')
        env = await report.render(player=player)
        await interaction.response.send_message(embed=env.embed, ephemeral=utils.get_ephemeral(interaction))

    @command(description=_('Shows player information'))
    @utils.app_has_role('DCS Admin')
    @app_commands.guild_only()
    async def info(self, interaction: discord.Interaction,
                   member: app_commands.Transform[discord.Member | str, utils.UserTransformer]):
        await self._info(interaction, member)

    @staticmethod
    def format_unmatched(data, _marker, _marker_emoji):
        embed = discord.Embed(title=_('Unlinked Players'), color=discord.Color.blue())
        embed.description = _('These players could be possibly linked:')
        ids = players = members = ''
        for i in range(0, len(data)):
            ids += (chr(0x31 + i) + '\u20E3' + '\n')
            players += "{}\n".format(utils.escape_string(data[i]['name']))
            members += f"{data[i]['match'].display_name}\n"
        embed.add_field(name=_('ID'), value=ids)
        embed.add_field(name=_('DCS Player'), value=players)
        embed.add_field(name=_('Member'), value=members)
        embed.set_footer(text=_('Press a number to link this specific user.'))
        return embed

    @command(description=_('Show players that could be linked'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def linkcheck(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True)
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                # check all unmatched players
                unmatched: list[dict] = []
                async for row in await cursor.execute("""
                    SELECT ucid, name FROM players 
                    WHERE discord_id = -1 AND name IS NOT NULL 
                    ORDER BY last_seen DESC
                """):
                    matched_member = self.bot.match_user(dict(row), True)
                    if matched_member:
                        unmatched.append({"name": row['name'], "ucid": row['ucid'], "match": matched_member})
            if len(unmatched) == 0:
                await interaction.followup.send(_('No unmatched member could be matched.'), ephemeral=True)
                return
        n = await utils.selection_list(interaction, unmatched, self.format_unmatched)
        if n != -1:
            async with self.apool.connection() as conn:
                await conn.execute('UPDATE players SET discord_id = %s, manual = TRUE WHERE ucid = %s',
                                   (unmatched[n]['match'].id, unmatched[n]['ucid']))
                await self.bot.audit(
                    f"linked ucid {unmatched[n]['ucid']} to user {unmatched[n]['match'].display_name}.",
                    user=interaction.user)
                await interaction.followup.send(
                    _("DCS player {player} linked to member {member}.").format(
                        player=utils.escape_string(unmatched[n]['name']),
                        member=unmatched[n]['match'].display_name), ephemeral=True)

    @staticmethod
    def format_suspicious(data, _marker, _marker_emoji):
        embed = discord.Embed(title=_('Possible Mislinks'), color=discord.Color.blue())
        embed.description = _('These players could be possibly mislinked:')
        ids = players = members = ''
        for i in range(0, len(data)):
            ids += (chr(0x31 + i) + '\u20E3' + '\n')
            players += f"{data[i]['name']}\n"
            members += f"{data[i]['mismatch'].display_name}\n"
        embed.add_field(name=_('ID'), value=ids)
        embed.add_field(name=_('DCS Player'), value=players)
        embed.add_field(name=_('Member'), value=members)
        embed.set_footer(text=_('Press a number to unlink this specific user.'))
        return embed

    @command(description=_('Show possibly mislinked players'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def mislinks(self, interaction: discord.Interaction):
        await interaction.response.defer(thinking=True)
        async with self.apool.connection() as conn:
            async with conn.cursor(row_factory=dict_row) as cursor:
                # check all matched members
                suspicious: list[dict] = []
                for member in self.bot.get_all_members():
                    # ignore bots
                    if member.bot:
                        continue
                    await cursor.execute("""
                        SELECT ucid, name FROM players 
                        WHERE discord_id = %s AND name IS NOT NULL AND manual = FALSE 
                        ORDER BY last_seen DESC
                    """, (member.id, ))
                    async for row in cursor:
                        matched_member = self.bot.match_user(dict(row), True)
                        if not matched_member:
                            suspicious.append({"name": row['name'], "ucid": row['ucid'], "mismatch": member})
                        elif matched_member.id != member.id:
                            suspicious.append({"name": row['name'], "ucid": row['ucid'], "mismatch": member,
                                               "match": matched_member})
                if len(suspicious) == 0:
                    await interaction.followup.send(_('No mislinked players found.'), ephemeral=True)
                    return
        n = await utils.selection_list(interaction, suspicious, self.format_suspicious)
        if n != -1:
            ephemeral = utils.get_ephemeral(interaction)
            async with self.apool.connection() as conn:
                await conn.execute('UPDATE players SET discord_id = %s, manual = %s WHERE ucid = %s',
                                   (suspicious[n]['match'].id if 'match' in suspicious[n] else -1,
                                    'match' in suspicious[n], suspicious[n]['ucid']))
                await self.bot.audit(
                    f"unlinked ucid {suspicious[n]['ucid']} from user {suspicious[n]['mismatch'].display_name}.",
                    user=interaction.user)
                if 'match' in suspicious[n]:
                    await self.bot.audit(
                        f"linked ucid {suspicious[n]['ucid']} to user {suspicious[n]['match'].display_name}.",
                        user=interaction.user)
                    await interaction.followup.send(
                        _("UCID {ucid} transferred from member {old_member} to member {new_member}.").format(
                            ucid=suspicious[n]['ucid'],
                            old_member=utils.escape_string(suspicious[n]['mismatch'].display_name),
                            new_member=utils.escape_string(suspicious[n]['match'].display_name)),
                        ephemeral=ephemeral)
                else:
                    await interaction.followup.send(_("Member {name} unlinked from UCID {ucid}.").format(
                        name=utils.escape_string(suspicious[n]['mismatch'].display_name),
                        ucid=suspicious[n]['ucid']), ephemeral=ephemeral)

    @command(description=_('Link your DCS and Discord user'))
    @app_commands.guild_only()
    async def linkme(self, interaction: discord.Interaction):
        async def send_token(token: str):
            await interaction.followup.send(
                _("**Your secure TOKEN is: {token}**\n"
                  "To link your user, type in the following into the in-game chat of one of our DCS servers:"
                  "```{prefix}linkme {token}```\n\n"
                  "**The TOKEN will expire in 2 days!**").format(token=token, prefix=self.eventlistener.prefix),
                ephemeral=True)

        await interaction.response.defer(ephemeral=True)
        member = DataObjectFactory().new(Member, name=interaction.user.name, node=self.node, member=interaction.user)
        if member.ucid and not utils.is_ucid(member.ucid):
            await send_token(member.ucid)
            return
        if utils.is_ucid(member.ucid) and member.verified:
            if not await utils.yn_question(interaction,
                                           _("You already have a verified DCS account!\n"
                                             "Are you sure you want to re-link your account? "
                                             "(Ex: Switched from Steam to Standalone)"), ephemeral=True):
                await interaction.followup.send(_('Aborted.'))
                return
            member.unlink()

        # generate the TOKEN
        async with self.apool.connection() as conn:
            async with conn.cursor() as cursor:
                # in the unlikely event that we had a token already and a linked user
                await cursor.execute("""
                    DELETE FROM players WHERE discord_id = %s AND length(ucid) = 4
                """, (interaction.user.id,))
                # in the very unlikely event that we have generated the very same random number twice
                while True:
                    try:
                        token = str(random.randrange(1000, 9999))
                        await cursor.execute("""
                            INSERT INTO players (ucid, discord_id, last_seen) 
                            VALUES (%s, %s, NOW() AT TIME ZONE 'UTC')
                        """, (token, interaction.user.id))
                        break
                    except psycopg.errors.UniqueViolation:
                        pass
            await send_token(token)

    @player.command(description=_('Shows inactive users'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def inactive(self, interaction: discord.Interaction, period: Literal['days', 'weeks', 'months', 'years'],
                       number: Range[int, 1]):
        report = Report(self.bot, self.plugin_name, 'inactive.json')
        env = await report.render(period=f"{number} {period}")
        await interaction.response.send_message(embed=env.embed, ephemeral=utils.get_ephemeral(interaction))

    @player.command(description=_('Analyses two suspects of being the same person'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def compare(self, interaction: discord.Interaction,
                      player1: app_commands.Transform[discord.Member | str, utils.UserTransformer],
                      player2: app_commands.Transform[discord.Member | str, utils.UserTransformer]):
        await interaction.response.defer()
        ephemeral = utils.get_ephemeral(interaction)
        if isinstance(player1, discord.Member):
            ucid1 = await self.bot.get_ucid_by_member(member=player1, verified=True)
        else:
            ucid1 = player1
        if isinstance(player2, discord.Member):
            ucid2 = await self.bot.get_ucid_by_member(member=player2, verified=True)
        else:
            ucid2 = player2

        if ucid1 == ucid2:
            await interaction.followup.send(_("You have provided the same UCID twice."), ephemeral=ephemeral)
            return

        async with self.apool.connection() as conn:
            cursor = await conn.execute("""
                SELECT 
                    t1.player_ucid AS player1,
                    t2.player_ucid AS player2,
                    GREATEST(t1.hop_on, t2.hop_on) AS overlap_start,
                    LEAST(t1.hop_off, t2.hop_off) AS overlap_end
                FROM 
                    statistics t1
                JOIN 
                    statistics t2
                ON 
                    t1.player_ucid = %s
                    AND t2.player_ucid = %s
                WHERE 
                    t1.hop_off > t2.hop_on
                    AND t1.hop_on < t2.hop_off
                ORDER BY overlap_start
            """, (ucid1, ucid2))
            if cursor.rowcount > 0:
                await interaction.followup.send(_("The players played at the same time. "
                                                  "It is unlikely that they are the same person."), ephemeral=ephemeral)
                return

            # Inform about their non-matching playtimes
            await interaction.followup.send(_("The players never played at the same time."))

            # check both player names
            cursor = await conn.execute("""
                SELECT DISTINCT name FROM (
                    SELECT name FROM players WHERE ucid = %(ucid)s
                    UNION
                    SELECT name FROM players_hist WHERE ucid = %(ucid)s
                ) x
            """, {"ucid": ucid1})
            names1 = [x[0] for x in await cursor.fetchall()]

            cursor = await conn.execute("""
                SELECT DISTINCT name FROM (
                    SELECT name FROM players WHERE ucid = %(ucid)s
                    UNION
                    SELECT name FROM players_hist WHERE ucid = %(ucid)s
                ) x
            """, {"ucid": ucid2})
            names2 = [x[0] for x in await cursor.fetchall()]

            same_names = utils.find_similar_names(names1, names2, threshold=85)
            if same_names:
                embed = discord.Embed(
                    description="The players used similar names in the past:",
                    color=discord.Color.blue()
                )
                names1, names2, scores = zip(*same_names)
                embed.add_field(name="Player 1", value='\n'.join(map(str, names1)))
                embed.add_field(name="Player 2", value='\n'.join(map(str, names2)))
                embed.add_field(name="Confidence", value='\n'.join(f"{score}%" for score in scores))
                await interaction.followup.send(embed=embed, ephemeral=ephemeral)

        period = PeriodFilter()
        for ucid in [ucid1, ucid2]:
            report = Report(self.bot, 'userstats', 'userstats.json')
            env = await report.render(member=ucid, member_name=ucid, server_name=None, period=period.period, flt=period)
            try:
                file = discord.File(fp=env.buffer, filename=env.filename)
                await interaction.followup.send(embed=env.embed, file=file, ephemeral=ephemeral)
            finally:
                if env.buffer:
                    env.buffer.close()

    # New command group "/mission"
    menu = Group(name="menu", description=_("Commands to manage mission menus"))

    @menu.command(description=_('Validate the menu.yaml'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS Admin')
    async def validate(self, interaction: discord.Interaction):
        menu_file = os.path.join(self.node.config_dir, 'menus.yaml')
        if os.path.exists(menu_file):
            await interaction.response.defer(ephemeral=True)
            try:
                utils.validate(menu_file, ['schemas/menus_schema.yaml'], raise_exception=True)
                await interaction.followup.send("Schema valid.", ephemeral=True)
            except Exception as ex:
                self.log.exception(ex)
                message = traceback.format_exc()
                await interaction.followup.send(message[:2000])
        else:
            await interaction.response.send_message(_("No menus.yaml found."), ephemeral=True)

    @command(description=_('Convert values'))
    @app_commands.guild_only()
    @utils.app_has_role('DCS')
    @describe(lat="Latitude in DD or DMS (N491012.12)",
              lon="Longitude in DD or DMS (E011012.34)",
              mgrs="MGRS coordinates")
    async def convert(self, interaction: discord.Interaction, mode: Literal['Lat/Lon => MGRS', 'MGRS => Lat/Lon'],
                      lat: app_commands.Transform[float, DDTransformer] | None = None,
                      lon: app_commands.Transform[float, DDTransformer] | None = None,
                      mgrs: str | None = None):
        if mode == 'Lat/Lon => MGRS':
            if not lat or not lon:
                await interaction.response.send_message("Latitude and longitude must be provided", ephemeral=True)
                return
            mgrs = utils.dd_to_mgrs(lat, lon)
            await interaction.response.send_message(f"**MGRS**: {mgrs}")
        elif mode == 'MGRS => Lat/Lon':
            if not mgrs:
                await interaction.response.send_message("MGRS must be provided", ephemeral=True)
                return
            mgrs = mgrs.replace(' ', '')
            lat, lon = utils.mgrs_to_dd(mgrs)
            d, m, s, f = utils.dd_to_dms(lat)
            lat_dms = ('N' if d > 0 else 'S') + ' {:02d}°{:02d}\'{:02d}.{:02d}"'.format(
                int(abs(d)), int(abs(m)), int(abs(s)), int(abs(f)))
            d, m, s, f = utils.dd_to_dms(lon)
            lon_dms = ('E' if d > 0 else 'W') + ' {:03d}°{:02d}\'{:02d}.{:02d}"'.format(
                int(abs(d)), int(abs(m)), int(abs(s)), int(abs(f)))
            await interaction.response.send_message(f"**DD**: {lat:.7f}, {lon:.7f}\n"
                                                    f"**DMS**: {lat_dms}, {lon_dms}\n"
                                                    f"**DDM**: {utils.dd_to_dmm(lat, lon)}")
        else:
            await interaction.response.send_message("Invalid conversion mode", ephemeral=True)

    @tasks.loop(hours=1)
    async def expire_token(self):
        async with self.apool.connection() as conn:
            await conn.execute("""
                DELETE FROM players 
                WHERE LENGTH(ucid) = 4 AND last_seen < (DATE(now() AT TIME ZONE 'utc') - interval '2 days')
            """)

    @expire_token.before_loop
    async def before_expire(self):
        await self.bot.wait_until_ready()

    @tasks.loop(minutes=1.0)
    async def check_for_unban(self):
        async with self.apool.connection() as conn:
            cursor = await conn.execute("""
                SELECT ucid FROM bans 
                WHERE banned_by <> 'cloud'
                AND banned_until < (NOW() AT TIME ZONE 'utc')
            """)
            rows = await cursor.fetchall()
            for row in rows:
                for server in self.bot.servers.values():
                    if server.status not in [Status.PAUSED, Status.RUNNING, Status.STOPPED]:
                        continue
                    await server.send_to_dcs({
                        "command": "unban",
                        "ucid": row[0]
                    })
                    player = server.get_player(ucid=row[0])
                    if player:
                        player.banned = False
                # delete unbanned accounts from the database
                await conn.execute("DELETE FROM bans WHERE ucid = %s", (row[0], ))

    @check_for_unban.before_loop
    async def before_check_unban(self):
        await self.bot.wait_until_ready()

    @tasks.loop(minutes=5.0)
    async def update_channel_name(self):
        # might happen during a restart
        if not self.bot.member:
            return
        for server_name, server in self.bot.servers.items():
            if server.status == Status.UNREGISTERED:
                continue
            try:
                channel_id = server.channels.get(Channel.STATUS, -1)
                if channel_id == -1:
                    continue
                channel = self.bot.get_channel(channel_id)
                if not channel:
                    channel = await self.bot.fetch_channel(server.channels[Channel.STATUS])
                # name changes of the status channel will only happen with the correct permission
                if not channel.permissions_for(self.bot.member).manage_channels:
                    return
                if channel.type in [discord.ChannelType.forum, discord.ChannelType.public_thread]:
                    continue
# TODO: Alternative implementation, if Discord decides to no longer use system messages for a thread rename
#                    for thread in channel.threads:
#                        if thread.name.startswith(server_name):
#                            channel = thread
#                            break
#                    else:
#                        continue
                name = channel.name
                if server.status in [Status.STOPPED, Status.SHUTDOWN, Status.LOADING, Status.SHUTTING_DOWN]:
                    if name.find('［') == -1:
                        name = name + '［-］'
                    else:
                        name = re.sub('［.*］', f'［-］', name)
                else:
                    players = server.get_active_players()
                    current = len(players) + 1
                    max_players = server.settings.get('maxPlayers') or 0
                    if name.find('［') == -1:
                        name = name + f'［{current}／{max_players}］'
                    else:
                        name = re.sub('［.*］', f'［{current}／{max_players}］', name)
                if name != channel.name:
                    await channel.edit(name=name)
            except discord.Forbidden:
                pass
            except Exception:
                self.log.debug(f"Exception in update_channel_name() for server {server_name}", exc_info=True)

    @update_channel_name.before_loop
    async def before_update_channel_name(self):
        await self.bot.wait_until_ready()

    @tasks.loop(minutes=1.0)
    async def afk_check(self):
        try:
            for server in self.bot.servers.values():
                config = server.locals.get('afk', {})
                max_time = config.get('afk_time', -1)
                if not config or max_time == -1 or server.status != Status.RUNNING:
                    continue
                for ucid, dt in server.afk.copy().items():
                    player = server.get_player(ucid=ucid, active=True)
                    exemptions = config.get('exemptions', {})
                    if 'discord' in exemptions:
                        exemptions['discord'] = list(set(exemptions['discord']) | {"DCS Admin", "GameMaster"})
                        if server.locals.get('managed_by'):
                            exemptions['discord'].extend(server.locals.get('managed_by'))
                    if not player or player.check_exemptions(exemptions):
                        continue
                    if (datetime.now(timezone.utc) - dt).total_seconds() > max_time:
                        msg = server.locals.get('afk', {}).get(
                            'message_afk', '{player.name}, you have been kicked for being AFK for more than {time}.'
                        ).format(player=player, time=utils.format_time(max_time))
                        await server.kick(player, msg)
                        server.afk.pop(ucid, None)
        except Exception as ex:
            self.log.exception(ex)

    @afk_check.before_loop
    async def before_afk_check(self):
        await self.bot.wait_until_ready()

    @tasks.loop(minutes=5.0, count=2)
    async def check_roles(self):
        if not self.bot.is_ready():
            return

        role = self.bot.get_role(self.bot.locals.get('autorole', {}).get('online'))
        if role:
            online_members: set[discord.Member] = set()
            for server in self.bot.servers.values():
                for player in server.get_active_players():
                    if player.member:
                        online_members.add(player.member)
            try:
                # check who needs to lose the role
                for member in (set(role.members) - online_members):
                    await member.remove_roles(role)
            except discord.Forbidden:
                await self.bot.audit('permission "Manage Roles" missing.', user=self.bot.member)
                return
        role = self.bot.get_role(self.bot.locals.get('autorole', {}).get('linked'))
        if role:
            linked_members: set[discord.Member] = set()
            async with self.apool.connection() as conn:
                async for row in await conn.execute("""
                    SELECT DISTINCT discord_id FROM players 
                    WHERE discord_id <> -1 AND manual IS TRUE
                """):
                    member = self.bot.guilds[0].get_member(row[0])
                    if member:
                        linked_members.add(member)
            for member in (linked_members - set(role.members)):
                await member.add_roles(role)
                self.log.debug(f"=> Member {member.display_name} is linked and got the {role.name} role.")

    async def handle_miz_uploads(self, message: discord.Message):
        pattern = ['.miz', '.sav']
        config = self.get_config().get('uploads', {})
        if not MissionUploadHandler.is_valid(message, pattern, config.get('discord', self.bot.roles['DCS Admin'])):
            return
        # check if upload is enabled
        if not config.get('enabled', True):
            self.log.warning("Mission upload is disabled!")
            return

        # check if we are in the correct channel
        server = None
        for node_name, node in self.locals.items():
            if node_name == 'commands':
                continue
            elif node_name == DEFAULT_TAG:
                channel = node.get('uploads', {}).get('channel')
                if channel:
                    if message.channel.id == channel:
                        server = await MissionUploadHandler.get_server(message, channel_id=channel)
                    break
            elif 'uploads' in node:
                channel = node.get('uploads', {}).get('channel')
                if message.channel.id == channel:
                    server = next((
                        server for server in self.bot.servers.values()
                        if server.instance.name == node_name
                    ), None)
                    break
            else:
                for instance_name, instance in node.items():
                    channel = instance.get('uploads', {}).get('channel')
                    if message.channel.id == channel:
                        server = next((
                            server for server in self.bot.servers.values()
                            if server.instance.name == instance_name
                        ), None)
                        break
                else:
                    continue
                break
        else:
            server = await MissionUploadHandler.get_server(message)

        if not server:
            self.log.debug("Mission upload: No server found, you are in the wrong channel!")
            return

        try:
            self.log.debug(f"Uploading mission {message.attachments[0].filename} to server {server.name} ...")
            handler = MissionUploadHandler(plugin=self, server=server, message=message, pattern=pattern)
            base_dir = await handler.server.get_missions_dir()
            ignore = ['.dcssb', 'Saves', 'Scripts']
            if server.locals.get('ignore_dirs'):
                ignore.extend(server.locals['ignore_dirs'])
            await handler.upload(base_dir, ignore_list=ignore)
        except Exception as ex:
            self.log.exception(ex)
        finally:
            with suppress(discord.errors.NotFound):
                await message.delete()

    @staticmethod
    async def load_warehouse_data(buffer: BytesIO) -> dict[str, dict]:
        xlsx = pd.ExcelFile(buffer)
        sheets: dict[str, dict] = {}
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            for key, title in SHEET_TITLES.items():
                df = pd.read_excel(xlsx, sheet_name=title, dtype={"Name": str, "Count": object})
                data = df.set_index("Name")["Count"].to_dict()
                if key == "liquids":
                    data = {
                        REVERSE_LIQUIDS.get(name, name): value
                        for name, value in data.items()
                    }
                sheets[key] = data

        return sheets

    @staticmethod
    async def upload_warehouse_data(channel: discord.TextChannel, server: Server, sheets: dict[str, dict],
                                    airports: list[str]) -> None:
        for airport in airports:
            await channel.send(_("Loading warehouse for airport {} ...").format(airport))
            for key, values in sheets.items():
                await channel.send(_("> uploading {} information ...").format(SHEET_TITLES[key].lower()))
                tasks = []
                for k, v in values.items():
                    if key != 'liquids':
                        cmd = "setWarehouseItem"
                    else:
                        cmd = "setWarehouseLiquid"
                    tasks.append(server.send_to_dcs_sync({
                        "command": cmd,
                        "name": airport,
                        "item": k,
                        "value": v
                    }, timeout=60))
                await asyncio.gather(*tasks)
            await channel.send(_("Warehouse at {} updated.").format(airport))

    async def handle_warehouse_uploads(self, message: discord.Message):
        if not utils.check_roles(set(self.bot.roles['DCS Admin'] + self.bot.roles['GameMaster']), message.author):
            await message.channel.send(_("You need to be DCS Admin or GameMaster to upload data."))
            return

        ctx = await self.bot.get_context(message)
        server = self.bot.get_server(message, admin_only=True)
        if not server:
            server = await utils.server_selection(self.bot, ctx, title=_("To which server do you want to upload?"))

        if not server:
            await message.channel.send(_("Aborted."))
            return

        if server.status not in [Status.PAUSED, Status.RUNNING]:
            await message.channel.send(_("Server {} has to be running or paused.").format(server.display_name))
            return

        att = message.attachments[0]
        filename = att.filename.lower()
        match = re.match(r'^warehouse-([^.]*)\.xlsx?$', filename)
        if not match:
            coalition = await utils.selection(
                ctx,
                title=_("Upload to all warehouses of this coalition:"),
                options=[SelectOption(label="Blue", value="BLUE"), SelectOption(label="Red", value="RED")]
            )
            if not coalition:
                await message.channel.send(_("Aborted."))
                return

            data = await server.send_to_dcs_sync({"command": "getMissionSituation"}, timeout=60)
            airports = data.get('coalitions', {}).get(coalition, {}).get('airbases')
        elif match.group(1).upper() in ['RED', 'BLUE']:
            data = await server.send_to_dcs_sync({"command": "getMissionSituation"}, timeout=60)
            airports = data.get('coalitions', {}).get(match.group(1).upper(), {}).get('airbases')
        elif len(match.group(1)) == 4:
            icao = match.group(1).upper()
            airport = next((x for x in server.current_mission.airbases if x.get('code', '') == icao), None)
            if not airport:
                await message.channel.send(_("Airport with ICAO {} not found.").format(icao))
                return
            airports = [airport['name']]
        else:
            name = utils.slugify(match.group(1)).casefold()
            airport = next((x for x in server.current_mission.airbases
                            if utils.slugify(x.get('name')).casefold() == name), None)
            if not airport:
                await message.channel.send(_("Airport with name {} not found.").format(name))
                return
            airports = [airport['name']]

        if not await utils.yn_question(
                ctx,
                question=_("Do you want to load a new warehouse configuration?"),
                message=_("This will replace the warehouse configuration for:\n{}").format(','.join(airports))
        ):
            await message.channel.send(_("Aborted."))
            return

        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(att.url, proxy=self.node.proxy, proxy_auth=self.node.proxy_auth) as response:
                    response.raise_for_status()
                    sheets = await self.load_warehouse_data(BytesIO(await response.read()))

            await self.upload_warehouse_data(message.channel, server, sheets, airports)
            await message.channel.send(_("All data uploaded."))
        except Exception as ex:
            self.log.exception(ex)
            await message.channel.send(_("Error while processing the file: {}").format(ex))
        finally:
            with suppress(discord.errors.NotFound):
                await message.delete()

    @commands.Cog.listener()
    async def on_message(self, message: discord.Message):
        if message.author.bot or not message.attachments:
            return

        att = message.attachments[0]
        filename = att.filename.lower()
        filetype = filename.lower().split('.')[-1]

        if filetype in ['miz', 'sav']:
            await self.handle_miz_uploads(message)
        elif filename.startswith('warehouse') and filetype.startswith('xls'):
            await self.handle_warehouse_uploads(message)

    @commands.Cog.listener()
    async def on_member_ban(self, _: discord.Guild, member: discord.Member):
        self.bot.log.debug(f"Member {member.display_name} has been banned.")
        if not self.bot.locals.get('no_dcs_autoban', False):
            ucid = await self.bot.get_ucid_by_member(member)
            if ucid:
                await self.bus.ban(ucid, 'Discord',
                                   self.bot.locals.get('message_ban', 'User has been banned on Discord.'))

    @commands.Cog.listener()
    async def on_member_join(self, member: discord.Member):
        ucid = await self.bot.get_ucid_by_member(member, verified=True)
        autorole = self.bot.locals.get('autorole', {}).get('linked')
        if ucid and autorole:
            try:
                role = self.bot.get_role(autorole)
                await member.add_roles(role)
                self.log.debug(f"=> Rejoined member {member.display_name} got their role {role.name} back.")
            except discord.Forbidden:
                await self.bot.audit(_('permission "Manage Roles" missing.'), user=self.bot.member)

    @commands.Cog.listener()
    async def on_member_update(self, before: discord.Member, after: discord.Member):
        # did a member change their roles?
        if before.roles == after.roles:
            return
        for server in self.bot.servers.values():
            player: Player = server.get_player(discord_id=before.id)
            if player and player.verified:
                await server.send_to_dcs({
                    'command': 'uploadUserRoles',
                    'ucid': player.ucid,
                    'discord_id': player.member.id,
                    'roles': [x.id for x in after.roles]
                })

    @commands.Cog.listener()
    async def on_interaction(self, interaction: discord.Interaction):
        if (interaction.type is not discord.InteractionType.component or
                not utils.check_roles(self.bot.roles['DCS Admin'], interaction.user)):
            return

        custom_id = interaction.data.get('custom_id')
        if custom_id.startswith('whitelist_'):
            name = custom_id[len('whitelist_'):]
            async with self.lock:
                if not self.eventlistener.whitelist:
                    self.eventlistener.whitelist = await asyncio.to_thread(self.eventlistener._read_whitelist)
                if name not in self.eventlistener.whitelist:
                    self.eventlistener.whitelist.add(name)
                    whitelist = Path(self.node.config_dir) / 'whitelist.txt'
                    async with aiofiles.open(whitelist, mode="a", encoding='utf-8') as f:
                        await f.write(f"{name}\n")
            for server in self.bus.servers.values():
                if server.status in [Status.RUNNING, Status.PAUSED]:
                    await server.send_to_dcs({
                        "command": "uploadWhitelist",
                        "name": name
                    })
            await interaction.response.edit_message(view=None)
            await interaction.message.add_reaction('✅')

        elif custom_id.startswith('ban_'):
            config = self.get_config()
            if custom_id.startswith('ban_profanity_'):
                ucid = custom_id[len('ban_profanity_'):]
                await self.bus.ban(
                    ucid, interaction.user.display_name,
                    config.get('messages', {}).get('ban_username',
                                                   'Inappropriate username, please contact an admin.')
                )
            elif custom_id.startswith('ban_evade_'):
                ucid = custom_id[len('ban_evade_'):]
                await self.bus.ban(
                    ucid, interaction.user.display_name,
                    config.get('messages', {}).get('ban_evasion',
                                                   'Trying to evade a ban with a 2nd account.')
                )
            await interaction.response.edit_message(view=None)
            await interaction.message.add_reaction('🚫')

        elif custom_id.startswith('message_profanity_'):
            ucid = custom_id[len('message_profanity_'):]
            async with self.apool.connection() as conn:
                await conn.execute("""
                    INSERT INTO messages (sender, player_ucid, message, ack)
                    VALUES (%s, %s, %s, %s)
                """, (interaction.user.display_name, ucid,
                      'Please change your playername to something more appropriate.', True))
            await interaction.response.edit_message(view=None)
            await interaction.message.add_reaction('✉️')

        elif custom_id == 'cancel':
            await interaction.response.edit_message(view=None)


async def setup(bot: DCSServerBot):
    if 'gamemaster' not in bot.plugins:
        raise PluginRequiredError('gamemaster')
    await bot.add_cog(Mission(bot, MissionEventListener))
